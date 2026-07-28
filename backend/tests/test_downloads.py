"""
Tests for the raw-data download center (routes/downloads.py).

Verifies the catalog lists core-ledger partners + module products, the Excel export honours
the business-date range for both a core (Transaction) source and a module source, and that
every export writes an audit-log row.
"""
import io
import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models.database import Base, User, Transaction, SBIBankTransaction, AuditLog
from routes.downloads import download_catalog, download_export, _build_export

USER = User(id="u1", username="raj", role="admin")


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    s = Session()
    yield s
    s.close()


def _txn(db, partner, side, date, amount):
    db.add(Transaction(partner=partner, side=side, recon_date=date, amount=amount, row_type="txn"))
    db.commit()


def test_catalog_lists_core_partner_and_module(db):
    _txn(db, "fino", "bank", "2026-06-20", 100)
    _txn(db, "fino", "internal", "2026-06-20", 100)
    db.add(SBIBankTransaction(txn_date="2026-06-20", debit=50)); db.commit()
    out = download_catalog(db=db, current_user=USER)
    by_key = {p["key"]: p for p in out["products"]}
    assert "fino" in by_key and by_key["fino"]["group"] == "core"
    assert by_key["fino"]["sides"]["bank"]["count"] == 1
    assert "kiosk" in by_key and by_key["kiosk"]["group"] == "module"
    assert by_key["kiosk"]["sides"]["bank"]["count"] == 1


def _wb_rows(content):
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Data"]
    return ws.max_row - 1   # minus header


def test_export_core_honours_date_range(db):
    for d, amt in [("2026-06-20", 100), ("2026-06-21", 200), ("2026-06-25", 300)]:
        _txn(db, "fino", "bank", d, amt)
    content, n = _build_export(db, "fino", "bank", "2026-06-20", "2026-06-21")
    assert n == 2 and _wb_rows(content) == 2


def test_export_module_source(db):
    db.add(SBIBankTransaction(txn_date="2026-06-20", debit=50, ref_number="X1"))
    db.add(SBIBankTransaction(txn_date="2026-06-21", debit=60, ref_number="X2"))
    db.commit()
    content, n = _build_export(db, "kiosk", "bank")   # no dates → everything
    assert n == 2 and _wb_rows(content) == 2


def test_export_writes_audit_row(db):
    _txn(db, "fino", "bank", "2026-06-20", 100)
    resp = download_export(product="fino", side="bank", date_from=None, date_to=None,
                           db=db, current_user=USER)
    assert resp.status_code == 200
    assert db.query(AuditLog).filter(AuditLog.action == "data_download").count() == 1


def test_bad_side_rejected(db):
    from fastapi import HTTPException
    with pytest.raises(HTTPException) as e:
        _build_export(db, "fino", "sideways")
    assert e.value.status_code == 400


def _wb_header(content):
    import io as _io, openpyxl as _ox
    ws = _ox.load_workbook(_io.BytesIO(content))["Data"]
    return [c.value for c in ws[1]]


def test_core_export_expands_raw_data_and_drops_plumbing(db):
    import json as _j
    db.add(Transaction(partner="aeps", side="bank", recon_date="2026-07-20", amount=3500, row_type="txn",
                       recon_status="matched", match_id="APS-1",
                       raw_data=_j.dumps({"RRN": "620114085924", "Bank Name": "SBI",
                                          "Transaction Amount": "3500"})))
    db.commit()
    hdr = _wb_header(_build_export(db, "aeps", "bank")[0])
    assert "RRN" in hdr and "Bank Name" in hdr            # original statement columns present
    assert "Recon Status" in hdr                          # readable recon context appended
    assert not ({"raw_data", "upload_session_id", "id", "matched_with_id"} & set(hdr))  # plumbing dropped


def test_account_filter_core(db):
    for acct, amt in [("ACC-1", 100), ("ACC-1", 200), ("ACC-2", 300)]:
        db.add(Transaction(partner="axis", side="bank", recon_date="2026-07-20", amount=amt,
                           row_type="txn", bank_account=acct, raw_data="{}"))
    db.commit()
    assert _build_export(db, "axis", "bank")[1] == 3
    assert _build_export(db, "axis", "bank", account="ACC-1")[1] == 2


def test_catalog_lists_accounts(db):
    for acct in ("919020056664138", "916020056063631"):
        db.add(Transaction(partner="axis", side="bank", recon_date="2026-07-20", amount=1,
                           row_type="txn", bank_account=acct, raw_data="{}"))
    db.commit()
    ax = next(p for p in download_catalog(db=db, current_user=USER)["products"] if p["key"] == "axis")
    assert set(ax["sides"]["bank"]["accounts"]) == {"919020056664138", "916020056063631"}
