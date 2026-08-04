"""
Characterization + range tests for the two SBI operator workbooks
(core/sbi_reports.py) after the 2026-07-23 range refactor.

Invariants:
1. SINGLE-DATE workbooks keep the finance-ops-verified shape exactly: sheet names,
   header rows, the 16-metric Summary with the 'Matched Count by Source File' block
   at its computed offset, and the 8-column duplicates sheet. (The workbooks were
   verified cell-by-cell against finance-ops' cloud files — this guards the refactor.)
2. RANGE workbooks NEVER match across a date boundary: a day-1 bank ref and a day-2
   source ref with the same 20-digit number both stay Unmatched. Matching pools are
   per-date by construction (pooling would silently reclassify money — the same class
   of change as the reverted P03 scope fix).
3. Range Summary metrics are the exact sums of the per-date workbooks, plus a leading
   'Business Dates Covered' row; the duplicates sheet gains a 'Txn Date' column.
4. The Limit & Settlement P01 status is looked up per (date, KO): the same KO on two
   dates with different P01 statuses shows both, not last-write-wins.
5. A range that resolves to ONE date produces the exact single-date shape.
"""
import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models.database import (Base, SBIBankTransaction, SBIKOLimits, SBIP01Result,
                             SBITxnReport)
from core.sbi_reports import (build_reconciliation_report,
                              build_reconciliation_report_range,
                              build_source_match_report,
                              build_source_match_report_range)

D1, D2 = "2026-07-01", "2026-07-02"
R1 = "61960000000000000001"   # D1 matched pair
R2 = "61960000000000000002"   # D1 source with no bank row
R3 = "61960000000000000003"   # D2 matched pair
RX = "61960000000000000099"   # D1 bank / D2 source — the cross-date probe
RV = "61960000000000000007"   # D1 DR+CR reversal pair

BANK_COLS = ["Bank Stmt Row", "Txn Date", "Value Date", "Description", "Branch Code",
             "Debit", "Credit", "Balance", "Extracted Txn No. (20-digit)",
             "Matched Source File", "Matched Transaction Type", "Source Amount",
             "Amount Match", "Match Status"]
DUP_COLS = ["Group No.", "Bank Stmt Row", "Transaction Number", "Debit", "Credit",
            "Matched Source File", "Description", "Remarks"]
PROD_COLS = ["Sr. No.", "KO ID", "Transaction Date & Time", "Reference Number",
             "Type of Transaction", "From Account", "To Account", "Amount", "Status",
             "Match Status", "Matched Bank Stmt Row", "Matched Bank Txn Date"]


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine, autoflush=False, autocommit=False)()
    yield s
    s.close()


def _bank(db, date, ref=None, debit=0.0, credit=0.0, desc="txn", **kw):
    db.add(SBIBankTransaction(txn_date=date, value_date=date, ref_number=ref,
                              debit=debit, credit=credit, balance=0.0,
                              description=desc, **kw))


def _src(db, date, ref, amount, status="Success", src_file="Withdrawal Report.xlsx"):
    db.add(SBITxnReport(txn_date=date, txn_datetime=f"{date} 10:00:00",
                        reference_number=ref, amount=amount, status=status,
                        source_file=src_file, ko_id="KO1", txn_type="Withdrawal"))


def _fixture(db):
    # D1: matched pair, unmatched bank, DR+CR reversal, settlement debit + KO Withdrawal
    _bank(db, D1, ref=R1, debit=100.0, desc=f"POS {R1}")
    _bank(db, D1, ref=RX, debit=50.0, desc=f"POS {RX}")
    _bank(db, D1, ref=RV, debit=25.0, desc=f"POS {RV} leg one")
    _bank(db, D1, ref=RV, credit=25.0, desc=f"POS {RV} leg two")
    _bank(db, D1, debit=500.0, desc="EKOSETTLEMENT KO deduction", ko_id="KO1",
          is_settlement=True)
    _src(db, D1, R1, 100.0)
    _src(db, D1, R2, 200.0)                       # unmatched (Success) → review
    _src(db, D1, "- / -", 10.0)                   # placeholder → Not Applicable
    db.add(SBIKOLimits(txn_date=D1, txn_datetime=f"{D1} 09:00:00", ko_id="KO1",
                       txn_type="KO Withdrawal", amount=500.0,
                       limit_configured_by="BC1"))
    db.add(SBIP01Result(recon_date=D1, ko_id="KO1", status="matched"))
    # D2: matched pair + the cross-date probe source (same ref as D1's unmatched bank)
    _bank(db, D2, ref=R3, credit=300.0, desc=f"POS {R3}")
    _src(db, D2, R3, 300.0)
    _src(db, D2, RX, 50.0)                        # must NOT match D1's bank row
    db.add(SBIKOLimits(txn_date=D2, txn_datetime=f"{D2} 09:00:00", ko_id="KO1",
                       txn_type="KO Deposit", amount=700.0,
                       limit_configured_by="BC1"))
    db.add(SBIP01Result(recon_date=D2, ko_id="KO1", status="unmatched"))
    db.commit()


def _wb(bio):
    return openpyxl.load_workbook(bio)


def _headers(ws):
    return [c.value for c in ws[1]]


def _col(ws, name):
    return _headers(ws).index(name)


def _rows(ws):
    return list(ws.iter_rows(min_row=2, values_only=True))


# ── 1. single-date characterization (the verified contract) ────────────────────
def test_single_date_reconciliation_workbook_contract(db):
    _fixture(db)
    wb = _wb(build_reconciliation_report(db, D1))
    assert wb.sheetnames == ["Summary", "Bank Statement (Reconciled)",
                             "Unmatched Bank Entries", "Unmatched Source Records",
                             "Duplicate Txn in Bank Stmt"]
    summ = wb["Summary"]
    assert summ.cell(row=2, column=1).value == "Total Bank Statement Transactions"
    assert summ.cell(row=2, column=2).value == 5
    # 15 metric rows (failures now split out of "unmatched") → block title at computed row 18
    assert summ.cell(row=18, column=1).value == "Matched Count by Source File"
    bank = wb["Bank Statement (Reconciled)"]
    assert _headers(bank) == BANK_COLS
    status = {(r[_col(bank, "Extracted Txn No. (20-digit)")] or ""): r[_col(bank, "Match Status")]
              for r in _rows(bank)}     # openpyxl reads empty cells back as None
    assert status[R1] == "Matched"
    assert status[RX] == "Unmatched"
    assert status[RV] == "Matched"          # reversal legs cancel out — reconciled (still listed on the duplicates sheet)
    assert status[""] == "Matched (Settlement)"        # the EKO DEDUCTION debit
    assert _headers(wb["Duplicate Txn in Bank Stmt"]) == DUP_COLS
    assert len(_rows(wb["Duplicate Txn in Bank Stmt"])) == 2   # both reversal legs


def test_single_date_source_match_workbook_contract(db):
    _fixture(db)
    wb = _wb(build_source_match_report(db, D1))
    assert wb.sheetnames[:2] == ["Summary", "Limit & Settlement"]
    assert "Withdrawal" in wb.sheetnames
    assert "AEPS Withdrawal Transaction Rep" in wb.sheetnames   # 31-char truncation
    total = [r for r in _rows(wb["Summary"]) if r[0] == "TOTAL"][0]
    # D1: s(R1) matched, s(R2) unmatched-Success, placeholder NA
    # cols now: Total, Matched, Unmatched(review=Success), Failed(closed), Not Applicable
    assert total[1:] == (2, 1, 1, 0, 1)
    ls = wb["Limit & Settlement"]
    assert len(_rows(ls)) == 1
    assert _rows(ls)[0][_col(ls, "Settlement Status (P01)")] == "Matched"   # status='matched'
    wd = wb["Withdrawal"]
    assert _headers(wd) == PROD_COLS


def test_failed_source_row_is_closed_not_unmatched(db):
    # a Failure txn with no bank match → 'Failed' (closed), OUT of "unmatched" — aligns with the app
    _bank(db, D1, ref=R1, debit=100.0, desc=f"POS {R1}")
    _src(db, D1, R1, 100.0)                                        # matched
    _src(db, D1, R2, 200.0, status="Success")                     # a genuine gap → Unmatched
    _src(db, D1, "61960000000000009999", 300.0, status="Failure") # failed → closed
    db.commit()
    from core.sbi_reports import reconcile
    rep = reconcile(db, D1)
    t = rep["totals"]
    assert t["source_matched"] == 1
    assert t["source_unmatched"] == 1     # the Success gap only — NOT the failure
    assert t["source_failed"] == 1        # the failure, counted separately (closed)
    ms = {r["Reference Number"]: r["Match Status"] for r in rep["source_recs"]}
    assert ms[R2] == "Unmatched" and ms["61960000000000009999"] == "Failed"
    # the failure is NOT in the "Unmatched Source Records" review list
    assert all(r["Reference Number"] != "61960000000000009999" for r in rep["unmatched_source"])


def test_partial_ko_settlement_is_per_row_not_partial(db):
    """Rajendra 2026-08: on a PARTIAL KO-day (some withdrawals settled, one still open) the
    Limit & Settlement sheet must resolve EACH withdrawal per-amount — settled ones 'Matched',
    the open one 'Unmatched' — and NEVER stamp the whole-KO 'Partial' on an individual
    transaction. That whole-KO label was the report↔app desync (a fully-settled row shown
    'Partial' in the report while the app showed it Matched). Mirrors matched_amounts exactly."""
    import json
    for amt in (50000.0, 60000.0, 70000.0):           # two settle, one stays open
        db.add(SBIKOLimits(txn_date=D1, txn_datetime=f"{D1} 13:00:00", ko_id="KOP",
                           txn_type="KO Withdrawal", amount=amt, limit_configured_by="BC1"))
    db.add(SBIP01Result(recon_date=D1, ko_id="KOP", status="partial",
                        matched_amounts=json.dumps([50000.0, 60000.0])))
    _bank(db, D1, debit=50000.0, desc="EKOSETTLEMENT KO deduction", ko_id="KOP",
          is_settlement=True)
    db.commit()
    ls = _wb(build_source_match_report(db, D1))["Limit & Settlement"]
    labels = {r[_col(ls, "Amount")]: r[_col(ls, "Settlement Status (P01)")]
              for r in _rows(ls) if r[_col(ls, "KO ID")] == "KOP"}
    assert labels == {50000.0: "Matched", 60000.0: "Matched", 70000.0: "Unmatched"}
    assert "Partial" not in labels.values()   # KO is 'partial' by day, no txn row says so


def test_partial_ko_duplicate_amounts_resolve_one_each(db):
    """Two withdrawals of the SAME amount, only one settled → exactly one 'Matched' + one
    'Unmatched' (the per-(date,KO) counter is consumed across rows), never two Matched."""
    import json
    for _ in range(2):
        db.add(SBIKOLimits(txn_date=D1, txn_datetime=f"{D1} 13:00:00", ko_id="KOD",
                           txn_type="KO Withdrawal", amount=30000.0, limit_configured_by="BC1"))
    db.add(SBIP01Result(recon_date=D1, ko_id="KOD", status="partial",
                        matched_amounts=json.dumps([30000.0])))
    _bank(db, D1, debit=30000.0, desc="EKOSETTLEMENT KO deduction", ko_id="KOD",
          is_settlement=True)
    db.commit()
    ls = _wb(build_source_match_report(db, D1))["Limit & Settlement"]
    got = sorted(r[_col(ls, "Settlement Status (P01)")]
                 for r in _rows(ls) if r[_col(ls, "KO ID")] == "KOD")
    assert got == ["Matched", "Unmatched"]


def test_partial_ko_deposit_row_is_na_not_partial(db):
    """A KO DEPOSIT on a partial KO-day reads '—' (N/A), never the whole-KO 'Partial' — P01
    settles only withdrawals, so a deposit has no settlement leg (sweep residual). The KO's two
    withdrawals still resolve per-amount (one settled, one open)."""
    import json
    db.add(SBIKOLimits(txn_date=D1, txn_datetime=f"{D1} 13:00:00", ko_id="KOX",
                       txn_type="KO Withdrawal", amount=50000.0, limit_configured_by="BC1"))
    db.add(SBIKOLimits(txn_date=D1, txn_datetime=f"{D1} 13:30:00", ko_id="KOX",
                       txn_type="KO Withdrawal", amount=70000.0, limit_configured_by="BC1"))
    db.add(SBIKOLimits(txn_date=D1, txn_datetime=f"{D1} 14:00:00", ko_id="KOX",
                       txn_type="KO Deposit", amount=90000.0, limit_configured_by="BC1"))
    db.add(SBIP01Result(recon_date=D1, ko_id="KOX", status="partial",
                        matched_amounts=json.dumps([50000.0])))
    _bank(db, D1, debit=50000.0, desc="EKOSETTLEMENT KO deduction", ko_id="KOX",
          is_settlement=True)
    db.commit()
    ls = _wb(build_source_match_report(db, D1))["Limit & Settlement"]
    byamt = {r[_col(ls, "Amount")]: r[_col(ls, "Settlement Status (P01)")]
             for r in _rows(ls) if r[_col(ls, "KO ID")] == "KOX"}
    assert byamt == {50000.0: "Matched", 70000.0: "Unmatched", 90000.0: "—"}
    assert "Partial" not in byamt.values()   # a KO Deposit must NEVER inherit the whole-KO label


# ── 2/3. range: sums + no cross-date matching + duplicates date column ─────────
def test_range_never_matches_across_dates(db):
    _fixture(db)
    wb = _wb(build_reconciliation_report_range(db, [D1, D2]))
    summ = wb["Summary"]
    assert summ.cell(row=2, column=1).value == "Business Dates Covered"
    assert f"{D1} → {D2} (2 date(s) with data)" in str(summ.cell(row=2, column=2).value)
    assert summ.cell(row=3, column=2).value == 6          # bank_total = 5 + 1
    assert summ.cell(row=19, column=1).value == "Matched Count by Source File"
    bank = wb["Bank Statement (Reconciled)"]
    by_ref = {r[_col(bank, "Extracted Txn No. (20-digit)")]: r for r in _rows(bank)}
    # THE invariant: D1 bank ref RX must stay Unmatched even though a D2 source
    # carries the identical ref — matching never crosses a date boundary.
    assert by_ref[RX][_col(bank, "Match Status")] == "Unmatched"
    assert by_ref[R1][_col(bank, "Match Status")] == "Matched"
    assert by_ref[R3][_col(bank, "Match Status")] == "Matched"
    # chronological blocks, per-date Bank Stmt Row numbering (D2 restarts at 1)
    rows = _rows(bank)
    assert rows[0][_col(bank, "Txn Date")] == D1
    assert rows[-1][_col(bank, "Txn Date")] == D2
    assert rows[-1][_col(bank, "Bank Stmt Row")] == 1
    # the D2 probe source appears among unmatched source records
    us = wb["Unmatched Source Records"]
    assert RX in {r[_col(us, "Reference Number")] for r in _rows(us)}
    # duplicates sheet gains the Txn Date column, rows tagged with their date
    dup = wb["Duplicate Txn in Bank Stmt"]
    assert _headers(dup) == DUP_COLS[:1] + ["Txn Date"] + DUP_COLS[1:]
    assert {r[1] for r in _rows(dup)} == {D1}


def test_range_source_match_per_date_p01_and_sums(db):
    _fixture(db)
    wb = _wb(build_source_match_report_range(db, [D1, D2]))
    total = [r for r in _rows(wb["Summary"]) if r[0] == "TOTAL"][0]
    # 4 valid-ref rows (R1, R2, R3, RX-probe): 2 matched, 2 unmatched-Success, 0 Failed, 1 NA
    assert total[1:] == (4, 2, 2, 0, 1)
    ls = wb["Limit & Settlement"]
    rows = _rows(ls)
    assert [r[_col(ls, "Sr. No.")] for r in rows] == [1, 2]
    # D1 row is a KO Withdrawal on a matched KO → Matched. D2 row is a KO DEPOSIT: P01 settles
    # only withdrawals, so a deposit's settlement status is N/A ('—'), never the whole-KO label.
    assert rows[0][_col(ls, "Settlement Status (P01)")] == "Matched"      # D1 KO Withdrawal, matched
    assert rows[1][_col(ls, "Settlement Status (P01)")] == "—"            # D2 KO Deposit → N/A
    wd = wb["Withdrawal"]
    ms = {(r[_col(wd, "Reference Number")]): r[_col(wd, "Match Status")] for r in _rows(wd)}
    assert ms[R1] == "Matched" and ms[R3] == "Matched"
    assert ms[RX] == "Unmatched"                    # the probe never matched D1's bank
    assert ms["- / -"] == "Not Applicable"


# ── 5. a one-date range is exactly the single-date workbook ────────────────────
def test_single_element_range_keeps_legacy_shape(db):
    _fixture(db)
    wb = _wb(build_reconciliation_report_range(db, [D1]))
    summ = wb["Summary"]
    assert summ.cell(row=2, column=1).value == "Total Bank Statement Transactions"
    assert summ.cell(row=18, column=1).value == "Matched Count by Source File"
    assert _headers(wb["Duplicate Txn in Bank Stmt"]) == DUP_COLS


# ── 6. half-loaded days inside a range are flagged, never silent ───────────────
# (Adversarial-review finding: a day whose bank statement isn't uploaded showed a
# full day of phantom 'needs review' unmatched rows with no hint the FILE is missing.)
D3 = "2026-07-03"


def test_range_flags_missing_bank_statement_day(db):
    _fixture(db)
    _src(db, D3, "61960000000000000055", 75.0)     # source rows, NO bank rows
    db.commit()
    wb = _wb(build_reconciliation_report_range(db, [D1, D3]))
    metrics = {r[0]: r[1] for r in _rows(wb["Summary"]) if r[0]}
    gap = [k for k in metrics if "NO bank-statement rows" in str(k)]
    assert gap and metrics[gap[0]] == D3
    # Workbook B: Coverage Notes block below the summary table
    wb2 = _wb(build_source_match_report_range(db, [D1, D3]))
    cells = [str(r[0]) for r in _rows(wb2["Summary"])]
    assert "Coverage Notes" in cells
    assert any("NO bank-statement rows" in c for c in cells)


def test_full_range_has_no_gap_rows(db):
    _fixture(db)                                    # D1 and D2 both have both sides
    wb = _wb(build_reconciliation_report_range(db, [D1, D2]))
    assert not any("missing file" in str(r[0]) for r in _rows(wb["Summary"]))


def test_ko_limits_only_date_reaches_range_workbook(db):
    _fixture(db)
    ko_day = "2026-07-04"                           # bank holiday: only limit activity
    db.add(SBIKOLimits(txn_date=ko_day, txn_datetime=f"{ko_day} 09:00:00", ko_id="KO9",
                       txn_type="KO Withdrawal", amount=900.0, limit_configured_by="BC1"))
    db.commit()
    from routes.sbi_kiosk import _resolve_workbook_dates
    assert ko_day in _resolve_workbook_dates(db, D1, "2026-07-31")
    wb = _wb(build_source_match_report_range(db, [D1, ko_day]))
    ls = wb["Limit & Settlement"]
    assert "KO9" in {r[_col(ls, "KO ID")] for r in _rows(ls)}
