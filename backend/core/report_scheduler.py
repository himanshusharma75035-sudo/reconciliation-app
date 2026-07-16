"""
core/report_scheduler.py

Scheduler logic for personal report subscriptions.
- register_report_job()    — add/update a cron job
- unregister_report_job()  — remove a cron job
- execute_report_subscription() — run a single subscription now
- reload_all_report_subscriptions() — called on startup

Date window logic converts labels like "yesterday", "last_7_days" into
concrete from/to date strings at job runtime, so reports always cover
the right period regardless of when the job fires.
"""
import datetime
import io
import json
import logging

from apscheduler.triggers.cron import CronTrigger

logger = logging.getLogger("eko_recon.report_scheduler")

REPORT_LABELS = {
    "open_items":    "Open Items",
    "summary":       "Recon Summary",
    "eod":           "EOD Summary",
    "matched_pairs": "Matched Pairs",
    "ageing":        "Ageing Report",
    "src":           "SRC Report",
    "evalue_summary":     "E-Value Summary",
    "evalue_exceptions":  "E-Value Exceptions",
    "evalue_unmatched":   "E-Value Unmatched",
    "bbps_summary":       "BBPS Summary",
    "bbps_exceptions":    "BBPS Exceptions",
    "bbps_unmatched":     "BBPS Unmatched",
    "sbi_daily_pack":     "SBI Kiosk Daily Pack",
    "sbi_summary":        "SBI Kiosk MIS Summary",
    "sbi_exceptions":     "SBI Kiosk Exceptions",
    "funds_position":     "Funds Position",
}

DATE_RANGE_LABELS = {
    "today":        "Today",
    "yesterday":    "Yesterday",
    "last_7_days":  "Last 7 Days",
    "this_week":    "This Week",
    "this_month":   "This Month",
    "last_month":   "Last Month",
}


# ─── Date window resolver ──────────────────────────────────────────────────────
def _resolve_dates(date_range: str) -> tuple[str, str]:
    """Return (from_date, to_date) as YYYY-MM-DD strings for the given window."""
    today = datetime.date.today()

    if date_range == "today":
        return str(today), str(today)
    if date_range == "yesterday":
        d = today - datetime.timedelta(days=1)
        return str(d), str(d)
    if date_range == "last_7_days":
        return str(today - datetime.timedelta(days=6)), str(today)
    if date_range == "this_week":
        # Monday → today
        monday = today - datetime.timedelta(days=today.weekday())
        return str(monday), str(today)
    if date_range == "this_month":
        first = today.replace(day=1)
        return str(first), str(today)
    if date_range == "last_month":
        first_this = today.replace(day=1)
        last_prev  = first_this - datetime.timedelta(days=1)
        first_prev = last_prev.replace(day=1)
        return str(first_prev), str(last_prev)
    # Fallback: yesterday
    d = today - datetime.timedelta(days=1)
    return str(d), str(d)


# ─── Report generator ─────────────────────────────────────────────────────────
def _generate_evalue_report(report_type, filters, from_date, to_date, db):
    """E-Value scheduled report (summary / exceptions / unmatched) over E-Value tables."""
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill, Font as XFont
    from models.database import EvalueBankTxn, EvalueWalletLoad
    from sqlalchemy import func

    reco = filters.get("reco_acc_no") or None
    # Wallet loads window on the VALUE date (fallback txn) — the date the load reconciles
    # on, matching every interactive view (behavior-contract item 13).
    _load_date = func.coalesce(func.nullif(EvalueWalletLoad.value_date, ""),
                               EvalueWalletLoad.transaction_date)
    bq = db.query(EvalueBankTxn).filter(EvalueBankTxn.txn_date >= from_date,
                                        EvalueBankTxn.txn_date <= to_date)
    lq = db.query(EvalueWalletLoad).filter(_load_date >= from_date,
                                           _load_date <= to_date)
    if reco:
        bq = bq.filter(EvalueBankTxn.reco_acc_no == reco)
        lq = lq.filter(EvalueWalletLoad.reco_acc_no == reco)
    bank = bq.all(); loads = lq.all()

    def brow(b):
        return {"Account": b.reco_acc_no, "Bank": b.bank_name, "Txn Date": b.txn_date,
                "Description": b.description, "Dr/Cr": b.dr_cr, "Amount": b.amount,
                "Channel": b.channel, "UTR": b.utr, "Recon Status": b.recon_status, "Match ID": b.match_id}

    def style(ws, fill="1E3A5F"):
        for col in ws.columns:
            m = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(m + 2, 46)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=fill); cell.font = XFont(bold=True, color="FFFFFF")

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        if report_type == "evalue_summary":
            STAT = ["matched_online", "matched_cash", "matched_manual", "interbank_matched",
                    "twice_credit", "wrong_amount", "unmatched_bank", "transaction_fee", "bank_debit"]
            accts = sorted(set(b.reco_acc_no for b in bank) | set(l.reco_acc_no for l in loads))
            rows = []
            for a in accts:
                bk = [b for b in bank if b.reco_acc_no == a]
                ld = [l for l in loads if l.reco_acc_no == a]
                c = {s: sum(1 for b in bk if b.recon_status == s) for s in STAT}
                credits = sum(1 for b in bk if b.dr_cr == "CR")
                matched = c["matched_online"] + c["matched_cash"] + c["matched_manual"] + c["interbank_matched"]
                rows.append({"Account": a, "Bank Credits": credits, "Loads": len(ld),
                             "Unmatched Load": sum(1 for l in ld if l.recon_status == "unmatched_load"),
                             **{k.replace("_", " ").title(): v for k, v in c.items()},
                             "Match Rate %": round(matched / (credits or 1) * 100, 1)})
            (pd.DataFrame(rows) if rows else pd.DataFrame({"info": ["no data"]})).to_excel(w, index=False, sheet_name="Summary")
            style(w.sheets["Summary"])
        elif report_type == "evalue_exceptions":
            exc = [brow(b) for b in bank if b.recon_status in ("twice_credit", "wrong_amount", "unmatched_bank")]
            unl = [{"Account": l.reco_acc_no, "Value Date": l.value_date, "Txn Date": l.transaction_date,
                    "CSP": l.csp_code,
                    "Amount": l.amount, "Mode": l.load_mode, "UTR": l.utr_number}
                   for l in loads if l.recon_status == "unmatched_load"]
            (pd.DataFrame(exc) if exc else pd.DataFrame({"info": ["none"]})).to_excel(w, index=False, sheet_name="Bank Exceptions")
            style(w.sheets["Bank Exceptions"], "B45309")
            (pd.DataFrame(unl) if unl else pd.DataFrame({"info": ["none"]})).to_excel(w, index=False, sheet_name="Unmatched Loads")
            style(w.sheets["Unmatched Loads"], "B45309")
        else:  # evalue_unmatched
            ub = [brow(b) for b in bank if b.recon_status == "unmatched_bank"]
            (pd.DataFrame(ub) if ub else pd.DataFrame({"info": ["none"]})).to_excel(w, index=False, sheet_name="Unmatched Bank")
            style(w.sheets["Unmatched Bank"], "B45309")
    out.seek(0)
    fn = f"{report_type}_{reco or 'all'}_{from_date}_to_{to_date}.xlsx".replace(" ", "_")
    return out.getvalue(), fn


def _generate_bbps_report(report_type, filters, from_date, to_date, db):
    """BBPS scheduled report (summary / exceptions / unmatched) over BBPS tables."""
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill, Font as XFont
    from models.database import BbpsBankTxn, BbpsInternal

    provider = filters.get("provider") or None
    bq = db.query(BbpsBankTxn).filter(BbpsBankTxn.transaction_date >= from_date,
                                      BbpsBankTxn.transaction_date <= to_date)
    iq = db.query(BbpsInternal).filter(BbpsInternal.transaction_date >= from_date,
                                       BbpsInternal.transaction_date <= to_date)
    if provider:
        bq = bq.filter(BbpsBankTxn.provider == provider)
        iq = iq.filter(BbpsInternal.provider == provider)
    bank = bq.all(); internal = iq.all()

    def brow(b):
        return {"Provider": b.provider, "Client Ref": b.client_ref, "Operator Ref": b.operator_ref,
                "Amount": b.amount, "Status": b.status, "Date": b.transaction_date,
                "Service": b.service_name, "Operator": b.operator_name, "Reason": b.reason,
                "Recon Status": b.recon_status, "Match ID": b.match_id}

    def irow(i):
        return {"Eko TID": i.eko_trxn_id, "Provider": i.provider, "Amount": i.amount,
                "Status": i.status, "Refunded": i.is_refunded, "Date": i.transaction_date,
                "CSP": i.csp_code, "Merchant": i.merchant_name, "Recon Status": i.recon_status, "Match ID": i.match_id}

    def style(ws, fill="1E3A5F"):
        for col in ws.columns:
            m = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(m + 2, 46)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=fill); cell.font = XFont(bold=True, color="FFFFFF")

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        if report_type == "bbps_summary":
            provs = sorted({b.provider for b in bank})
            rows = []
            for p in provs:
                rs = [b for b in bank if b.provider == p]
                recon = sum(1 for b in rs if b.recon_status in ("matched", "failed_refunded"))
                rows.append({"Provider": p, "Operator Rows": len(rs), "Reconciled": recon,
                             "Match Rate %": round(recon / (len(rs) or 1) * 100, 1)})
            (pd.DataFrame(rows) if rows else pd.DataFrame({"info": ["no data"]})).to_excel(w, index=False, sheet_name="Summary")
            style(w.sheets["Summary"])
        elif report_type == "bbps_exceptions":
            exc = [brow(b) for b in bank if b.recon_status in ("failed_pending_refund", "refunded_but_success", "amount_mismatch")]
            (pd.DataFrame(exc) if exc else pd.DataFrame({"info": ["none"]})).to_excel(w, index=False, sheet_name="Exceptions")
            style(w.sheets["Exceptions"], "B45309")
        else:  # bbps_unmatched
            ub = [brow(b) for b in bank if b.recon_status == "unmatched_bank"]
            ui = [irow(i) for i in internal if i.recon_status == "unmatched_internal"]
            (pd.DataFrame(ub) if ub else pd.DataFrame({"info": ["none"]})).to_excel(w, index=False, sheet_name="Unmatched Operator")
            style(w.sheets["Unmatched Operator"], "B45309")
            (pd.DataFrame(ui) if ui else pd.DataFrame({"info": ["none"]})).to_excel(w, index=False, sheet_name="Unmatched Internal")
            style(w.sheets["Unmatched Internal"], "B45309")
    out.seek(0)
    fn = f"{report_type}_{provider or 'all'}_{from_date}_to_{to_date}.xlsx".replace(" ", "_")
    return out.getvalue(), fn


def _generate_report(report_type: str, filters: dict, date_range: str, db) -> tuple[bytes, str]:
    """
    Generate the report Excel and return (bytes, filename).
    Imports pandas + openpyxl inline to avoid circular imports.
    """
    import pandas as pd
    from models.database import Transaction, ReconStatus
    from openpyxl.styles import PatternFill, Font as XFont

    from_date, to_date = _resolve_dates(date_range)

    # ── E-Value scheduled reports (own tables) ────────────────────────────────
    if report_type.startswith("evalue_"):
        return _generate_evalue_report(report_type, filters, from_date, to_date, db)

    # ── BBPS scheduled reports (own tables) ───────────────────────────────────
    if report_type.startswith("bbps_"):
        return _generate_bbps_report(report_type, filters, from_date, to_date, db)

    # ── SBI Kiosk scheduled reports (own process tables) ──────────────────────
    # Shares the report-library builder with the on-demand /sbi/report route. The SBI
    # type is the suffix (sbi_daily_pack -> daily_pack); date-scoped reports use to_date,
    # range-scoped use from_date..to_date (the builder picks by the report's scope).
    if report_type.startswith("sbi_"):
        from routes.sbi_kiosk import build_sbi_report_xlsx
        data, fname, _tag = build_sbi_report_xlsx(
            db, report_type[len("sbi_"):], recon_date=to_date,
            date_from=from_date, date_to=to_date)
        return data, fname

    # ── Funds Position (EOD) — director/CFO daily balance email ───────────────
    if report_type == "funds_position":
        import io as _io
        import datetime as _dt
        from core.funds import get_funds_position
        # T+1: the scheduled email reports the previous day's statements (today's
        # statements are not uploaded yet when the morning schedule fires).
        d = str(_dt.date.today() - _dt.timedelta(days=1))
        pos = get_funds_position(db, d)
        cols = ["product", "partner", "bank_account", "statement_date",
                "opening_balance", "total_dr", "total_cr", "closing_balance",
                "delta", "prev_date", "txn_count", "source"]
        totals = [{"product": p, **t} for p, t in sorted(pos["totals_by_product"].items())]
        out = _io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            pd.DataFrame(pos["rows"], columns=cols).to_excel(w, sheet_name="Funds Position", index=False)
            pd.DataFrame(totals, columns=["product", "accounts", "closing_total"]
                         ).to_excel(w, sheet_name="Totals by Product", index=False)
            for ws in w.sheets.values():
                fill = PatternFill("solid", fgColor="094053")
                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = XFont(color="FFFFFF", bold=True)
        out.seek(0)
        return out.read(), f"funds_position_{d}.xlsx"

    # ── EOD summary — the rich 3-sheet workbook, shared with the on-demand export
    # (previously the scheduled 'eod' emitted only a plain per-partner grid). A daily EOD
    # is single-date: use the end of the resolved range; partner from filters (falsy => all).
    if report_type == "eod":
        from routes.reports import build_eod_summary_xlsx
        return build_eod_summary_xlsx(db, filters.get("partner"), to_date)

    partner   = filters.get("partner") or None
    side      = filters.get("side")    or None
    src_code  = filters.get("src_code")or None
    status    = filters.get("recon_status") or None

    q = db.query(Transaction)

    # Date range
    q = q.filter(Transaction.recon_date >= from_date,
                 Transaction.recon_date <= to_date)

    # Common filters — expand a PRODUCT-level partner (e.g. 'dmt') to its member
    # banks so a scheduled DMT report isn't an empty workbook (parity with the
    # on-demand Reports page's _apply_partner). A 1:1 product returns [partner].
    if partner:
        from routes.reports import _partner_slugs
        _slugs = _partner_slugs(partner)
        if _slugs:
            q = q.filter(Transaction.partner.in_(_slugs))
    if side:     q = q.filter(Transaction.side    == side)
    if src_code: q = q.filter(Transaction.src_code == src_code)

    if report_type == "open_items":
        if status and status not in ("", "all"):
            q = q.filter(Transaction.recon_status == status)
        else:
            q = q.filter(Transaction.recon_status.in_(
                [ReconStatus.unmatched, ReconStatus.src_assigned]))
        q = q.filter(Transaction.row_type == "txn")

    elif report_type == "matched_pairs":
        q = q.filter(Transaction.recon_status.in_(
            [ReconStatus.matched, ReconStatus.manual_matched]))
        q = q.filter(Transaction.row_type == "txn")

    elif report_type == "ageing":
        q = q.filter(Transaction.recon_status.in_(
            [ReconStatus.unmatched, ReconStatus.src_assigned]))
        q = q.filter(Transaction.row_type == "txn")

    elif report_type == "src":
        q = q.filter(Transaction.recon_status == ReconStatus.src_assigned)
        q = q.filter(Transaction.row_type == "txn")

    elif report_type == "summary":
        pass  # all rows, aggregate below

    elif report_type == "eod":
        pass  # full summary

    txns = q.order_by(Transaction.recon_date, Transaction.partner, Transaction.side).all()

    # Build DataFrame
    if report_type in ("summary", "eod"):
        from collections import defaultdict
        agg = defaultdict(lambda: {"total": 0, "matched": 0, "unmatched": 0, "amount": 0.0})
        for t in txns:
            key = (t.partner, t.recon_date)
            agg[key]["total"] += 1
            agg[key]["amount"] += float(t.amount or 0)
            s = str(t.recon_status)
            if s in ("matched", "manual_matched", "interbank_matched"):
                agg[key]["matched"] += 1
            elif s in ("unmatched", "src_assigned", "amount_mismatch"):
                agg[key]["unmatched"] += 1
        rows = []
        for (partner_k, date_k), v in sorted(agg.items()):
            rate = round(v["matched"] / max(v["total"], 1) * 100, 1)
            rows.append({"Partner": partner_k, "Date": date_k,
                         "Total": v["total"], "Matched": v["matched"],
                         "Unmatched": v["unmatched"], "Match Rate %": rate,
                         "Amount (₹)": round(v["amount"], 2)})
        sheet_name = "Summary"
    else:
        rows = [{
            "Partner":         t.partner,
            "Side":            t.side,
            "Recon Date":      t.recon_date,
            "Transaction Date":t.transaction_date or "",
            "Recon Status":    t.recon_status,
            "Eko TID":         t.eko_tid         or "",
            "Tracking Number": t.tracking_number or "",
            "UTR Number":      t.utr_number      or "",
            "Amount (₹)":      t.amount,
            "DR/CR":           t.dr_cr            or "",
            "Txn Status":      t.status           or "",
            "Match ID":        t.match_id         or "",
            "SRC Code":        t.src_code         or "",
            "SRC Note":        t.src_note         or "",
        } for t in txns]
        sheet_name = REPORT_LABELS.get(report_type, report_type.title())

    df = pd.DataFrame(rows) if rows else pd.DataFrame()

    # Style + write
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        ws = writer.sheets[sheet_name[:31]]
        # Autofit
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        # Header styling
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            cell.font = XFont(bold=True, color="FFFFFF")

    filename = (
        f"{report_type}_{partner or 'all'}_{from_date}_to_{to_date}.xlsx"
        .replace(" ", "_")
    )
    return output.getvalue(), filename


# ─── Email delivery ───────────────────────────────────────────────────────────
def _send_report(
    to_list: list[str],
    subject: str,
    body_html: str,
    attachment_bytes: bytes,
    filename: str,
):
    """Send an email with the Excel file attached."""
    import os, smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    SMTP_HOST = os.getenv("SMTP_HOST", "")
    SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
    SMTP_USER = os.getenv("SMTP_USER", "")
    SMTP_PASS = os.getenv("SMTP_PASS", "")
    SMTP_FROM = os.getenv("SMTP_FROM", "recon@eko.co.in")

    if not SMTP_HOST or not to_list:
        logger.info(f"[report_scheduler] Email skipped (SMTP not configured): {subject}")
        return "skipped_no_smtp"

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"]    = SMTP_FROM
    msg["To"]      = ", ".join(to_list)
    msg.attach(MIMEText(body_html, "html"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
        server.ehlo(); server.starttls()
        if SMTP_USER:
            server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_FROM, to_list, msg.as_string())
    return "sent"


# ─── Executive dashboard email block ──────────────────────────────────────────
def _fmt_inr_short(n) -> str:
    n = float(n or 0)
    if abs(n) >= 1e7: return f"₹{n / 1e7:.2f} Cr"
    if abs(n) >= 1e5: return f"₹{n / 1e5:.2f} L"
    return f"₹{n:,.0f}"


def _analytics_email_html(db, from_date: str, to_date: str) -> str:
    """Executive analytics dashboard as email-safe HTML (inline styles, table layout,
    pure-CSS bars — no JS/SVG, renders in Gmail/Outlook). Covers the same date window
    as the host report. Returns '' on ANY error so it can never block the email."""
    try:
        import os
        from core.analytics import build_analytics
        a = build_analytics(db, date_from=from_date, date_to=to_date)
        t = a.get("totals", {})
        groups = a.get("by_group", [])
        day_label = from_date if from_date == to_date else f"{from_date} → {to_date}"
        base = os.getenv("APP_PUBLIC_URL", "https://122.176.147.78:8443/recon").rstrip("/")
        exec_url = base + "/exec"

        kpis = [
            ("Total txns",     f'{t.get("transactions", 0):,}',        "#111827"),
            ("Matched",        f'{t.get("matched", 0):,}',             "#059669"),
            ("Unmatched",      f'{t.get("unmatched", 0):,}',           "#dc2626"),
            ("Match rate",     f'{t.get("match_rate", 0)}%',           "#4f46e5"),
            ("Matched volume", _fmt_inr_short(t.get("matched_volume", 0)), "#111827"),
        ]
        kpi_cells = "".join(
            f'<td style="padding:8px 6px;text-align:center;border:1px solid #eef2f7">'
            f'<div style="font-size:10px;color:#6b7280;text-transform:uppercase;letter-spacing:.03em">{lbl}</div>'
            f'<div style="font-size:17px;font-weight:800;color:{color};margin-top:2px">{val}</div></td>'
            for lbl, val, color in kpis)

        def _bar(rate):
            rate = max(0.0, min(100.0, float(rate or 0)))
            col = "#059669" if rate >= 85 else ("#d97706" if rate >= 50 else "#dc2626")
            return (f'<span style="background:#eef2f7;border-radius:4px;height:8px;width:80px;'
                    f'display:inline-block;vertical-align:middle">'
                    f'<span style="background:{col};height:8px;width:{rate:.0f}%;'
                    f'border-radius:4px;display:inline-block"></span></span>')

        rows = "".join(
            '<tr>'
            f'<td style="padding:6px 8px;border-bottom:1px solid #f1f5f9;font-weight:600;color:#334155">{g.get("label","")}</td>'
            f'<td style="padding:6px 8px;border-bottom:1px solid #f1f5f9;text-align:right;color:#475569">{g.get("transactions",0):,}</td>'
            f'<td style="padding:6px 8px;border-bottom:1px solid #f1f5f9;text-align:right;color:#059669;font-weight:600">{g.get("matched",0):,}</td>'
            f'<td style="padding:6px 8px;border-bottom:1px solid #f1f5f9;text-align:right;color:#dc2626">{g.get("unmatched",0):,}</td>'
            f'<td style="padding:6px 8px;border-bottom:1px solid #f1f5f9;text-align:right;white-space:nowrap">{_bar(g.get("match_rate",0))} '
            f'<span style="font-size:11px;color:#64748b">{g.get("match_rate",0)}%</span></td>'
            '</tr>'
            for g in groups)
        if not rows:
            rows = ('<tr><td colspan="5" style="padding:10px;text-align:center;color:#9ca3af;'
                    'font-style:italic">No reconciliation activity for this window.</td></tr>')

        return f"""
        <div style="margin-top:22px">
          <div style="background:#065f46;color:white;padding:13px 18px;border-radius:8px 8px 0 0">
            <div style="font-size:11px;opacity:.75;letter-spacing:.04em">EXECUTIVE RECONCILIATION DASHBOARD</div>
            <div style="font-size:15px;font-weight:700;margin-top:3px">Activity for {day_label} · all products</div>
          </div>
          <div style="border:1px solid #e5e7eb;border-top:none;padding:16px 18px;border-radius:0 0 8px 8px;background:#fff">
            <table width="100%" style="border-collapse:collapse;margin-bottom:14px"><tr>{kpi_cells}</tr></table>
            <table width="100%" style="border-collapse:collapse;font-size:12.5px">
              <tr style="color:#94a3b8;text-align:left">
                <th style="padding:6px 8px;font-weight:600;border-bottom:2px solid #e5e7eb">Product</th>
                <th style="padding:6px 8px;font-weight:600;text-align:right;border-bottom:2px solid #e5e7eb">Txns</th>
                <th style="padding:6px 8px;font-weight:600;text-align:right;border-bottom:2px solid #e5e7eb">Matched</th>
                <th style="padding:6px 8px;font-weight:600;text-align:right;border-bottom:2px solid #e5e7eb">Unmatched</th>
                <th style="padding:6px 8px;font-weight:600;text-align:right;border-bottom:2px solid #e5e7eb">Match rate</th>
              </tr>
              {rows}
            </table>
            <div style="margin-top:16px;text-align:center">
              <a href="{exec_url}" style="display:inline-block;background:#059669;color:#fff;text-decoration:none;font-size:13px;font-weight:600;padding:9px 20px;border-radius:8px">View the live dashboard →</a>
            </div>
          </div>
        </div>"""
    except Exception as e:
        logger.warning(f"[report_scheduler] analytics dashboard block skipped: {e}")
        return ""


# ─── Main job function ────────────────────────────────────────────────────────
def execute_report_subscription(sub_id: str) -> dict:
    """
    Run a single report subscription: generate Excel + send email.
    Called by the scheduler cron job or by the run-now endpoint.
    """
    from models.database import SessionLocal, ReportSubscription
    db = SessionLocal()
    try:
        sub = db.query(ReportSubscription).filter(ReportSubscription.id == sub_id).first()
        if not sub:
            return {"status": "error", "message": "Subscription not found"}

        filters   = json.loads(sub.filters or "{}")
        from_date, to_date = _resolve_dates(sub.date_range)

        excel_bytes, filename = _generate_report(sub.report_type, filters, sub.date_range, db)

        # Build email
        report_label = REPORT_LABELS.get(sub.report_type, sub.report_type)
        date_label   = DATE_RANGE_LABELS.get(sub.date_range, sub.date_range)
        subject = f"[Eko Recon] {sub.name} — {report_label} {from_date} to {to_date}"
        partner_str = filters.get("partner", "All Partners").title()
        # Opt-in executive dashboard block (same date window as this report).
        dashboard_html = (_analytics_email_html(db, from_date, to_date)
                          if getattr(sub, "include_dashboard", False) else "")

        html = f"""
        <html><body style="font-family:sans-serif;color:#1f2937;max-width:600px;margin:auto;padding:24px">
          <div style="background:#1e3a5f;color:white;padding:16px 20px;border-radius:8px 8px 0 0">
            <div style="font-size:11px;opacity:0.7">Eko Bharat Ventures — Scheduled Report</div>
            <div style="font-size:18px;font-weight:700;margin-top:4px">{sub.name}</div>
          </div>
          <div style="background:#f9fafb;padding:16px 20px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 8px 8px">
            <table style="font-size:13px;width:100%">
              <tr><td style="color:#6b7280;padding:3px 0;width:130px">Report Type:</td><td><strong>{report_label}</strong></td></tr>
              <tr><td style="color:#6b7280;padding:3px 0">Period:</td><td><strong>{date_label}</strong> ({from_date} → {to_date})</td></tr>
              <tr><td style="color:#6b7280;padding:3px 0">Partner:</td><td><strong>{partner_str}</strong></td></tr>
            </table>
            <div style="margin-top:12px;font-size:12px;color:#6b7280">
              The full report is attached as an Excel file.<br>
              To manage your report subscriptions, open the Recon App → Reports → My Scheduled Reports.
            </div>
          </div>
          {dashboard_html}
          <div style="margin-top:12px;font-size:11px;color:#9ca3af;text-align:center">
            Auto-generated by Eko Recon · {datetime.datetime.now().strftime('%d %b %Y %H:%M IST')}
          </div>
        </body></html>"""

        to_list = [e.strip() for e in (sub.email_to or "").split(",") if e.strip()]
        if not to_list:
            sub.last_run_status  = "error"
            sub.last_run_message = "No email address configured"
            sub.last_run_at      = datetime.datetime.utcnow()
            db.commit()
            return {"status": "error", "message": "No email configured"}

        result = _send_report(to_list, subject, html, excel_bytes, filename)

        sub.last_run_at      = datetime.datetime.utcnow()
        sub.last_run_status  = "success" if result == "sent" else result
        sub.last_run_message = f"Sent to {', '.join(to_list)}" if result == "sent" else "SMTP not configured"
        db.commit()
        return {"status": sub.last_run_status, "filename": filename, "rows": "ok"}

    except Exception as e:
        logger.error(f"[report_scheduler] Subscription {sub_id} failed: {e}")
        try:
            sub = db.query(ReportSubscription).filter(ReportSubscription.id == sub_id).first()
            if sub:
                sub.last_run_at      = datetime.datetime.utcnow()
                sub.last_run_status  = "error"
                sub.last_run_message = str(e)[:500]
                db.commit()
        except Exception:
            pass
        return {"status": "error", "message": str(e)}
    finally:
        db.close()


# ─── APScheduler wiring ───────────────────────────────────────────────────────
def _job_id(sub_id: str) -> str:
    return f"report_sub_{sub_id}"


def register_report_job(sub):
    """Add or replace a cron job for the subscription."""
    from core.scheduler import get_scheduler
    scheduler = get_scheduler()
    jid = _job_id(sub.id)

    if scheduler.get_job(jid):
        scheduler.remove_job(jid)

    if not sub.is_active:
        return

    if sub.frequency == "daily":
        trigger = CronTrigger(hour=sub.hour, minute=sub.minute, timezone="Asia/Kolkata")
    elif sub.frequency == "weekly":
        dow = sub.day_of_week if sub.day_of_week is not None else 0
        trigger = CronTrigger(day_of_week=dow, hour=sub.hour, minute=sub.minute, timezone="Asia/Kolkata")
    elif sub.frequency == "monthly":
        dom = sub.day_of_month if sub.day_of_month is not None else 1
        trigger = CronTrigger(day=dom, hour=sub.hour, minute=sub.minute, timezone="Asia/Kolkata")
    else:
        return

    scheduler.add_job(
        execute_report_subscription,
        trigger=trigger,
        id=jid,
        args=[sub.id],
        replace_existing=True,
        misfire_grace_time=3600,
    )
    logger.info(f"[report_scheduler] Registered job {jid} ({sub.frequency} at {sub.hour:02d}:{sub.minute:02d})")


def unregister_report_job(sub_id: str):
    from core.scheduler import get_scheduler
    jid = _job_id(sub_id)
    scheduler = get_scheduler()
    if scheduler.get_job(jid):
        scheduler.remove_job(jid)
        logger.info(f"[report_scheduler] Removed job {jid}")


def reload_all_report_subscriptions(db):
    """Called on startup — register all active subscriptions."""
    from models.database import ReportSubscription
    subs = db.query(ReportSubscription).filter(ReportSubscription.is_active == True).all()
    for sub in subs:
        try:
            register_report_job(sub)
        except Exception as e:
            logger.error(f"[report_scheduler] Failed to register {sub.id}: {e}")
    logger.info(f"[report_scheduler] Loaded {len(subs)} subscription(s)")
