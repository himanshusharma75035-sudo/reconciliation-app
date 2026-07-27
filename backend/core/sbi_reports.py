"""SBI Kiosk reference-based reconciliation + the two operator output workbooks.

This is the engine behind the two files finance ops reconcile against by hand:
  * Reconciliation_Report      — bank-statement-centric (which source product each
                                 bank row matched)
  * Source_Files_Match_Status  — source-file-centric (per product: matched / unmatched,
                                 split by transaction Status)

It is **read-only** over `sbi_bank_transactions` + `sbi_txn_reports`. It does NOT touch
the P01-P04 result tables or any matching-engine state — it is an additive reporting
layer. The matching rule is the finance-ops-approved one (see docs/sbi-kiosk.md § Resolved):

  Match the bank statement's 20-digit transaction number (extracted from the
  Description) against the pooled six source files' Reference Number, one-to-one,
  amount agreeing within ₹0.01. Non-20-digit / placeholder references are tagged
  Not Applicable (never Unmatched). Unmatched source rows are split by their source
  Status: Failure = expected (never posted to the bank), Success = a real gap to review.
"""
import io
import re
from collections import defaultdict

import pandas as pd

from models.database import SBIBankTransaction, SBITxnReport, SBIKOLimits, SBIP01Result

# ── matching constants ────────────────────────────────────────────────────────
TOL = 0.01                                   # SBI paisa tolerance (contract-wide)
# A valid SBI txn number is a 20-digit number. Its leading digits encode the date, so the
# prefix is NOT fixed — 15 Jul refs are 6196…, 20 Jul refs are 6201/6202…. (An earlier
# 61-only regex silently rejected every non-61 date; do not reintroduce it.)
_REF_RE = re.compile(r"^\d{20}$")            # a valid SBI txn number: exactly 20 digits
_REF_SCAN = re.compile(r"\d{20}")            # find one inside a bank Description
_PLACEHOLDER = {"", "-", "- / -", "-/-", "- /-", "-/ -", "n/a", "na"}

# canonical product names — these are the six source files + the labels the manual
# report uses. Ingestion stores the raw *filename* in source_file; we normalise it.
PRODUCTS = [
    "AEPS Withdrawal Transaction Report",
    "AePS Onus Deposit",
    "Deposit",
    "Money Transfer",
    "Other Transactions",
    "Withdrawal",
]


def canonical_product(source_file: str) -> str:
    """Map a raw upload filename to one of the six canonical product names.
    Order matters: the specific AEPS/Onus files must win over the generic
    Withdrawal/Deposit ones."""
    s = (source_file or "").lower()
    if "aeps" in s and "transaction" in s:
        return "AEPS Withdrawal Transaction Report"
    if "onus" in s and "deposit" in s:
        return "AePS Onus Deposit"
    if "money" in s and "transfer" in s:
        return "Money Transfer"
    if "other" in s:
        return "Other Transactions"
    if "deposit" in s:
        return "Deposit"
    if "withdrawal" in s:
        return "Withdrawal"
    return (source_file or "Unknown").strip() or "Unknown"


def _valid_ref(ref) -> bool:
    return bool(ref) and bool(_REF_RE.match(str(ref).strip()))


def _extract_bank_ref(b) -> str:
    """The bank row's 20-digit txn number. Ingestion already extracts one into
    ref_number; re-validate it, and fall back to scanning the Description so a
    liberal earlier extraction (which populated cash rows too) can't leak a bad ref."""
    r = (b.ref_number or "").strip()
    if _valid_ref(r):
        return r
    m = _REF_SCAN.search(b.description or "")
    return m.group(0) if m else ""


def _bank_amount(b) -> float:
    d = b.debit or 0
    c = b.credit or 0
    return d if d > 0 else c


def _dedup(rows, keyfn):
    """Keep the first row per key; drop later exact duplicates. Order-preserving."""
    seen = set()
    out = []
    for r in rows:
        k = keyfn(r)
        if k in seen:
            continue
        seen.add(k)
        out.append(r)
    return out


# ── the reconciliation ─────────────────────────────────────────────────────────
def reconcile(db, recon_date: str) -> dict:
    """Reference-match the bank statement against the pooled six source files for one
    business date. Returns a rich dict consumed by both report builders. Read-only."""
    bank = (db.query(SBIBankTransaction)
              .filter(SBIBankTransaction.txn_date == recon_date)
              .order_by(SBIBankTransaction.created_at, SBIBankTransaction.id).all())
    src = (db.query(SBITxnReport)
             .filter(SBITxnReport.txn_date == recon_date)
             .order_by(SBITxnReport.source_file, SBITxnReport.txn_datetime, SBITxnReport.id).all())

    # Defensive exact-duplicate dedup (a file uploaded twice — e.g. a re-export whose
    # SHA differs from the first — doubles a product and floods it with phantom
    # 'Unmatched-Success' rows). Collapse only rows identical on the full business tuple,
    # so genuinely-distinct rows (incl. many sharing the '- / -' placeholder ref) survive.
    # Non-destructive: the DB is untouched; this only cleans the in-memory view.
    src = _dedup(src, lambda s: (canonical_product(s.source_file), (s.reference_number or "").strip(),
                                 round(s.amount or 0, 2), (s.status or "").strip(),
                                 (s.ko_id or "").strip(), (s.txn_datetime or "").strip(),
                                 (s.to_account or "").strip()))
    bank = _dedup(bank, lambda b: (b.txn_date, (b.ref_number or "").strip(),
                                   round(b.debit or 0, 2), round(b.credit or 0, 2),
                                   b.balance, (b.description or "").strip()))

    # index source rows by valid reference; carry per-row derived fields
    src_by_ref = defaultdict(list)
    src_meta = {}                                   # id(s) -> dict of derived state
    for s in src:
        ref = (s.reference_number or "").strip()
        valid = _valid_ref(ref)
        src_meta[id(s)] = {
            "product": canonical_product(s.source_file),
            "valid": valid,
            "match_bank_seq": None,
            "match_bank_date": None,
        }
        if valid:
            src_by_ref[ref].append(s)

    # duplicate txn numbers *within the bank statement* (same ref as both DR and CR)
    ref_bank = defaultdict(list)
    for b in bank:
        r = _extract_bank_ref(b)
        if r:
            ref_bank[r].append(b)
    dup_refs = {r: rows for r, rows in ref_bank.items()
                if len(rows) >= 2 and any((x.debit or 0) > 0 for x in rows)
                and any((x.credit or 0) > 0 for x in rows)}

    # KO-Limits "KO Withdrawal" entries for the date, for SETTLEMENT matching. An EKO
    # DEDUCTION bank debit (is_settlement, no 20-digit ref) reconciles against a KO
    # Withdrawal by (KO id, amount) — the finance-ops rule (date already scoped by
    # recon_date; KO id + deduct_date are parsed from the description at ingestion). Keyed
    # for one-to-one consumption so two equal withdrawals can't both claim one debit.
    ko_wdl = defaultdict(list)   # (ko_id, amount) -> [SBIKOLimits rows]
    for k in db.query(SBIKOLimits).filter(SBIKOLimits.txn_date == recon_date,
                                          SBIKOLimits.txn_type == "KO Withdrawal").all():
        ko_wdl[((k.ko_id or "").strip(), round(k.amount or 0, 2))].append(k)
    used_kw = set()

    bank_recs = []
    used_src = set()
    n_with_ref = n_matched = n_ref_not_found = n_no_ref = n_amt_mismatch = n_reversal = 0
    n_settle = n_settle_matched = n_settle_open = 0

    for i, b in enumerate(bank, start=1):
        ref = _extract_bank_ref(b)
        amt = _bank_amount(b)
        matched = None
        amount_match = ""
        settle = None                                # matched KO Withdrawal (settlement)
        is_settle_debit = bool(b.is_settlement) and (b.debit or 0) > 0
        if ref:
            n_with_ref += 1
            cands = [s for s in src_by_ref.get(ref, []) if id(s) not in used_src]
            # prefer an amount-agreeing source row; else take any remaining same-ref row
            pick = next((s for s in cands if abs((s.amount or 0) - amt) <= TOL), None)
            if pick is None and cands:
                pick = cands[0]
            if pick is not None:
                matched = pick
                used_src.add(id(pick))
                m = src_meta[id(pick)]
                m["match_bank_seq"] = i
                m["match_bank_date"] = b.txn_date
                amount_match = "Yes" if abs((pick.amount or 0) - amt) <= TOL else "No"
                n_matched += 1
                if amount_match == "No":
                    n_amt_mismatch += 1
            elif ref in dup_refs:
                n_reversal += 1            # sibling leg of a DR+CR reversal — explained, not a gap
            else:
                n_ref_not_found += 1
        elif is_settle_debit:
            n_settle += 1
            cands = [k for k in ko_wdl.get(((b.ko_id or "").strip(), round(b.debit or 0, 2)), [])
                     if id(k) not in used_kw]
            if cands:
                settle = cands[0]; used_kw.add(id(settle)); n_settle_matched += 1
            else:
                n_settle_open += 1        # settlement debit with no matching KO Withdrawal (review)
        else:
            n_no_ref += 1

        if matched is not None:
            status = "Matched"
        elif settle is not None:
            status = "Matched (Settlement)"
        elif is_settle_debit:
            status = "Unmatched Settlement"
        elif ref in dup_refs:
            status = "Reversal"
        elif ref:
            status = "Unmatched"
        else:
            status = "No Txn Number"
        bank_recs.append({
            "Bank Stmt Row": i,
            "Txn Date": b.txn_date,
            "Value Date": b.value_date or "",
            "Description": b.description or "",
            "Branch Code": b.branch_code or "",
            "Debit": b.debit or 0,
            "Credit": b.credit or 0,
            "Balance": b.balance if b.balance is not None else "",
            "Extracted Txn No. (20-digit)": ref,
            "Matched Source File": ("KO Limits" if settle else
                                    (canonical_product(matched.source_file) if matched else "")),
            "Matched Transaction Type": ("KO Withdrawal" if settle else
                                         ((matched.txn_type or "") if matched else "")),
            "Source Amount": (settle.amount if settle else ((matched.amount or 0) if matched else "")),
            "Amount Match": ("Yes" if settle else amount_match),
            "Match Status": status,
            "_is_dup": ref in dup_refs,
        })

    # ── source-side classification ─────────────────────────────────────────────
    per_product = {p: {"total": 0, "matched": 0, "unmatched": 0,
                       "unmatched_success": 0, "unmatched_failure": 0, "not_applicable": 0}
                   for p in PRODUCTS}
    source_recs = []                                # every source row, annotated
    src_seq = defaultdict(int)                      # per-product running Sr. No. (proxy)
    for s in src:
        m = src_meta[id(s)]
        prod = m["product"]
        pp = per_product.setdefault(prod, {"total": 0, "matched": 0, "unmatched": 0,
                                           "unmatched_success": 0, "unmatched_failure": 0,
                                           "not_applicable": 0})
        src_seq[prod] += 1
        status = (s.status or "").strip()
        if not m["valid"]:
            match_status = "Not Applicable"          # GUID/placeholder ref — not bank-reconcilable
            pp["not_applicable"] += 1
        elif m["match_bank_seq"] is not None:
            match_status = "Matched"
            pp["matched"] += 1
        else:
            match_status = "Unmatched"
            pp["unmatched"] += 1
            if status.lower() == "success":
                pp["unmatched_success"] += 1
            else:
                pp["unmatched_failure"] += 1
        # 'total' counts only bank-reconcilable rows (exclude Not Applicable from denom)
        if m["valid"]:
            pp["total"] += 1
        source_recs.append({
            "product": prod,
            "Sr. No.": src_seq[prod],
            "KO ID": s.ko_id or "",
            "Transaction Date & Time": s.txn_datetime or "",
            "Reference Number": s.reference_number or "",
            "Type of Transaction": s.txn_type or "",
            "From Account": s.from_account or "",
            "To Account": s.to_account or "",
            "Amount": s.amount or 0,
            "Status": status,
            "Match Status": match_status,
            "Matched Bank Stmt Row": m["match_bank_seq"] if m["match_bank_seq"] else "",
            "Matched Bank Txn Date": m["match_bank_date"] or "",
        })

    totals = {
        "bank_total": len(bank),
        "bank_with_ref": n_with_ref,
        "bank_matched": n_matched,
        "bank_ref_not_found": n_ref_not_found,
        "bank_reversal_legs": n_reversal,
        "bank_settlement_debits": n_settle,
        "bank_settlement_matched": n_settle_matched,
        "bank_settlement_open": n_settle_open,
        "bank_no_ref": n_no_ref,
        "amount_mismatches": n_amt_mismatch,
        "source_total": sum(pp["total"] for pp in per_product.values()),
        "source_matched": sum(pp["matched"] for pp in per_product.values()),
        "source_unmatched": sum(pp["unmatched"] for pp in per_product.values()),
        "source_unmatched_success": sum(pp["unmatched_success"] for pp in per_product.values()),
        "source_unmatched_failure": sum(pp["unmatched_failure"] for pp in per_product.values()),
        "source_not_applicable": sum(pp["not_applicable"] for pp in per_product.values()),
    }

    # unmatched source rows, Success first (the review items float to the top)
    unmatched_source = [r for r in source_recs if r["Match Status"] == "Unmatched"]
    unmatched_source.sort(key=lambda r: (0 if (r["Status"] or "").lower() == "success" else 1,
                                         r["product"], r["Sr. No."]))
    # unmatched bank rows needing review: a ref with no source match, OR a settlement debit
    # with no matching KO Withdrawal
    unmatched_bank = [r for r in bank_recs
                      if r["Match Status"] in ("Unmatched", "Unmatched Settlement")]

    # duplicate groups for the bank sheet
    dup_rows = []
    for gno, (r, rows) in enumerate(sorted(dup_refs.items()), start=1):
        for b in rows:
            seq = next((br["Bank Stmt Row"] for br in bank_recs
                        if br["Extracted Txn No. (20-digit)"] == r
                        and br["Description"] == (b.description or "")), "")
            dup_rows.append({
                "Group No.": gno,
                "Bank Stmt Row": seq,
                "Transaction Number": r,
                "Debit": b.debit or 0,
                "Credit": b.credit or 0,
                "Matched Source File": canonical_product(
                    next((s for s in src_by_ref.get(r, [])), None).source_file)
                    if src_by_ref.get(r) else "",
                "Description": b.description or "",
                "Remarks": "Same txn number posted as both Debit and Credit (reversal pattern)",
            })

    return {
        "recon_date": recon_date,
        "bank_recs": bank_recs,
        "source_recs": source_recs,
        "per_product": per_product,
        "totals": totals,
        "unmatched_source": unmatched_source,
        "unmatched_bank": unmatched_bank,
        "duplicates": dup_rows,
        "dup_ref_count": len(dup_refs),
    }


# ── workbook styling helpers ───────────────────────────────────────────────────
_HEADER_FILL = "094053"      # Eko teal
_GREEN = "C6E0B4"            # matched
_ORANGE = "F8CBAD"          # unmatched-Success → needs review
_RED = "FFC7CE"             # unmatched-Failure → expected
_GREY = "D9D9D9"            # not applicable


def _style_header(ws, ncols):
    from openpyxl.styles import PatternFill, Font
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    for i in range(1, ncols + 1):
        col = ws.cell(row=1, column=i)
        ws.column_dimensions[col.column_letter].width = max(11, min(46, len(str(col.value or "")) + 3))
    ws.freeze_panes = "A2"


def _write_sheet(writer, name, rows, cols):
    df = pd.DataFrame(rows, columns=cols)
    df.to_excel(writer, sheet_name=name[:31], index=False)
    ws = writer.sheets[name[:31]]
    _style_header(ws, len(cols))
    return ws


def _highlight_by_match(ws, rows, match_col_idx, status_col_idx):
    """Colour each data row by its Match Status / Status (green/orange/red/grey)."""
    from openpyxl.styles import PatternFill
    fills = {k: PatternFill("solid", fgColor=v) for k, v in
             {"g": _GREEN, "o": _ORANGE, "r": _RED, "x": _GREY}.items()}
    for i, r in enumerate(rows, start=2):
        ms = r.get("Match Status", "")
        st = (r.get("Status", "") or "").lower()
        key = None
        if ms in ("Matched", "Matched (Settlement)"):
            key = "g"
        elif ms == "Not Applicable":
            key = "x"
        elif ms == "Unmatched Settlement":
            key = "o"                       # settlement debit with no KO Withdrawal — review
        elif ms == "Unmatched":
            key = "o" if st == "success" else "r"
        if key:
            ws.cell(row=i, column=match_col_idx).fill = fills[key]


# ── merge helpers (range mode) ─────────────────────────────────────────────────
# A date RANGE workbook is per-date reconcile() results concatenated chronologically.
# Matching deliberately never crosses a date boundary: pooling the range into one
# reconcile() call would let day-1 bank debits consume day-2 source rows / KO
# Withdrawals and pair reversal legs across days — a silent semantics change of the
# same class as the reverted P03 scope change. Counters are purely additive, so the
# Summary sheets are exact sums of the per-date workbooks.
def _sum_totals(results):
    return {k: sum(r["totals"][k] for r in results) for k in results[0]["totals"]}


def _sum_per_product(results):
    merged = {}
    for r in results:
        for p, pp in r["per_product"].items():
            m = merged.setdefault(p, {k: 0 for k in pp})
            for k, v in pp.items():
                m[k] = m.get(k, 0) + v
    return merged


def _coverage_gaps(results):
    """Dates inside a range whose bank statement or transaction reports are missing.
    On such half-loaded days every row of the present side shows as Unmatched — a
    missing FILE, not a recon failure — so the range workbooks must say so up front
    (the single-date path can't hit this: its resolver requires bank rows).
    Derived from the reconcile results themselves: zero bank rows for a resolved
    date ⇔ no bank statement that day, and likewise for source rows."""
    no_bank = [r["recon_date"] for r in results
               if r["totals"]["bank_total"] == 0 and r["source_recs"]]
    no_src = [r["recon_date"] for r in results
              if r["totals"]["bank_total"] > 0 and not r["source_recs"]]
    return no_bank, no_src


_GAP_NO_BANK = ("⚠ Dates with source files but NO bank-statement rows "
                "(their source rows show Unmatched — missing file, not a recon failure)")
_GAP_NO_SRC = ("⚠ Dates with bank rows but NO transaction reports "
               "(their bank rows show Unmatched — missing file, not a recon failure)")


# ── Report A: Reconciliation_Report (bank-centric, 5 sheets) ────────────────────
def build_reconciliation_report(db, recon_date: str) -> io.BytesIO:
    return _write_reconciliation_workbook([reconcile(db, recon_date)])


def build_reconciliation_report_range(db, dates) -> io.BytesIO:
    """Range variant: one workbook covering several business dates, each date
    reconciled independently (see merge-helper note) and concatenated in date order.
    Sheet shapes match the single-date workbook exactly, plus (range only): a
    'Business Dates Covered' summary row, and a 'Txn Date' column on the duplicates
    sheet — the one sheet whose rows otherwise would not identify their date."""
    return _write_reconciliation_workbook([reconcile(db, d) for d in sorted(set(dates))])


def _write_reconciliation_workbook(results) -> io.BytesIO:
    multi = len(results) > 1
    t = _sum_totals(results)
    per_product = _sum_per_product(results)
    dup_ref_count = sum(r["dup_ref_count"] for r in results)
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Summary — in range mode, one extra leading row naming the dates covered.
        # (The single-date shape below is the finance-ops-verified contract; the
        # 'Matched Count by Source File' block offset is computed from len(summ),
        # so the extra range row shifts it safely.)
        summ = []
        if multi:
            ds = [r["recon_date"] for r in results]
            summ.append({"Metric": "Business Dates Covered",
                         "Value": f"{ds[0]} → {ds[-1]} ({len(ds)} date(s) with data)"})
            no_bank, no_src = _coverage_gaps(results)
            if no_bank:
                summ.append({"Metric": _GAP_NO_BANK, "Value": ", ".join(no_bank)})
            if no_src:
                summ.append({"Metric": _GAP_NO_SRC, "Value": ", ".join(no_src)})
        summ += [
            {"Metric": "Total Bank Statement Transactions", "Value": t["bank_total"]},
            {"Metric": "Bank Rows with a 20-digit Txn Number", "Value": t["bank_with_ref"]},
            {"Metric": "Bank Rows Matched to a Source File", "Value": t["bank_matched"]},
            {"Metric": "Bank Rows with Txn No. but NOT Found in Any Source File", "Value": t["bank_ref_not_found"]},
            {"Metric": "Bank Reversal Legs (DR+CR, same txn no.)", "Value": t["bank_reversal_legs"]},
            {"Metric": "Settlement Debits (EKO DEDUCTION) Matched to a KO Withdrawal", "Value": t["bank_settlement_matched"]},
            {"Metric": "  — Settlement Debits with NO matching KO Withdrawal (review)", "Value": t["bank_settlement_open"]},
            {"Metric": "Bank Rows with No Txn Number (cash / bank-only, mostly credits)", "Value": t["bank_no_ref"]},
            {"Metric": "Amount Mismatches Among Matched Rows", "Value": t["amount_mismatches"]},
            {"Metric": "", "Value": ""},
            {"Metric": "Total Records Across All 6 Source Files (excl. Not Applicable)", "Value": t["source_total"]},
            {"Metric": "Source Records NOT Found in Bank Statement", "Value": t["source_unmatched"]},
            {"Metric": "  — of which Status=Success (needs review)", "Value": t["source_unmatched_success"]},
            {"Metric": "  — of which Status=Failure (expected)", "Value": t["source_unmatched_failure"]},
            {"Metric": "Not-Applicable source rows (non-20-digit ref)", "Value": t["source_not_applicable"]},
            {"Metric": "Duplicate Txn Numbers in Bank Stmt (DR+CR)", "Value": dup_ref_count},
        ]
        _write_sheet(writer, "Summary", summ, ["Metric", "Value"])
        # Matched-by-source-file block appended below the metrics
        ws = writer.sheets["Summary"]
        start = len(summ) + 3
        ws.cell(row=start, column=1, value="Matched Count by Source File").font = \
            __import__("openpyxl").styles.Font(bold=True)
        hdr = ["Source File", "Matched Bank Rows", "Total Records in File", "Unmatched (mostly Failed)"]
        for j, h in enumerate(hdr, start=1):
            ws.cell(row=start + 1, column=j, value=h).font = __import__("openpyxl").styles.Font(bold=True)
        for k, p in enumerate(PRODUCTS, start=start + 2):
            pp = per_product.get(p, {})
            ws.cell(row=k, column=1, value=p)
            ws.cell(row=k, column=2, value=pp.get("matched", 0))
            ws.cell(row=k, column=3, value=pp.get("total", 0))
            ws.cell(row=k, column=4, value=pp.get("unmatched", 0))

        # Bank Statement (Reconciled) — per-date blocks concatenated chronologically.
        # 'Bank Stmt Row' deliberately keeps its per-date meaning (the row's position in
        # THAT day's statement, the operator's cross-reference into the raw file); the
        # adjacent 'Txn Date' column disambiguates in range mode.
        bank_cols = ["Bank Stmt Row", "Txn Date", "Value Date", "Description", "Branch Code",
                     "Debit", "Credit", "Balance", "Extracted Txn No. (20-digit)",
                     "Matched Source File", "Matched Transaction Type", "Source Amount",
                     "Amount Match", "Match Status"]
        bank_rows = [r for R in results for r in R["bank_recs"]]
        wsb = _write_sheet(writer, "Bank Statement (Reconciled)", bank_rows, bank_cols)
        _highlight_by_match(wsb, bank_rows, len(bank_cols), None)

        # Unmatched Bank Entries
        _write_sheet(writer, "Unmatched Bank Entries",
                     [{"Bank Stmt Row": r["Bank Stmt Row"], "Txn Date": r["Txn Date"],
                       "Debit": r["Debit"], "Credit": r["Credit"], "Balance": r["Balance"],
                       "Extracted Txn No.": r["Extracted Txn No. (20-digit)"],
                       "Description": r["Description"],
                       "Remarks": ("Settlement debit (EKO DEDUCTION) with no matching KO Withdrawal"
                                   if r["Match Status"] == "Unmatched Settlement"
                                   else "Has a 20-digit txn no. but no matching source record")}
                      for R in results for r in R["unmatched_bank"]],
                     ["Bank Stmt Row", "Txn Date", "Debit", "Credit", "Balance",
                      "Extracted Txn No.", "Description", "Remarks"])

        # Unmatched Source Records (per-date blocks, Success first within each date)
        us_cols = ["Sr. No.", "KO ID", "Transaction Date & Time", "Reference Number",
                   "Type of Transaction", "From Account", "To Account", "Amount", "Status"]
        _write_sheet(writer, "Unmatched Source Records",
                     [{"Source File": r["product"], **{c: r[c] for c in us_cols}}
                      for R in results for r in R["unmatched_source"]],
                     ["Source File"] + us_cols)

        # Duplicate Txn in Bank Stmt — range mode adds a Txn Date column (these rows
        # otherwise carry no date) right after Group No.; single-date shape unchanged.
        dup_cols = ["Group No.", "Bank Stmt Row", "Transaction Number", "Debit", "Credit",
                    "Matched Source File", "Description", "Remarks"]
        if multi:
            dup_rows = [{"Group No.": d["Group No."], "Txn Date": R["recon_date"],
                         **{k: d[k] for k in dup_cols[1:]}}
                        for R in results for d in R["duplicates"]]
            dup_cols = dup_cols[:1] + ["Txn Date"] + dup_cols[1:]
        else:
            dup_rows = results[0]["duplicates"]
        _write_sheet(writer, "Duplicate Txn in Bank Stmt", dup_rows, dup_cols)
    out.seek(0)
    return out


# ── Report B: Source_Files_Match_Status (source-centric, per-product sheets) ────
_P01_LABEL = {"matched": "Matched", "unmatched": "Unmatched",
              # legacy statuses (pre-2026-07-27) fold into the two-state model
              "CREDITED": "Matched", "PENDING": "Unmatched",
              "EXCESS": "Unmatched", "PARTIAL": "Unmatched"}


def build_source_match_report(db, recon_date: str) -> io.BytesIO:
    return _write_source_match_workbook(db, [reconcile(db, recon_date)])


def build_source_match_report_range(db, dates) -> io.BytesIO:
    """Range variant: per-date reconcile() results concatenated in date order (matching
    never crosses a date boundary — see merge-helper note). Sheet shapes are identical
    to the single-date workbook; the Limit & Settlement sheet groups rows by date with
    the P01 status looked up per (date, KO) so the same KO on two dates can't collide."""
    return _write_source_match_workbook(db, [reconcile(db, d) for d in sorted(set(dates))])


def _write_source_match_workbook(db, results) -> io.BytesIO:
    per_product = _sum_per_product(results)
    t = _sum_totals(results)
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Summary — per source file split by Success/Failure
        rows = []
        for p in PRODUCTS:
            pp = per_product.get(p, {})
            rows.append({
                "Source File": p, "Total Records": pp.get("total", 0),
                "Matched": pp.get("matched", 0), "Unmatched": pp.get("unmatched", 0),
                "Unmatched - Success": pp.get("unmatched_success", 0),
                "Unmatched - Failure": pp.get("unmatched_failure", 0),
                "Not Applicable": pp.get("not_applicable", 0),
            })
        rows.append({"Source File": "TOTAL", "Total Records": t["source_total"],
                     "Matched": t["source_matched"], "Unmatched": t["source_unmatched"],
                     "Unmatched - Success": t["source_unmatched_success"],
                     "Unmatched - Failure": t["source_unmatched_failure"],
                     "Not Applicable": t["source_not_applicable"]})
        _write_sheet(writer, "Summary", rows,
                     ["Source File", "Total Records", "Matched", "Unmatched",
                      "Unmatched - Success", "Unmatched - Failure", "Not Applicable"])
        # Range mode: flag half-loaded days below the table (same computed-offset
        # pattern as Report A's matched-by-source-file block) so a missing upload
        # can't read as a day of real gaps.
        if len(results) > 1:
            no_bank, no_src = _coverage_gaps(results)
            if no_bank or no_src:
                ws = writer.sheets["Summary"]
                r0 = len(rows) + 3
                ws.cell(row=r0, column=1, value="Coverage Notes").font = \
                    __import__("openpyxl").styles.Font(bold=True)
                for note, ds in ((_GAP_NO_BANK, no_bank), (_GAP_NO_SRC, no_src)):
                    if ds:
                        r0 += 1
                        ws.cell(row=r0, column=1, value=note)
                        ws.cell(row=r0, column=2, value=", ".join(ds))

        # Limit & Settlement — raw KO-limits per date + our current P01 status.
        # NOTE: P01 settlement recon is a finance-ops "still open" item (docs/sbi-kiosk.md);
        # the Status shown is our existing P01 output, not a re-validated settlement match.
        # Queried per date (chronological blocks, amount-ordered within a date; Sr. No.
        # runs across the whole sheet) with the P01 lookup scoped to that date.
        ls_rows = []
        for R in results:
            d = R["recon_date"]
            ko = (db.query(SBIKOLimits).filter(SBIKOLimits.txn_date == d)
                    .order_by(SBIKOLimits.amount).all())
            p01 = {r.ko_id: r.status for r in
                   db.query(SBIP01Result).filter(SBIP01Result.recon_date == d).all()}
            ls_rows += [{
                "Sr. No.": len(ls_rows) + j + 1, "Date of Transaction": r.txn_datetime or "",
                "Limit Configured By": r.limit_configured_by or "", "KO ID": r.ko_id or "",
                "Opening Limit": r.opening_limit if r.opening_limit is not None else "",
                "Type of Transaction": r.txn_type or "", "Amount": r.amount or 0,
                "Closing Limit": r.closing_limit if r.closing_limit is not None else "",
                "Settlement Status (P01)": _P01_LABEL.get(p01.get(r.ko_id, ""), p01.get(r.ko_id, "") or "—"),
            } for j, r in enumerate(ko)]
        _write_sheet(writer, "Limit & Settlement", ls_rows,
                     ["Sr. No.", "Date of Transaction", "Limit Configured By", "KO ID",
                      "Opening Limit", "Type of Transaction", "Amount", "Closing Limit",
                      "Settlement Status (P01)"])

        # one sheet per product (rows in date order; Sr. No. restarts per date, the
        # Transaction Date & Time column disambiguates in range mode)
        by_prod = defaultdict(list)
        for R in results:
            for r in R["source_recs"]:
                by_prod[r["product"]].append(r)
        prod_cols = ["Sr. No.", "KO ID", "Transaction Date & Time", "Reference Number",
                     "Type of Transaction", "From Account", "To Account", "Amount", "Status",
                     "Match Status", "Matched Bank Stmt Row", "Matched Bank Txn Date"]
        for p in PRODUCTS:
            recs = by_prod.get(p, [])
            ws = _write_sheet(writer, p, [{c: r[c] for c in prod_cols} for r in recs], prod_cols)
            _highlight_by_match(ws, recs, prod_cols.index("Match Status") + 1, None)
    out.seek(0)
    return out
