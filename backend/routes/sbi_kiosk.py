import logging
logger = logging.getLogger("eko_recon")

"""
routes/sbi_kiosk.py

SBI Kiosk Banking — Full Reconciliation Module (4 Processes)

P01 — SBI Settlement Reconciliation
   KO Limits Config (KO Withdrawal) ↔ Bank Statement (EKOSETTLEMENT)
   Match each withdrawal by amount + the settlement's deduct (business) date; per KO → matched / unmatched

P02 — Bank Statement & Transaction Report Reconciliation
   7 Transaction Report files ↔ Bank Statement
   Match on 20-digit Reference Number; detect Reversals (same ref, DR+CR)

P03 — CSP-Transaction-Bank Reconciliation
   CSP Master + Transaction Report (money OUT) ↔ Bank Statement (money IN)
   CSP Code + Amount with D+1 / D-1 date-shift logic (4-priority)

P04 — CSP Wallet Balance Reconciliation
   Limit Update Failure Report ↔ KO Cash Holding
   Identify under/overfunded wallets; track DEPOSIT / WITHDRAWAL actions
"""

import io
import re
import json
import datetime
from typing import Optional, List

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import func, or_

from models.database import (
    get_db, generate_id, AuditLog,
    SBIBankTransaction, SBITxnReport, SBIKOLimits,
    SBIKOCashHolding, SBILimitFailure, SBICSPMaster,
    SBIP01Result, SBIP02Result, SBIP03Result, SBIP04Result,
    SBIManualMatch, SBISrcAssignment, SBIManualPair,
)
from core.auth import get_current_user, require_permission, require_product_access
from core import maker_checker
from core.sbi_reports import canonical_product, TOL

# Gate on the "kiosk" product id directly (what Users.jsx writes into allowed_products),
# so access does not depend on the PartnerConfig(slug='sbi'->product='kiosk') seed row
# existing — check_product_access falls back to product==partner when no slug matches.
router = APIRouter(prefix="/api/sbi", tags=["sbi-kiosk"],
                   dependencies=[Depends(require_product_access("kiosk"))])

# ── Helpers ───────────────────────────────────────────────────────────────────

def _clip(v, n: int) -> str:
    """Clip a string to a column's max length. MySQL ENFORCES VARCHAR lengths
    (SQLite ignores them), so an over-long parsed value — e.g. a settlement
    description leaking into ref_number — fails the whole upload on prod with
    'Data too long for column …'. Clipping at ingest makes uploads robust on both."""
    s = '' if v is None else str(v)
    return s[:n] if n and len(s) > n else s


def _sf(v, default=0.0) -> float:
    try:
        s = str(v).replace(',', '').strip()
        return float(s) if s and s.lower() not in ('nan', 'none', '') else default
    except Exception:
        return default

def _clean(v) -> str:
    s = str(v).strip() if v is not None else ''
    if s.startswith("'"):                 # Excel text-format marker — artifact, never data
        s = s.lstrip("'").strip()         # (P01–P03 match on RAW equality; see routes.upload._clean)
    return '' if s.lower() in ('nan', 'none') else s

def _nd(v) -> str:
    """Normalise date string to YYYY-MM-DD."""
    import re as _re
    s = _clean(v)
    if not s: return ''
    # DD/MM/YYYY or DD-MM-YYYY
    m = _re.match(r'^(\d{2})[-/](\d{2})[-/](\d{4})', s)
    if m: return f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
    # YYYY-MM-DD already
    if _re.match(r'^\d{4}-\d{2}-\d{2}', s): return s[:10]
    # DD-Mon-YYYY (e.g. 30-May-2026)
    try:
        import dateutil.parser
        return str(dateutil.parser.parse(s).date())
    except Exception:
        return s

def _read_xls_rows(content: bytes) -> list:
    """Read the first sheet of an uploaded workbook into a list of row-lists (cells as
    stripped strings). Tries three readers so a source-format change can't hard-block a
    day's upload (see docs/skills.md):
      1. xlrd      — classic BIFF .xls (what most SBI exports are)
      2. calamine  — robust reader for Java/jExcelApi-written .xls that xlrd rejects with
                     'BOF not workbook/worksheet' (the Limit Update Failure Report is one)
      3. openpyxl  — a genuine .xlsx that arrived with an .xls name
    Returns [] for a validly-parsed but empty sheet; raises only when no reader can open it."""
    import io
    errs = []
    try:
        import xlrd
        sh = xlrd.open_workbook(file_contents=content).sheet_by_index(0)
        return [[str(sh.cell_value(r, c)).strip() for c in range(sh.ncols)]
                for r in range(sh.nrows)]
    except Exception as e:
        errs.append(f"xlrd: {str(e)[:80]}")
    try:
        from python_calamine import CalamineWorkbook
        rows = CalamineWorkbook.from_filelike(io.BytesIO(content)).get_sheet_by_index(0).to_python()
        return [[("" if c is None else str(c)).strip() for c in row] for row in rows]
    except Exception as e:
        errs.append(f"calamine: {str(e)[:80]}")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [[("" if c is None else str(c)).strip() for c in row]
                for row in ws.iter_rows(values_only=True)]
    except Exception as e:
        errs.append(f"openpyxl: {str(e)[:80]}")
    raise ValueError("no reader could open the file — " + " | ".join(errs))

# Empty markers seen in the bank statement's "Ref No./Cheque No." column — never store these
# as a reference (cash-deposit/CDM credits carry no 20-digit ref there, so the column is a bare
# placeholder). Mirrors core.sbi_reports._PLACEHOLDER.
_REFNO_PLACEHOLDER = {"", "-", "- / -", "-/-", "- /-", "-/ -", "n/a", "na"}

def _extract_bank_ref(desc: str) -> str:
    """Extract 20-digit reference number from bank description."""
    m = re.search(r'(?:TO|BY)\s+TRANSFER-(\d{20})', desc)
    return m.group(1) if m else ''

def _clean_refno(ref_no: str) -> str:
    """The bank 'Ref No./Cheque No.' column, blanked when it's just a placeholder like '- / -'."""
    return "" if (ref_no or "").strip().lower() in _REFNO_PLACEHOLDER else ref_no

def _extract_ko_id(desc: str) -> str:
    """Extract KO ID from bank description (after TXN@KO or @KO )."""
    m = re.search(r'TXN@KO\s*(\w+)[-\s]', desc)
    if m: return m.group(1)
    m = re.search(r'@KO\s*(\w+)', desc)
    return m.group(1) if m else ''

def _extract_settlement_info(desc: str):
    """
    For EKOSETTLEMENT rows, extract:
    - KO ID from: EKO DEDUCTION-{KO_ID}.EKOSETTLEMENT
    - Deduction date from: EKO DEDUCTION_{DD-MM-YYYY}
    Returns (ko_id, deduct_date)
    """
    ko_m = re.search(r'EKO DEDUCTION-(\w+)\.EKOSETTLEMENT', desc, re.I)
    dt_m = re.search(r'EKO[_ ]DEDUCTION[_-](\d{2}-\d{2}-\d{4})', desc, re.I)
    ko_id = ko_m.group(1) if ko_m else ''
    deduct_date = _nd(dt_m.group(1)) if dt_m else ''
    return ko_id, deduct_date

def _extract_txn_type_from_bank(desc: str) -> str:
    """Derive transaction type label from bank description."""
    desc_u = desc.upper()
    if 'EKOSETTLEMENT' in desc_u or 'EKO DEDUCTION' in desc_u: return 'Settlement'
    if 'AEPSOFFUSWDL' in desc_u: return 'AEPS OFFUS Withdrawal'
    if 'AEPSWDL' in desc_u: return 'AEPS Withdrawal'
    if 'AEPSDEPOSIT' in desc_u: return 'AEPS Deposit'
    if 'MONEYTRF' in desc_u: return 'Money Transfer'
    if 'WITHDRAWAL' in desc_u: return 'Withdrawal'
    if 'LOAN' in desc_u: return 'Loan'
    if 'RUPAYWDL' in desc_u: return 'Rupay Withdrawal'
    if 'INITIAL' in desc_u: return 'Initial Deposit'
    return 'Other'

def _audit(db: Session, user, action: str, detail: dict, action_type: str = "human"):
    # SBI uploads + P01–P04 runs are deliberate admin actions → "human" (not "app").
    try:
        db.add(AuditLog(
            user_id=user.id, username=user.username,
            action=action, entity_type="sbi_kiosk", entity_id=None,
            action_type=action_type,
            detail=json.dumps(detail),
        ))
        db.commit()
    except Exception:
        db.rollback()


def _sbi_business_dates(db):
    """Every business (txn) date present across the SBI sources, sorted. Drives the
    'reconcile all dates' orchestration so a single bulk upload reconciles each day it
    contains — instead of everything collapsing under one upload-batch recon_date."""
    dates = set()
    # Include KO-Limits business dates and settlement DEDUCT dates: a D+1 settlement's
    # original business date may have no bank/txn rows of its own, but P01 must still
    # reconcile it there (match by deduct_date). Without these, "reconcile all dates"
    # would skip that day.
    for col in (SBIBankTransaction.txn_date, SBITxnReport.txn_date,
                SBIBankTransaction.deduct_date, SBIKOLimits.txn_date):
        for (d,) in db.query(col).filter(col.isnot(None)).distinct().all():
            if d and len(d) >= 10 and d[:4].isdigit():
                dates.add(d[:10])
    return sorted(dates)


# ── P01: Bank Statement Upload ────────────────────────────────────────────────

@router.post("/upload/bank-statement")
async def upload_bank_statement(
    file: UploadFile = File(...),
    recon_date: str = "",
    force: bool = False,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("upload")),
):
    """
    Upload SBI Bank Statement (settlement account).

    Accepts BOTH shapes SBI exports — same 8 columns, same order, in both:
      • tab-separated TEXT carrying a .xls extension (the historical export), and
      • a real .xlsx workbook (account/balance metadata block, then the 'Txn Date'
        header row, then the transactions).
    Both are normalised to rows-of-parts here so the row-building loop below stays a
    SINGLE code path (behaviour for the text export is unchanged).

    Parses all transactions; auto-detects settlement rows via EKOSETTLEMENT keyword.
    Extracts KO ID and deduction date from settlement description.

    This endpoint APPENDS (SBI statements are per-day and accumulate), so re-uploading the
    same file would silently DOUBLE the data — which is exactly what happened on 2026-07-17
    while recovering from an accidental clear. A SHA-256 duplicate-file guard now blocks the
    identical bytes with a 409; pass force=true to re-apply on purpose (e.g. restoring rows
    that were removed since).
    """
    raw = await file.read()

    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "sbi", "bank", raw, file.filename or "", current_user, force=force)

    def _cell(v) -> str:
        """Excel cell → the plain string the tab-separated path would have carried."""
        if v is None: return ''
        if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime('%Y-%m-%d')
        if isinstance(v, float) and v.is_integer(): return str(int(v))   # 10521.0 → "10521"
        s = str(v).strip()
        return '' if s.lower() in ('nan', 'nat', 'none') else s

    # A real workbook starts with the zip magic (xlsx) or the OLE magic (legacy .xls);
    # the historical "tab-separated .xls" is neither — it's plain text.
    if raw[:2] == b'PK' or raw[:4] == b'\xd0\xcf\x11\xe0':
        try:
            df = pd.read_excel(io.BytesIO(raw), header=None, dtype=object)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read the Excel workbook: {e}")
        grid = [[_cell(v) for v in r] for r in df.values.tolist()]
        header_idx = next((i for i, r in enumerate(grid)
                           if any(c.strip().lower() == 'txn date' for c in r)
                           and any(c.strip().lower() == 'description' for c in r)), None)

        def _trim(r):
            """Drop trailing empty cells so a workbook row looks exactly like a tab-split
            line. Without this, pandas pads every row to the full grid width and the
            trailing footer/legend line ("**This is a computer generated statement…")
            survives the `len(parts) < 6` guard below and lands in the ledger as a bogus
            transaction (also poisoning the funds-position balance with a 0.0 line)."""
            r = list(r)
            while r and r[-1] == '':
                r.pop()
            return r

        rows = [_trim(r) for r in grid[header_idx + 1:]] if header_idx is not None else []
    else:
        lines = raw.decode('utf-8', errors='replace').splitlines()
        header_idx = next((i for i, l in enumerate(lines) if 'Txn Date' in l and 'Description' in l), None)
        rows = [l.strip().split('\t') for l in lines[header_idx + 1:]] if header_idx is not None else []

    if header_idx is None:
        raise HTTPException(status_code=400, detail="Cannot find bank statement header row. Expected 'Txn Date' column.")

    inserted = 0
    errors = []
    today = str(datetime.date.today())
    _fund_rows = []   # funds-position lines (file order; reporting only)

    # ── Per-date REPLACE (idempotent ingest) ──────────────────────────────────
    # This endpoint used to pure-APPEND, so re-uploading a statement silently DOUBLED
    # the day's rows (the 2026-07-17 incident). Now it replaces only the dates the file
    # actually contains: delete existing bank rows for those txn_dates, then insert. A
    # re-upload of the same day therefore supersedes it instead of duplicating, and an
    # incremental (new-day) file leaves earlier days intact. Mirrors the E-Value bank fix
    # (behavior-contract item 13 spirit). Combined with the SHA-256 guard above, an
    # accidental same-file re-upload is blocked; a deliberate force re-upload is safe.
    file_dates = {d for d in (_nd((p[0] or "").strip()) for p in rows if p) if d}
    if file_dates:
        db.query(SBIBankTransaction).filter(
            SBIBankTransaction.txn_date.in_(file_dates)).delete(synchronize_session=False)

    for parts in rows:
        if len(parts) < 6: continue
        try:
            txn_date   = _nd(parts[0].strip()) if parts else ''
            value_date = _nd(parts[1].strip()) if len(parts) > 1 else ''
            desc       = parts[2].strip() if len(parts) > 2 else ''
            ref_no     = parts[3].strip().lstrip("'").strip() if len(parts) > 3 else ''  # raw path bypasses _clean
            branch     = parts[4].strip() if len(parts) > 4 else ''
            debit      = _sf(parts[5].strip()) if len(parts) > 5 else 0.0
            credit     = _sf(parts[6].strip()) if len(parts) > 6 else 0.0
            balance    = _sf(parts[7].strip()) if len(parts) > 7 else None

            if not txn_date and not desc: continue

            is_settlement = bool(re.search(r'EKOSETTLEMENT|EKO.DEDUCTION', desc, re.I))
            ko_id_extracted, deduct_date = ('', '')
            if is_settlement:
                ko_id_extracted, deduct_date = _extract_settlement_info(desc)
            else:
                ko_id_extracted = _extract_ko_id(desc)

            ref_extracted = _extract_bank_ref(desc)

            db.add(SBIBankTransaction(
                id            = generate_id(),
                upload_date   = today,
                txn_date      = _clip(txn_date, 10),
                value_date    = _clip(value_date, 10),
                description   = desc,                            # Text — no length cap
                ref_number    = _clip(ref_extracted or _clean_refno(ref_no), 100),
                branch_code   = _clip(branch, 20),
                debit         = debit,
                credit        = credit,
                balance       = balance,
                ko_id         = _clip(ko_id_extracted, 20),
                deduct_date   = _clip(deduct_date, 10),
                is_settlement = is_settlement,
                txn_type      = _clip(_extract_txn_type_from_bank(desc), 30),
            ))
            inserted += 1
            # Funds-position line (file order; reporting only)
            _fund_rows.append({"date": _nd(txn_date), "balance": balance,
                               "dr": debit or 0.0, "cr": credit or 0.0})
        except Exception as e:
            errors.append(f"Line error: {e}")

    db.commit()

    # ── Funds-position: EOD balance snapshots for the SBI settlement account ──
    try:
        from core.funds import record_snapshots
        record_snapshots(db, "kiosk", "sbi", "SBI settlement a/c", _fund_rows,
                         uploaded_by=current_user.username)
    except Exception as _e:
        logger.warning(f"routes/sbi_kiosk.py: funds snapshot skipped: {_e}")

    # M10: validate that we parsed a meaningful number of rows
    # SBI bank statements for a full day typically have hundreds of rows.
    # If inserted < 10, it likely means the file format changed or header detection failed.
    validation_warning = None
    if inserted < 10:
        validation_warning = (
            f"Only {inserted} rows were parsed. The file may have a different format "
            f"than expected. Verify the file is a valid SBI bank statement — either the "
            f"tab-separated export or an .xlsx workbook, each with a 'Txn Date' header row."
        )
        logger.warning(f"SBI bank statement: low row count ({inserted}) for {file.filename}")

    _audit(db, current_user, "sbi_upload_bank_statement",
           {"filename": file.filename, "inserted": inserted, "warning": validation_warning})
    settlement_count = db.query(SBIBankTransaction).filter(
        SBIBankTransaction.upload_date == today,
        SBIBankTransaction.is_settlement == True
    ).count()
    # Reconcile the business dates this statement contains — PLUS the settlement DEDUCT
    # dates (a settlement posted today may belong to an earlier business date; P01 now
    # matches by deduct_date, so its original date must be re-run to pick up the match).
    settle_dd = {d for (d,) in db.query(SBIBankTransaction.deduct_date).filter(
        SBIBankTransaction.upload_date == today,
        SBIBankTransaction.is_settlement == True,
        SBIBankTransaction.deduct_date.isnot(None),
        SBIBankTransaction.deduct_date != '').distinct().all() if d}
    auto_recon = _auto_run_after_upload(db, current_user, dates=sorted(set(file_dates) | settle_dd))
    return {
        "inserted": inserted,
        "settlement_rows": settlement_count,
        "errors": errors[:10],
        "filename": file.filename,
        "validation_warning": validation_warning,
        "auto_recon": auto_recon,
    }


# ── P01: KO Limits Config Upload ──────────────────────────────────────────────

@router.post("/upload/ko-limits")
async def upload_ko_limits(
    file: UploadFile = File(...),
    force: bool = False,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("upload")),
):
    """
    Upload KO Limits Configuration Report (real Excel, 3 metadata rows then header).
    Records every KO Deposit and KO Withdrawal.
    """
    content = await file.read()
    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "sbi", "ko_limits", content, file.filename or "", current_user, force=force)
    try:
        df = pd.read_excel(io.BytesIO(content), engine='xlrd', header=None)
    except Exception:
        try:
            df = pd.read_excel(io.BytesIO(content), header=None)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot parse file: {e}")

    # Header is row 3 (0-indexed)
    h = [_clean(c) for c in df.iloc[3].tolist()]
    data = df.iloc[4:].copy()
    data.columns = h
    today = str(datetime.date.today())
    inserted = 0

    # Parse first, insert after — this upload REPLACES its own scope instead of appending.
    parsed = []
    for _, row in data.iterrows():
        txn_type = _clean(row.get('Type of Transaction', ''))
        if txn_type not in ('KO Deposit', 'KO Withdrawal'): continue
        raw_dt = _clean(row.get('Date of Transaction', ''))
        parsed.append({
            "raw_dt": raw_dt, "td": _nd(raw_dt[:10] if raw_dt else ''),
            "configurer": _clean(row.get('Limit Configured By', '')),
            "ko": _clean(row.get('KO ID', '')), "ol": _sf(row.get('Opening Limit')),
            "tt": txn_type, "amt": _sf(row.get('Amount')), "cl": _sf(row.get('Closing Limit')),
        })
    file_dates = {p["td"] for p in parsed if p["td"]}
    configurers = {p["configurer"] for p in parsed if p["configurer"]}

    # REPLACE, not append: this file supersedes the same dates for the same configurer(s).
    # The old append meant a re-upload (or the force path) DOUBLED KO withdrawals — and
    # P01/P04 SUM per KO per date, so settlement totals doubled. Scoping by configurer
    # keeps another BC's file for the same date untouched (isolation of change).
    replaced = 0
    if file_dates:
        q = db.query(SBIKOLimits).filter(SBIKOLimits.txn_date.in_(file_dates))
        if configurers:
            q = q.filter(SBIKOLimits.limit_configured_by.in_(configurers))
        replaced = q.delete(synchronize_session=False)

    for p in parsed:
        try:
            db.add(SBIKOLimits(
                id                  = generate_id(),
                upload_date         = today,
                txn_datetime        = _clip(p["raw_dt"], 30),
                txn_date            = _clip(p["td"], 10),
                limit_configured_by = _clip(p["configurer"], 20),
                ko_id               = _clip(p["ko"], 20),
                opening_limit       = p["ol"],
                txn_type            = _clip(p["tt"], 30),
                amount              = p["amt"],
                closing_limit       = p["cl"],
            ))
            inserted += 1
        except Exception as _e:
            logger.warning(f"routes/sbi_kiosk.py: {_e}")  # db.commit()
    _audit(db, current_user, "sbi_upload_ko_limits",
           {"filename": file.filename, "inserted": inserted, "replaced": replaced})
    return {"inserted": inserted, "replaced": replaced, "filename": file.filename,
            "auto_recon": _auto_run_after_upload(db, current_user, dates=sorted(file_dates))}


# ── P02: Transaction Report Upload ────────────────────────────────────────────

def _sbi_settlement_accounts():
    """Configured SBI Kiosk settlement-account identifier(s), read from the
    gitignored instance/seed_accounts.json ("sbi_settlement_accounts": [...]).
    Real account numbers never live in the repo. Returns [] when unconfigured,
    which makes the Other-Transactions account filter FAIL OPEN (insert every
    row) so a mis-configured instance never silently drops the whole file."""
    try:
        from models.database import _load_instance_seed_accounts
        vals = _load_instance_seed_accounts().get("sbi_settlement_accounts", []) or []
        return [str(v).strip() for v in vals if str(v).strip()]
    except Exception:
        return []


def _keep_txn_report_row(my_product, fa, ta, settle_accts) -> bool:
    """Whether a Transaction-Report row should be ingested. For the "Other
    Transactions" file only, a row is kept when its From OR To account contains a
    configured SBI settlement account; every other file — and any file when the
    settlement account is unconfigured — keeps all rows (fail open)."""
    if my_product != "Other Transactions" or not settle_accts:
        return True
    return any(a in (fa or "") or a in (ta or "") for a in settle_accts)


@router.post("/upload/txn-report")
async def upload_txn_report(
    file: UploadFile = File(...),
    force: bool = False,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("upload")),
):
    """
    Upload one of the 7 BC Transaction Report files.
    All files share the same column structure (3 metadata rows + header row 4).
    'Other Txn' file has extra REVERSAL_STATUS, SETTELMENT_ACCOUNT_, KO HOLDING columns.

    This endpoint APPENDS (a day has several report files, so per-date replace can't be
    used — it would delete the day's earlier files). A SHA-256 duplicate-file guard blocks
    re-uploading the EXACT same file (which would double it); pass force=true to override.
    """
    content = await file.read()
    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "sbi", "txn_report", content, file.filename or "", current_user, force=force)
    try:
        df = pd.read_excel(io.BytesIO(content), engine='xlrd', header=None)
    except Exception:
        try:
            df = pd.read_excel(io.BytesIO(content), header=None)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot parse file: {e}")

    # Header is row 3
    raw_h = df.iloc[3].tolist()
    h = [_clean(c) for c in raw_h]
    data = df.iloc[4:].copy()
    data.columns = h

    # Normalize column names (some files have empty cols mixed in)
    col_map = {c.lower().replace(' ', '_').replace('/', '_'): c for c in h if c}
    def gc(key): return col_map.get(key, '')
    def gv(row, key): return _clean(row.get(gc(key), '')) if gc(key) else ''

    today = str(datetime.date.today())
    source = file.filename
    my_product = canonical_product(source)
    # Point-4 (Rajendra): in the "Other Transactions" file, only rows that touch our SBI
    # settlement account (From OR To account) are ours to reconcile — a customer↔customer
    # ATM fund-transfer must NOT be ingested. Applies to that file ONLY; fails open when
    # the account is unconfigured. The account value comes from gitignored instance config.
    settle_accts = _sbi_settlement_accounts()
    inserted = skipped_non_settlement = 0

    # Parse first, insert after — this upload REPLACES its own scope instead of appending.
    parsed = []
    file_dates = set()   # only reconcile the business dates THIS file touches (see below)
    for _, row in data.iterrows():
        # Skip metadata/empty rows
        sr = gv(row, 'sr._no') or gv(row, 'sr._no.')
        if not sr or not str(sr).replace('.', '').strip().isdigit(): continue
        raw_dt = gv(row, 'transaction_date_&_time') or gv(row, 'transaction_date')
        txn_date = _nd(raw_dt[:10]) if raw_dt else ''
        if txn_date: file_dates.add(txn_date)
        parsed.append((raw_dt, txn_date, row))

    # REPLACE, not append: a day has ~7 report files that legitimately share dates, so a
    # blanket per-date replace would destroy the day's other files (that append was
    # deliberate). Instead this file supersedes rows of the SAME PRODUCT for ITS dates:
    # product comes from canonical_product(source_file), which is stable across renamed
    # re-downloads ("report.xls" vs "report (1).xls") — so a re-upload (or the force
    # path) replaces its own prior rows and can add new entries, never double them.
    replaced = 0
    if file_dates:
        prior_files = [sf for (sf,) in db.query(SBITxnReport.source_file)
                       .filter(SBITxnReport.txn_date.in_(file_dates)).distinct().all()
                       if canonical_product(sf) == my_product]
        if prior_files:
            replaced = (db.query(SBITxnReport)
                        .filter(SBITxnReport.txn_date.in_(file_dates),
                                SBITxnReport.source_file.in_(prior_files))
                        .delete(synchronize_session=False))

    for raw_dt, txn_date, row in parsed:
        fa = _clip(gv(row, 'from_account'), 30)
        ta = _clip(gv(row, 'to_account'), 30)
        # Other-Transactions only: skip rows whose From AND To account both miss our
        # settlement account (customer↔customer transfers we don't reconcile).
        if not _keep_txn_report_row(my_product, fa, ta, settle_accts):
            skipped_non_settlement += 1
            continue
        try:
            db.add(SBITxnReport(
                id              = generate_id(),
                upload_date     = today,
                source_file     = _clip(source, 100),
                ko_id           = _clip(gv(row, 'ko_id'), 20),
                txn_datetime    = _clip(raw_dt, 30),
                txn_date        = _clip(txn_date, 10),
                reference_number= _clip(gv(row, 'reference_number'), 100),
                txn_type        = _clip(gv(row, 'type_of_transaction'), 80),
                from_account    = fa,
                to_account      = ta,
                amount          = _sf(row.get(gc('amount'), 0)),
                customer_charge = _sf(row.get(gc('customer_charge'), 0)),
                journal_number  = _clip(gv(row, 'journal_number'), 30),
                status          = _clip(gv(row, 'status'), 20),
                reversal_status = _clip(gv(row, 'reversal_status'), 20),
                settlement_acct = _clip(gv(row, 'settelment_account_'), 20),
                ko_holding      = _sf(row.get(gc('ko_holding'), None)),
            ))
            inserted += 1
        except Exception as _e:
            logger.warning(f"routes/sbi_kiosk.py: {_e}")  # db.commit()
    _audit(db, current_user, "sbi_upload_txn_report",
           {"filename": source, "inserted": inserted, "replaced": replaced,
            "skipped_non_settlement": skipped_non_settlement})
    # Reconcile ONLY the dates this file touched — not every business date. A full re-run
    # (22 dates × P01-P04) took minutes and blocked the single worker, freezing the app on
    # every upload. `dates=[]` (a file with no parseable dates) → skip, don't fall back to all.
    return {"inserted": inserted, "replaced": replaced, "filename": source,
            "skipped_non_settlement": skipped_non_settlement,
            "auto_recon": _auto_run_after_upload(db, current_user, dates=sorted(file_dates))}


# ── P03: CSP Master Sheet Upload ──────────────────────────────────────────────

@router.post("/upload/csp-master")
async def upload_csp_master(
    file: UploadFile = File(...),
    force: bool = False,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("upload")),
):
    """
    Upload CSP Master Sheet (CSP Code | Ref Number | Mode).
    Cash: fixed ref per CSP (same ref across multiple deposits).
    Electronic/CDM: each row is a unique transaction reference.
    """
    content = await file.read()
    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "sbi", "csp_master", content, file.filename or "", current_user, force=force)
    try:
        df = pd.read_excel(io.BytesIO(content), header=None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot parse: {e}")

    # Header in row 0
    h = [_clean(c) for c in df.iloc[0].tolist()]
    data = df.iloc[1:].copy()
    data.columns = h
    today = str(datetime.date.today())
    inserted = 0

    # DEDUP, not append: reference data must never double. Rows identical on the full
    # business tuple are skipped, so a re-upload (or the force path) is idempotent and a
    # second file with additional CSPs inserts only the new ones.
    existing = {(r.csp_code, r.ref_number, r.mode)
                for r in db.query(SBICSPMaster.csp_code, SBICSPMaster.ref_number,
                                  SBICSPMaster.mode).all()}
    skipped = 0
    for _, row in data.iterrows():
        csp = _clean(row.get('CSP Code', ''))
        if not csp: continue
        rec = (_clip(csp, 20),
               _clip(_clean(row.get('Ref.. Number', '') or row.get('Ref Number', '')), 100),
               _clip(_clean(row.get('Mood', '') or row.get('Mode', '')), 30))
        if rec in existing:
            skipped += 1
            continue
        existing.add(rec)
        db.add(SBICSPMaster(
            id          = generate_id(),
            upload_date = today,
            csp_code    = rec[0],
            ref_number  = rec[1],
            mode        = rec[2],
        ))
        inserted += 1

    db.commit()
    _audit(db, current_user, "sbi_upload_csp_master",
           {"filename": file.filename, "inserted": inserted, "skipped_duplicates": skipped})
    # CSP Master is reference data (no business date). It feeds P03's mode lookup across ALL
    # dates, so auto-reconciling it would mean a full re-run that freezes the worker — and
    # it changes rarely. Skip auto-recon; use "Reconcile all dates" to propagate a change.
    return {"inserted": inserted, "skipped_duplicates": skipped, "filename": file.filename,
            "auto_recon": _auto_run_after_upload(db, current_user, dates=[])}


# ── P04: KO Cash Holding Upload ───────────────────────────────────────────────

@router.post("/upload/ko-cash-holding")
async def upload_ko_cash_holding(
    file: UploadFile = File(...),
    report_date: str = "",
    force: bool = False,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("upload")),
):
    """Upload KO Cash Holding Report (KO ID + opening/closing balances)."""
    content = await file.read()
    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "sbi", "cash_holding", content, file.filename or "", current_user, force=force)
    try:
        df = pd.read_excel(io.BytesIO(content), engine='xlrd', header=None)
    except Exception:
        df = pd.read_excel(io.BytesIO(content), header=None)

    h = [_clean(c) for c in df.iloc[3].tolist()]
    data = df.iloc[4:].copy()
    data.columns = h
    today = str(datetime.date.today())
    r_date = _nd(report_date) if report_date else today
    inserted = 0

    # Parse first — this upload REPLACES the same report_date's rows for the KOs it
    # re-states (a re-upload/force must never double per-KO balances; another BC's file
    # for the same date is untouched).
    parsed = []
    for _, row in data.iterrows():
        ko = _clean(row.get('KO ID', ''))
        if not ko: continue
        parsed.append((ko, row))
    replaced = 0
    file_kos = {_clip(ko, 20) for ko, _ in parsed}
    if r_date and file_kos:
        replaced = (db.query(SBIKOCashHolding)
                    .filter(SBIKOCashHolding.report_date == _clip(r_date, 10),
                            SBIKOCashHolding.ko_id.in_(file_kos))
                    .delete(synchronize_session=False))

    for ko, row in parsed:
        db.add(SBIKOCashHolding(
            id              = generate_id(),
            upload_date     = today,
            report_date     = _clip(r_date, 10),
            ko_id           = _clip(ko, 20),
            limit           = _sf(row.get('Limit')),
            opening_balance = _sf(row.get('Opening Balance')),
            cash_receipts   = _sf(row.get('Cash Receipts')),
            cash_payments   = _sf(row.get('Cash Payments')),
            ko_deposit      = _sf(row.get('KO Deposit')),
            ko_withdrawal   = _sf(row.get('KO Withdrawal')),
            closing_balance = _sf(row.get('Closing Balance')),
        ))
        inserted += 1

    db.commit()
    _audit(db, current_user, "sbi_upload_ko_cash_holding",
           {"filename": file.filename, "inserted": inserted, "replaced": replaced})
    return {"inserted": inserted, "replaced": replaced, "filename": file.filename,
            "auto_recon": _auto_run_after_upload(db, current_user, dates=[r_date] if r_date else [])}


@router.post("/upload/limit-failures")
async def upload_limit_failures(
    file: UploadFile = File(...),
    force: bool = False,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("upload")),
):
    """Upload Limit Update Failure Report (CSPs whose wallet limit update failed)."""
    content = await file.read()
    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "sbi", "limit_failures", content, file.filename or "", current_user, force=force)
    try:
        rows = _read_xls_rows(content)   # xlrd → calamine (handles the JXL .xls) → openpyxl
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot parse Limit Fail file: {e}")

    # A Limit Update Failure Report with no failures that day is a VALID empty file — the
    # data rows start at row 4 (rows 0-2 metadata, row 3 header), so an empty/short file
    # simply inserts nothing rather than erroring.
    today = str(datetime.date.today())
    inserted = 0

    # Parse first — this upload REPLACES its own scope (its dates for its BC ids) instead
    # of appending, so a re-upload/force never doubles failures; another BC's file for
    # the same date is untouched.
    parsed = []
    file_dates = set()
    for row in rows[4:]:  # data starts at row 4
        if len(row) < 5: continue
        sr = row[0].replace('.', '').strip()
        if not sr or not sr.isdigit(): continue
        _td = _nd(row[1])
        if _td: file_dates.add(_td)
        parsed.append((_td, row))
    file_bcs = {_clip(r[3].strip(), 20) for _, r in parsed if len(r) > 3 and r[3].strip()}
    replaced = 0
    if file_dates:
        q = db.query(SBILimitFailure).filter(SBILimitFailure.txn_date.in_(file_dates))
        if file_bcs:
            q = q.filter(SBILimitFailure.bc_id.in_(file_bcs))
        replaced = q.delete(synchronize_session=False)

    for _td, row in parsed:
        try:
            db.add(SBILimitFailure(
                id          = generate_id(),
                upload_date = today,
                txn_date    = _clip(_td, 10),
                csp_code    = _clip(row[2].strip(), 20),
                bc_id       = _clip(row[3].strip(), 20),
                amount      = _sf(row[4]),
                user        = _clip(row[5].strip() if len(row) > 5 else '', 50),
            ))
            inserted += 1
        except Exception as _e:
            logger.warning(f"routes/sbi_kiosk.py: {_e}")  # db.commit()
    _audit(db, current_user, "sbi_upload_limit_failures",
           {"filename": file.filename, "inserted": inserted, "replaced": replaced})
    return {"inserted": inserted, "replaced": replaced, "filename": file.filename,
            "auto_recon": _auto_run_after_upload(db, current_user, dates=sorted(file_dates))}


# ── P01: Run Settlement Reconciliation ────────────────────────────────────────

@router.post("/run/p01")
def run_p01(
    recon_date: str,
    upload_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("run_recon")),
):
    """
    P01 — SBI Settlement Reconciliation → **matched / unmatched** (finance-ops rule,
    Rajendra 2026-07-27).

    Each KO wallet withdrawal (KO Limits Config, business date = recon_date) is matched
    to a bank settlement debit by EXACT amount (±0.01, the SBI tolerance) and the
    settlement's DEDUCT date — the date inside the EKO DEDUCTION description — NOT the
    bank's posting date. So a settlement that posts a day (or two) later still reconciles
    on the original business date (fixes the old D+1 → PENDING+EXCESS artifact).

    One result row per KO: 'matched' iff every withdrawal has a settlement AND no
    settlement is left over (amounts pair up 1:1); otherwise 'unmatched' (some wallet
    withdrawal not settled, some bank settlement with no withdrawal — e.g. the KO Limits
    file for that date isn't uploaded, or a real short/excess). No PARTIAL/PENDING/EXCESS.
    """
    from collections import defaultdict
    from sqlalchemy import or_, and_
    # KO Withdrawals for this BUSINESS date (txn_date IS the deduct/business date).
    wdl_by_ko = defaultdict(list)   # ko_id → [withdrawal amounts]
    for r in db.query(SBIKOLimits).filter(
        SBIKOLimits.txn_type == 'KO Withdrawal', SBIKOLimits.txn_date == recon_date).all():
        if r.ko_id:
            wdl_by_ko[r.ko_id].append(round(r.amount or 0, 2))

    # Bank settlements whose DEDUCT date (from the description) is this business date —
    # regardless of which day the bank posted them. Rows with no parseable deduct date
    # fall back to their posting date so nothing is silently dropped.
    setl_by_ko = defaultdict(list)  # ko_id → [(amount, posting_txn_date)]
    for r in db.query(SBIBankTransaction).filter(
        SBIBankTransaction.is_settlement == True,
        or_(SBIBankTransaction.deduct_date == recon_date,
            and_(or_(SBIBankTransaction.deduct_date.is_(None), SBIBankTransaction.deduct_date == ''),
                 SBIBankTransaction.txn_date == recon_date))).all():
        if r.ko_id:
            setl_by_ko[r.ko_id].append((round(r.debit or 0, 2), r.txn_date))

    # WIPE-GUARD: nothing to reconcile for this date → do NOT delete existing results.
    if not wdl_by_ko and not setl_by_ko:
        return {"recon_date": recon_date, "skipped": True,
                "reason": "no P01 settlement source data for this date", "total_kos": 0}
    db.query(SBIP01Result).filter(SBIP01Result.recon_date == recon_date).delete()

    TOL = 0.01
    results = []
    for ko in set(wdl_by_ko) | set(setl_by_ko):
        wdls  = sorted(wdl_by_ko.get(ko, []))
        setls = sorted(setl_by_ko.get(ko, []), key=lambda x: x[0])
        used  = [False] * len(setls)
        unmatched_wdl = 0
        for a in wdls:                                  # greedy 1:1 by amount
            hit = next((i for i, (s, _td) in enumerate(setls)
                        if not used[i] and abs(s - a) <= TOL), None)
            if hit is not None:
                used[hit] = True
            else:
                unmatched_wdl += 1
        unmatched_setl = sum(1 for u in used if not u)
        # matched only when BOTH sides pair up completely (and there was a withdrawal)
        status = 'matched' if (wdls and unmatched_wdl == 0 and unmatched_setl == 0) else 'unmatched'
        wallet_amt = round(sum(wdls), 2)
        bank_amt   = round(sum(s for s, _ in setls), 2)
        bank_td    = max((td for _, td in setls), default='') if setls else ''
        results.append(SBIP01Result(
            id=generate_id(), recon_date=recon_date, ko_id=ko,
            wallet_withdrawn=wallet_amt, bank_settled=bank_amt,
            difference=round(bank_amt - wallet_amt, 2), status=status,
            deduct_date=recon_date, bank_txn_date=bank_td,
            notes=(f"settled later ({bank_td}) — matched on business date {recon_date}"
                   if status == 'matched' and bank_td and bank_td != recon_date else ''),
        ))
    for r in results:
        db.add(r)
    db.commit()
    summary = {s: sum(1 for r in results if r.status == s) for s in ('matched', 'unmatched')}
    _audit(db, current_user, "sbi_run_p01", {"recon_date": recon_date, **summary})
    return {"recon_date": recon_date, "total_kos": len(results), "summary": summary}


# ── P02: Run Bank + Transaction Report Reconciliation ─────────────────────────

@router.post("/run/p02")
def run_p02(
    recon_date: str,
    upload_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("run_recon")),
):
    """
    P02 — Bank Statement & Transaction Report Reconciliation.
    Primary match: Reference Number (20-digit, extracted from bank description).
    Reversal: same ref appears as both DR and CR → tagged as Reversal Debit / Credit.
    Statuses: Matched / Unmatched / Partial / Reversal
    """
    # Source by the transaction's BUSINESS date (recon_date), not the upload batch.
    bank_txns = db.query(SBIBankTransaction).filter(
        SBIBankTransaction.txn_date == recon_date
    ).all()

    # WIPE-GUARD: no bank rows for this date → do NOT delete existing results.
    if not bank_txns:
        return {"recon_date": recon_date, "skipped": True,
                "reason": "no bank statement rows for this date", "total": 0}
    db.query(SBIP02Result).filter(SBIP02Result.recon_date == recon_date).delete()

    # Reference-number index from the transaction reports for this date
    txn_by_ref = {}  # ref → list of SBITxnReport rows
    for r in db.query(SBITxnReport).filter(SBITxnReport.txn_date == recon_date).all():
        if r.reference_number:
            txn_by_ref.setdefault(r.reference_number, []).append(r)

    # First pass: detect reversals (same ref, DR + CR pair). Only real 20-digit references
    # count — the "no reference" placeholder ('- / -') must NOT group here, or every no-ref
    # cash/settlement row (which naturally spans debits and credits) gets lumped into one
    # bogus "reversal" (it was inflating Reversal by ~977 rows across the data).
    from core.sbi_reports import _valid_ref
    ref_bank_map = {}  # ref → list of (bank_txn, type)
    for bt in bank_txns:
        if not _valid_ref((bt.ref_number or "").strip()): continue
        ref_bank_map.setdefault(bt.ref_number, []).append(bt)

    reversal_refs = {
        ref for ref, txns in ref_bank_map.items()
        if any(t.debit > 0 for t in txns) and any(t.credit > 0 for t in txns)
    }

    results = []
    used_txn_ids = set()

    for bt in bank_txns:
        ref = bt.ref_number
        bank_type = 'DR' if bt.debit > 0 else 'CR'
        bank_amount = bt.debit if bt.debit > 0 else bt.credit

        # Reversal detection
        is_reversal = ref in reversal_refs
        reversal_type = ''
        if is_reversal:
            reversal_type = 'Reversal Debit' if bank_type == 'DR' else 'Reversal Credit'

        # Match against transaction reports
        matched_txn = None
        if ref and ref in txn_by_ref:
            for txn in txn_by_ref[ref]:
                if txn.id not in used_txn_ids:
                    matched_txn = txn
                    used_txn_ids.add(txn.id)
                    break

        if is_reversal:
            # A DR+CR pair sharing one reference cancels out — it IS reconciled, not a separate
            # bucket. Classify as Matched (finance-ops: "either matched or unmatched, never
            # reversal"); the reversal nature is preserved in reversal_type for the Reversals report.
            match_status = 'Matched'
            success = 'Success'
        elif matched_txn:
            # Validate amount — SBI paisa tolerance (₹0.01, finance-ops-approved; was ₹1,
            # unsupported by the data: 0 amount mismatches across 27,187 validated matches).
            amt_diff = abs(bank_amount - (matched_txn.amount or 0))
            match_status = 'Matched' if amt_diff <= 0.01 else 'Partial'
            success = matched_txn.status or 'Success'
        elif not ref:
            match_status = 'Unmatched'
            success = 'Fail'
        else:
            match_status = 'Unmatched'
            success = 'Fail'

        result = SBIP02Result(
            id               = generate_id(),
            recon_date       = recon_date,
            bank_txn_id      = bt.id,
            txn_report_id    = matched_txn.id if matched_txn else None,
            reference_number = ref,
            ko_id            = bt.ko_id or (matched_txn.ko_id if matched_txn else ''),
            bank_amount      = bank_amount,
            bank_type        = bank_type,
            report_amount    = matched_txn.amount if matched_txn else None,
            report_txn_type  = matched_txn.txn_type if matched_txn else '',
            match_status     = match_status,
            reversal_type    = reversal_type,
            success_status   = success,
            notes            = bt.txn_type,
        )
        db.add(result)
        results.append(result)

    db.commit()
    summary = {s: sum(1 for r in results if r.match_status == s)
               for s in ('Matched', 'Unmatched', 'Partial')}
    # reversal legs are Matched now (they cancel out); surface the count separately for info.
    summary['Reversal'] = sum(1 for r in results if (r.reversal_type or '') not in ('', 'No'))
    _audit(db, current_user, "sbi_run_p02", {"recon_date": recon_date, **summary})
    return {"recon_date": recon_date, "total": len(results), "summary": summary}


# ── P03: Run CSP-Transaction-Bank Reconciliation ──────────────────────────────

@router.post("/run/p03")
def run_p03(
    recon_date: str,
    upload_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("run_recon")),
):
    """
    P03 — CSP-Transaction-Bank Reconciliation.
    Money OUT (Transaction Report debits to CSP) ↔ Money IN (Bank credits from CSP).
    Match key: CSP Code + Amount.
    4-priority date logic (SOP Section 4):
      P1: same-day | P2: D+1 bank→txn | P3: D+1 txn→bank | P4: D-1 txn→bank
    One-to-one matching (no duplicate assignment).
    """
    # CSP Master is reference data (mode lookup) — not date-scoped.
    csp_modes = {}
    for r in db.query(SBICSPMaster).all():
        if r.csp_code not in csp_modes:
            csp_modes[r.csp_code] = r.mode

    # Source by BUSINESS date (recon_date). NOTE: Phase 1 matches same-day only — the
    # cross-day D±1 shift window is a Phase-3 item (its ±2 window and cross-date one-to-one
    # need finance sign-off); running per business date with a window would otherwise let a
    # bank credit be reused across two dates' runs.
    txn_rows = []
    for r in db.query(SBITxnReport).filter(SBITxnReport.txn_date == recon_date).all():
        if r.status and r.status.lower() == 'success' and r.amount and r.amount > 0:
            txn_rows.append(r)

    bank_credits = []
    for r in db.query(SBIBankTransaction).filter(
        SBIBankTransaction.txn_date == recon_date,
        SBIBankTransaction.credit > 0,
        SBIBankTransaction.is_settlement == False,
    ).all():
        bank_credits.append(r)

    # WIPE-GUARD: nothing to reconcile for this date → do NOT delete existing results.
    if not txn_rows and not bank_credits:
        return {"recon_date": recon_date, "skipped": True,
                "reason": "no P03 (txn report / bank credit) source data for this date", "total": 0}
    db.query(SBIP03Result).filter(SBIP03Result.recon_date == recon_date).delete()

    # Build bank lookup: (ko_id, amount, date) → bank row
    # Date offsets checked: 0, +1, -1
    #
    # NOTE (2026-07-22): a reference-number match was tried here (docs/sbi-kiosk.md §Resolved
    # item 1) and REVERTED. It is more *precise* (the (ko,amount) key makes ~1,760 false
    # matches/day — money-IN txns, deposits/transfers, matched to unrelated credits), but P03
    # only looks at bank CREDITS, so those money-in txns (whose real counterpart is a bank
    # DEBIT) then read as Unmatched. Completing that fix = matching against ALL bank rows,
    # which is the P02/P03 MERGE that is still open and needs finance-ops confirmation. Until
    # then the fully-correct reconciliation is the reference-based report (core/sbi_reports.py).
    bank_by_ko_amt = {}
    for b in bank_credits:
        key = (b.ko_id, round(b.credit, 2))
        bank_by_ko_amt.setdefault(key, []).append(b)

    results = []
    used_bank_ids = set()

    def _find_bank(ko_id, amount, txn_date_str, max_shift=1):
        """Find best matching bank credit by KO ID + Amount with date shifting."""
        key = (ko_id, round(amount, 2))
        candidates = bank_by_ko_amt.get(key, [])
        # Sort candidates by date proximity
        best = None
        best_shift = 999
        for b in candidates:
            if b.id in used_bank_ids: continue
            try:
                from datetime import date
                t = date.fromisoformat(txn_date_str) if txn_date_str else None
                bd = date.fromisoformat(b.txn_date) if b.txn_date else None
                if t and bd:
                    shift = (bd - t).days
                    if abs(shift) <= max_shift and abs(shift) < abs(best_shift):
                        best = b
                        best_shift = shift
            except Exception:
                if best is None:
                    best = b
                    best_shift = 0
        return best, best_shift if best else 999

    # Match each transaction report row against bank credits
    for txn in txn_rows:
        ko = txn.ko_id
        amt = txn.amount or 0
        mode = csp_modes.get(ko, 'Unknown')

        bank_match, shift = _find_bank(ko, amt, txn.txn_date, max_shift=2)

        if bank_match:
            used_bank_ids.add(bank_match.id)
            if shift == 0: priority = 1
            elif shift == 1: priority = 2   # D+1 bank
            elif shift == -1: priority = 4   # D-1 bank
            else: priority = 3

            result = SBIP03Result(
                id              = generate_id(),
                recon_date      = recon_date,
                csp_code        = ko,
                mode            = mode,
                ref_number      = txn.reference_number,
                txn_amount      = amt,
                txn_date        = txn.txn_date,
                bank_credit_date= bank_match.txn_date,
                bank_amount     = bank_match.credit,
                match_status    = 'Matched',
                date_shift      = shift,
                match_priority  = priority,
                notes           = f"Shift {shift:+d}d" if shift != 0 else '',
            )
        else:
            result = SBIP03Result(
                id              = generate_id(),
                recon_date      = recon_date,
                csp_code        = ko,
                mode            = mode,
                ref_number      = txn.reference_number,
                txn_amount      = amt,
                txn_date        = txn.txn_date,
                bank_credit_date= None,
                bank_amount     = None,
                match_status    = 'Unmatched_TxnReport',
                date_shift      = None,
                match_priority  = None,
                notes           = 'No matching bank credit found (checked ±2 days)',
            )

        db.add(result)
        results.append(result)

    # Unmatched bank credits
    for b in bank_credits:
        if b.id not in used_bank_ids:
            mode = csp_modes.get(b.ko_id, 'Unknown')
            result = SBIP03Result(
                id              = generate_id(),
                recon_date      = recon_date,
                csp_code        = b.ko_id,
                mode            = mode,
                ref_number      = b.ref_number,
                txn_amount      = None,
                txn_date        = None,
                bank_credit_date= b.txn_date,
                bank_amount     = b.credit,
                match_status    = 'Unmatched_Bank',
                date_shift      = None,
                match_priority  = None,
                notes           = 'Bank credit with no matching transaction report entry',
            )
            db.add(result)
            results.append(result)

    db.commit()
    summary = {s: sum(1 for r in results if r.match_status == s)
               for s in ('Matched', 'Unmatched_TxnReport', 'Unmatched_Bank')}
    _audit(db, current_user, "sbi_run_p03", {"recon_date": recon_date, **summary})
    return {"recon_date": recon_date, "total": len(results), "summary": summary}


# ── P04: Run Wallet Balance Reconciliation ────────────────────────────────────

@router.post("/run/p04")
def run_p04(
    recon_date: str,
    upload_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("run_recon")),
):
    """
    P04 — CSP Wallet Balance Reconciliation.
    For each KO in the Limit Update Failure Report:
      - Find their Closing Balance in KO Cash Holding Report
      - Determine whether wallet needs DEPOSIT or WITHDRAWAL correction
    Action is flagged; team marks done after performing action in SBI portal.
    """
    # Source by BUSINESS date (recon_date), not the upload batch.
    # KO Cash Holding closing balances for this report date.
    cash_holding = {}
    for r in db.query(SBIKOCashHolding).filter(SBIKOCashHolding.report_date == recon_date).all():
        cash_holding[r.ko_id] = r.closing_balance or 0

    # Limit Failures for this date (P04 is driven by the failure report)
    failures = db.query(SBILimitFailure).filter(
        SBILimitFailure.txn_date == recon_date
    ).all()

    # WIPE-GUARD: no failures for this date → nothing to reconcile → don't delete existing.
    if not failures:
        return {"recon_date": recon_date, "skipped": True,
                "reason": "no limit-failure rows for this date", "total": 0}
    db.query(SBIP04Result).filter(SBIP04Result.recon_date == recon_date).delete()

    results = []
    for f in failures:
        closing = cash_holding.get(f.csp_code, 0)
        failed_amt = f.amount or 0  # negative = withdrawal attempt, positive = deposit attempt

        # Expected balance after the failed operation would have succeeded
        expected = closing + failed_amt   # if withdrawal failed, wallet has MORE than expected

        diff = closing - expected         # actual vs expected
        abs_diff = abs(diff)

        if abs_diff < 0.01:
            action = 'NONE'
            action_amt = 0.0
        elif diff > 0:
            # Wallet has MORE than expected → withdrawal needed
            action = 'WITHDRAWAL'
            action_amt = diff
        else:
            # Wallet has LESS than expected → deposit needed
            action = 'DEPOSIT'
            action_amt = abs_diff

        result = SBIP04Result(
            id              = generate_id(),
            recon_date      = recon_date,
            csp_code        = f.csp_code,
            failed_amount   = failed_amt,
            closing_balance = closing,
            expected_balance= expected,
            difference      = round(diff, 2),
            action_required = action,
            action_amount   = round(action_amt, 2),
            action_done     = False,
            notes           = f"Limit update failed on {f.txn_date}: {failed_amt:,.0f}. "
                              f"Current balance: {closing:,.0f}, Expected: {expected:,.0f}",
        )
        db.add(result)
        results.append(result)

    db.commit()
    summary = {a: sum(1 for r in results if r.action_required == a)
               for a in ('DEPOSIT', 'WITHDRAWAL', 'NONE')}
    _audit(db, current_user, "sbi_run_p04", {"recon_date": recon_date, **summary})
    return {"recon_date": recon_date, "total": len(results), "summary": summary}


# ── P04: Mark action done ─────────────────────────────────────────────────────

@router.patch("/p04/{result_id}/done")
def mark_p04_done(
    result_id: str,
    done: bool = True,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("upload")),
):
    """Mark a P04 wallet adjustment as completed (after team acts in SBI portal)."""
    r = db.query(SBIP04Result).filter(SBIP04Result.id == result_id).first()
    if not r: raise HTTPException(status_code=404, detail="Result not found")
    r.action_done = done
    db.commit()
    return {"message": "Updated"}


# ── Query Endpoints ───────────────────────────────────────────────────────────

@router.get("/p01/results")
def get_p01_results(
    recon_date: Optional[str] = None,
    status: Optional[str] = None,
    ko_id: Optional[str] = None,
    txn_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = db.query(SBIP01Result)
    if recon_date: q = q.filter(SBIP01Result.recon_date == recon_date)
    if status: q = q.filter(SBIP01Result.status == status)
    if ko_id:  q = q.filter(SBIP01Result.ko_id.like(f"%{ko_id}%"))
    if txn_date:  # match the bank statement date OR the wallet deduction date (D±1)
        q = q.filter((SBIP01Result.bank_txn_date == txn_date) | (SBIP01Result.deduct_date == txn_date))
    rows = q.order_by(SBIP01Result.ko_id).all()
    grand = {"wallet_withdrawn": 0.0, "bank_settled": 0.0, "difference": 0.0}
    data = []
    for r in rows:
        grand["wallet_withdrawn"] += r.wallet_withdrawn or 0
        grand["bank_settled"]     += r.bank_settled or 0
        grand["difference"]       += r.difference or 0
        data.append({k: getattr(r, k) for k in ('id','ko_id','wallet_withdrawn','bank_settled','difference','status','deduct_date','bank_txn_date','notes','recon_date')})
    data = _apply_src_assignments(db, "p01", recon_date, data)
    return {"rows": data, "grand": {k: round(v, 2) for k, v in grand.items()}, "count": len(data)}


@router.get("/p01/lines")
def get_p01_lines(
    recon_date: str,
    upload_date: Optional[str] = None,
    ko_id: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    P01 line-level (bank <-> data) drill — the P02-style view for P01: the individual
    EKOSETTLEMENT bank lines and KO-withdrawal lines that make up each KO's P01 totals.
    Source rows are keyed by the transaction's BUSINESS date (recon_date == txn_date),
    matching the runs.
    """
    p01q = db.query(SBIP01Result).filter(SBIP01Result.recon_date == recon_date)
    if ko_id:  p01q = p01q.filter(SBIP01Result.ko_id.like(f"%{ko_id}%"))
    if status: p01q = p01q.filter(SBIP01Result.status == status)
    p01 = {r.ko_id: r for r in p01q.all()}
    wanted = set(p01.keys())

    groups = {}
    def _g(ko):
        if ko not in groups:
            r = p01.get(ko)
            groups[ko] = {
                "ko_id": ko,
                "status": r.status if r else None,
                "wallet_withdrawn": r.wallet_withdrawn if r else 0.0,
                "bank_settled": r.bank_settled if r else 0.0,
                "difference": r.difference if r else 0.0,
                "deduct_date": r.deduct_date if r else "",
                "bank_txn_date": r.bank_txn_date if r else "",
                "withdrawals": [], "settlements": [],
            }
        return groups[ko]

    for w in db.query(SBIKOLimits).filter(
            SBIKOLimits.txn_type == "KO Withdrawal", SBIKOLimits.txn_date == recon_date).all():
        if wanted and w.ko_id not in wanted:
            continue
        _g(w.ko_id)["withdrawals"].append({
            "amount": w.amount, "txn_date": w.txn_date, "datetime": w.txn_datetime,
            "configured_by": w.limit_configured_by})
    from sqlalchemy import or_ as _or, and_ as _and
    # Settlements belong to the BUSINESS date via their deduct_date (from the description),
    # not the bank posting date — same key run_p01 matches on, so D+1 settlements show here.
    for s in db.query(SBIBankTransaction).filter(
            SBIBankTransaction.is_settlement == True,
            _or(SBIBankTransaction.deduct_date == recon_date,
                _and(_or(SBIBankTransaction.deduct_date.is_(None), SBIBankTransaction.deduct_date == ''),
                     SBIBankTransaction.txn_date == recon_date))).all():
        if not s.ko_id or (wanted and s.ko_id not in wanted):
            continue
        _g(s.ko_id)["settlements"].append({
            "amount": s.debit, "txn_date": s.txn_date, "deduct_date": s.deduct_date,
            "description": s.description, "ref": s.ref_number})

    for ko in wanted:               # KOs with a P01 row but no source lines still show
        _g(ko)

    rows = sorted(groups.values(), key=lambda x: x["ko_id"] or "")
    rows = _apply_src_assignments(db, "p01", recon_date, rows)   # KO-level SRC overlay
    return {"recon_date": recon_date, "upload_date": upload_date, "kos": rows, "count": len(rows)}


@router.get("/p02/results")
def get_p02_results(
    recon_date: Optional[str] = None,
    match_status: Optional[str] = None,
    ko_id: Optional[str] = None,
    reference: Optional[str] = None,
    bank_type: Optional[str] = None,
    page: int = 1,
    page_size: int = 100,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    # Filters that narrow which rows are shown (applied to the row list AND the
    # summary so counts match the filtered view). match_status is the summary's
    # grouping key, so it's NOT applied to the summary.
    def _common(query):
        if recon_date: query = query.filter(SBIP02Result.recon_date == recon_date)
        if ko_id:      query = query.filter(SBIP02Result.ko_id.like(f"%{ko_id}%"))
        if reference:  query = query.filter(SBIP02Result.reference_number.like(f"%{reference}%"))
        if bank_type:  query = query.filter(SBIP02Result.bank_type == bank_type)
        return query

    q = _common(db.query(SBIP02Result))
    if match_status: q = q.filter(SBIP02Result.match_status == match_status)
    total = q.count()
    # Summary from the filtered set (excluding the match_status filter)
    from sqlalchemy import func as sqlfunc
    summary_q = _common(db.query(SBIP02Result.match_status, sqlfunc.count(SBIP02Result.id)))
    summary = {s: c for s, c in summary_q.group_by(SBIP02Result.match_status).all()}
    rows = q.order_by(SBIP02Result.reference_number).offset((page-1)*page_size).limit(page_size).all()
    # Bank-statement narration (read-only) for this page's rows, via the
    # SBIP02Result → SBIBankTransaction FK. One batched lookup, no N+1.
    # Read-only enrichment for the paired view (purely additive — no match-logic change):
    # the bank line's date + narration (via bank_txn_id) and the matched transaction
    # report's date/journal/status/ref (via txn_report_id). Two batched lookups, no N+1.
    _bt_ids = [r.bank_txn_id for r in rows if r.bank_txn_id]
    _bank_map = {}
    if _bt_ids:
        for _bid, _bdesc, _bdate in db.query(
            SBIBankTransaction.id, SBIBankTransaction.description, SBIBankTransaction.txn_date
        ).filter(SBIBankTransaction.id.in_(_bt_ids)).all():
            _bank_map[_bid] = (_bdesc, _bdate)
    _tr_ids = [r.txn_report_id for r in rows if r.txn_report_id]
    _rep_map = {}
    if _tr_ids:
        for _rid, _rdate, _rjrnl, _rstat, _rref in db.query(
            SBITxnReport.id, SBITxnReport.txn_date, SBITxnReport.journal_number,
            SBITxnReport.status, SBITxnReport.reference_number
        ).filter(SBITxnReport.id.in_(_tr_ids)).all():
            _rep_map[_rid] = (_rdate, _rjrnl, _rstat, _rref)
    data = []
    for r in rows:
        _b = _bank_map.get(r.bank_txn_id) or ('', '')
        _rp = _rep_map.get(r.txn_report_id) or ('', '', '', '')
        data.append({
            **{k: getattr(r, k) for k in ('id','reference_number','ko_id','bank_amount','bank_type','report_amount','report_txn_type','match_status','reversal_type','success_status','notes','recon_date')},
            'bank_description': _b[0] or '',
            'bank_txn_date':   _b[1] or '',
            'report_date':     _rp[0] or '',
            'report_journal':  _rp[1] or '',
            'report_status':   _rp[2] or '',
            'report_ref':      _rp[3] or '',
        })
    data = _apply_manual_matches(db, "p02", recon_date, data)
    data = _apply_src_assignments(db, "p02", recon_date, data)
    _mmq = db.query(SBIManualMatch).filter(SBIManualMatch.process == "p02")
    if recon_date: _mmq = _mmq.filter(SBIManualMatch.recon_date == recon_date)
    return {"rows": data, "summary": summary, "manual_matched_count": _mmq.count(),
            "total": total, "page": page, "page_size": page_size}


@router.get("/p03/results")
def get_p03_results(
    recon_date: Optional[str] = None,
    match_status: Optional[str] = None,
    csp_code: Optional[str] = None,
    mode: Optional[str] = None,
    txn_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 100,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    def _common(query):
        if recon_date: query = query.filter(SBIP03Result.recon_date == recon_date)
        if csp_code:   query = query.filter(SBIP03Result.csp_code.like(f"%{csp_code}%"))
        if mode:       query = query.filter(SBIP03Result.mode == mode)
        if txn_date:   # the transaction date OR the bank credit date
            query = query.filter((SBIP03Result.txn_date == txn_date) | (SBIP03Result.bank_credit_date == txn_date))
        return query

    q = _common(db.query(SBIP03Result))
    if match_status: q = q.filter(SBIP03Result.match_status == match_status)
    total = q.count()
    from sqlalchemy import func as sqlfunc
    summary_q = _common(db.query(SBIP03Result.match_status, sqlfunc.count(SBIP03Result.id)))
    summary = {s: c for s, c in summary_q.group_by(SBIP03Result.match_status).all()}
    rows = q.order_by(SBIP03Result.csp_code).offset((page-1)*page_size).limit(page_size).all()
    data = [{k: getattr(r, k) for k in ('id','csp_code','mode','ref_number','txn_amount','txn_date','bank_credit_date','bank_amount','match_status','date_shift','match_priority','notes','recon_date')} for r in rows]
    data = _apply_manual_matches(db, "p03", recon_date, data)
    data = _apply_src_assignments(db, "p03", recon_date, data)
    _mmq = db.query(SBIManualMatch).filter(SBIManualMatch.process == "p03")
    if recon_date: _mmq = _mmq.filter(SBIManualMatch.recon_date == recon_date)
    return {"rows": data, "summary": summary, "manual_matched_count": _mmq.count(),
            "total": total, "page": page, "page_size": page_size}


@router.get("/p04/results")
def get_p04_results(
    recon_date: Optional[str] = None,
    action_required: Optional[str] = None,
    csp_code: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = db.query(SBIP04Result)
    if recon_date:      q = q.filter(SBIP04Result.recon_date == recon_date)
    if action_required: q = q.filter(SBIP04Result.action_required == action_required)
    if csp_code:        q = q.filter(SBIP04Result.csp_code.like(f"%{csp_code}%"))
    rows = q.order_by(SBIP04Result.csp_code).all()
    summary = {}
    for r in rows:
        summary[r.action_required] = summary.get(r.action_required, 0) + 1
    data = [{k: getattr(r, k) for k in ('id','csp_code','failed_amount','closing_balance','expected_balance','difference','action_required','action_amount','action_done','notes','recon_date')} for r in rows]
    data = _apply_src_assignments(db, "p04", recon_date, data)
    return {"rows": data, "summary": summary, "count": len(data)}


# ── Unified ledger — every bank & data entry + how each reconciled (P01/P02/P03) ──

_P01_UNIFIED_STATUS = {"matched": "Matched", "unmatched": "Unmatched",
                       # legacy statuses (pre-2026-07-27) fold into the two-state model
                       "CREDITED": "Matched", "PENDING": "Unmatched",
                       "PARTIAL": "Unmatched", "EXCESS": "Unmatched"}

def _r(n) -> str:
    try:
        return f"₹{float(n or 0):,.0f}"
    except (TypeError, ValueError):
        return "₹0"


# Rows whose unified status counts as reconciled (excluded from the pair-picker's open lists).
_MATCHED_UNIFIED = {"Matched", "Manual_Matched"}


def _row_disc(source, row):
    """A per-physical-row discriminator that is STABLE across recon re-runs and file
    re-uploads (source CONTENT, never the regenerated row id). Folded into _pair_key so two
    business-identical rows (e.g. two same-amount KO withdrawals for one KO on one day) get
    DISTINCT keys and a pair can't fan out across them. Bank rows use the running balance
    (unique per statement line); txn reports use datetime+journal; KO rows use the datetime."""
    if source in ("Bank Settlement", "Bank Statement"):
        bal = getattr(row, "balance", None)
        return "" if bal is None else f"{bal}"
    if source == "Txn Report":
        return f"{getattr(row, 'txn_datetime', '') or ''}|{getattr(row, 'journal_number', '') or ''}"
    if source in ("KO Withdrawal", "KO Deposit"):
        return getattr(row, "txn_datetime", "") or ""
    return ""


def _pair_key(side, source, ko, ref, date, drcr, amount, disc=""):
    """Stable business key for one side of a manual pair. Computed identically from a
    source row (write time) and from a unified entry (read time) so a pair survives recon
    re-runs AND file re-uploads (behavior-contract #17 — source ids regenerate on upload).
    `disc` (a stable per-row content discriminator) keeps business-identical rows distinct."""
    try:
        amt = f"{float(amount or 0):.2f}"
    except (TypeError, ValueError):
        amt = "0.00"
    tail = f"|d={disc or ''}"
    if side == "bank":
        tag = "bankset" if source == "Bank Settlement" else "bank"
        return f"{tag}|{ko or ''}|{ref or ''}|{date or ''}|{drcr or ''}|{amt}{tail}"
    tag = {"Txn Report": "txr", "KO Withdrawal": "kow", "KO Deposit": "kod"}.get(source, "dat")
    return f"{tag}|{ko or ''}|{ref or ''}|{date or ''}|{amt}{tail}"


def _apply_pairs(db, entries):
    """Overlay operator manual pairs onto unified entries (read-time). A row whose stable
    key is one side of a pair flips to Manual_Matched with the other side as its counterpart.
    Each stored pair is CONSUMED once per side, so even if two rows share a key a pair can
    never phantom-reconcile more than one row. No-op when no pairs exist, so every existing
    surface (the unified page, the all-entries report) stays byte-identical until a pair."""
    if not entries:
        return entries
    pairs = db.query(SBIManualPair).all()
    if pairs:
        by_bank = {p.bank_key: p for p in pairs if p.bank_key}
        by_data = {p.data_key: p for p in pairs if p.data_key}
        used_bank_ids, used_data_ids = set(), set()
        for e in entries:
            k = _pair_key(e["side"], e["source"], e["ko_csp"], e["ref"], e["date"],
                          e["drcr"], e["amount"], e.get("_disc"))
            if e["side"] == "bank":
                p = by_bank.get(k)
                if not p or p.id in used_bank_ids:
                    continue
                used_bank_ids.add(p.id)
                other_src, other_amt = p.data_source, p.data_amount
            else:
                p = by_data.get(k)
                if not p or p.id in used_data_ids:
                    continue
                used_data_ids.add(p.id)
                other_src, other_amt = p.bank_source, p.bank_amount
            e["status"] = "Manual_Matched"
            e["process"] = "Manual"
            e["manual_pair_id"] = p.id
            e["counterpart"] = f"{other_src or ''} {_r(other_amt)}".strip()
    for e in entries:
        e.pop("_disc", None)     # private discriminator — never serialised
    return entries


def _unified_entries(db, recon_date, include_deposits=False):
    """Build the flat list of every bank + data source entry for ONE business date, each
    tagged with recon status/process/counterpart and its manual-pair overlay. Returns
    entries carrying a private '_result' handle (caller must pop it). Shared by get_unified
    (the SBI Kiosk 'All Entries' view and the all-entries report) and the manual-match
    pair-picker. `include_deposits` adds KO Deposit rows — OFF by default so the unified
    page and the all-entries report stay byte-identical to before."""
    # P01 results indexed by (business_date, ko). A settlement posted on recon_date may
    # belong to an EARLIER business date (its deduct_date) where its P01 match now lives,
    # so load recon_date PLUS those deduct dates. Withdrawal side keys on recon_date;
    # settlement side keys on the row's deduct_date.
    _settle_dd = {d for (d,) in db.query(SBIBankTransaction.deduct_date).filter(
        SBIBankTransaction.txn_date == recon_date, SBIBankTransaction.is_settlement == True,
        SBIBankTransaction.deduct_date.isnot(None), SBIBankTransaction.deduct_date != '').distinct().all() if d}
    p01_idx = {}
    for x in db.query(SBIP01Result).filter(SBIP01Result.recon_date.in_({recon_date} | _settle_dd)).all():
        p01_idx[(x.recon_date, x.ko_id)] = x
    p01_by_ko = {k[1]: v for k, v in p01_idx.items() if k[0] == recon_date}
    p02_by_bank, p02_by_report = {}, {}
    for x in db.query(SBIP02Result).filter(SBIP02Result.recon_date == recon_date).all():
        if x.bank_txn_id:   p02_by_bank[x.bank_txn_id] = x
        if x.txn_report_id: p02_by_report[x.txn_report_id] = x
    # P03 has no source FK — link best-effort by CSP + amount. p03_matched (Matched only)
    # drives the "also in P03" overlay on bank credits; p03_by_txn (all statuses, money-out
    # side, keyed incl. ref) maps transaction-report lines to their P03 row for SRC/actions.
    p03_matched, p03_by_txn = {}, {}
    for x in db.query(SBIP03Result).filter(SBIP03Result.recon_date == recon_date).all():
        if not x.csp_code:
            continue
        if x.match_status == "Matched":
            a = x.txn_amount if x.txn_amount is not None else x.bank_amount
            if a is not None:
                p03_matched[(x.csp_code, round(float(a), 2))] = x
        if x.txn_amount is not None:
            p03_by_txn[(x.csp_code, round(float(x.txn_amount), 2), x.ref_number or "")] = x

    entries = []

    def _new(sidev, src, row_id, ref, ko, amt, date, drcr, narration, file="", disc=""):
        return {"side": sidev, "source": src, "id": row_id, "ref": ref or "", "ko_csp": ko or "",
                "amount": amt, "date": date or "", "drcr": drcr, "narration": narration or "",
                "file": file or src,   # fine-grained originating file/report (Txn Report → canonical product)
                "status": "Not reconciled", "process": "", "counterpart": None, "also_p03": False,
                "result_id": None, "result_process": None, "src_code": None, "src_note": None,
                "manual_pair_id": None, "_disc": disc, "_result": None}

    # ---- bank side ----
    for b in db.query(SBIBankTransaction).filter(SBIBankTransaction.txn_date == recon_date).all():
        drcr = "DR" if (b.debit or 0) > 0 else ("CR" if (b.credit or 0) > 0 else "")
        amt = (b.debit or 0) if (b.debit or 0) > 0 else (b.credit or 0)
        if b.is_settlement:
            e = _new("bank", "Bank Settlement", b.id, _clean_refno(b.ref_number), b.ko_id, amt, b.txn_date, drcr, b.description,
                     disc=_row_disc("Bank Settlement", b))
            # settlement reconciles under its business (deduct) date, not the posting date
            r = p01_idx.get(((b.deduct_date or recon_date), b.ko_id))
            if r:
                e.update(status=_P01_UNIFIED_STATUS.get(r.status, r.status), process="P01",
                         counterpart=f"Wallet out {_r(r.wallet_withdrawn)}", result_id=r.id,
                         result_process="p01", _result=r)
        else:
            e = _new("bank", "Bank Statement", b.id, _clean_refno(b.ref_number), b.ko_id, amt, b.txn_date, drcr, b.description,
                     disc=_row_disc("Bank Statement", b))
            r = p02_by_bank.get(b.id)
            if r:
                e.update(status=r.match_status, process="P02", result_id=r.id, result_process="p02", _result=r,
                         counterpart=(f"Report {_r(r.report_amount)} · {r.report_txn_type}" if r.report_amount is not None else None))
            if drcr == "CR" and (b.ko_id, round(float(amt), 2)) in p03_matched:
                e["also_p03"] = True
                if e["status"] in ("Unmatched", "Not reconciled"):
                    pr = p03_matched[(b.ko_id, round(float(amt), 2))]
                    e.update(status="Matched", process="P03", result_id=pr.id, result_process="p03", _result=pr,
                             counterpart=f"Txn out {_r(pr.txn_amount)}")
        entries.append(e)

    # ---- data side: transaction reports ----
    for t in db.query(SBITxnReport).filter(SBITxnReport.txn_date == recon_date).all():
        amt = t.amount or 0
        e = _new("data", "Txn Report", t.id, t.reference_number, t.ko_id, amt, t.txn_date, "", t.txn_type,
                 file=canonical_product(t.source_file), disc=_row_disc("Txn Report", t))
        e["status"] = "Unmatched"
        r = p02_by_report.get(t.id)
        if r:
            e.update(status=r.match_status, process="P02", result_id=r.id, result_process="p02", _result=r,
                     counterpart=f"Bank {r.bank_type} {_r(r.bank_amount)}")
        else:
            pr = p03_by_txn.get((t.ko_id, round(float(amt), 2), t.reference_number or ""))
            if pr:
                st = {"Unmatched_TxnReport": "Unmatched", "Unmatched_Bank": "Unmatched"}.get(pr.match_status, pr.match_status)
                e.update(status=st, process="P03", also_p03=(pr.match_status == "Matched"),
                         result_id=pr.id, result_process="p03", _result=pr,
                         counterpart=(f"Bank credit {_r(pr.bank_amount)}" if pr.bank_amount is not None else None))
        entries.append(e)

    # ---- data side: KO withdrawals ----
    for w in db.query(SBIKOLimits).filter(SBIKOLimits.txn_type == "KO Withdrawal",
                                          SBIKOLimits.txn_date == recon_date).all():
        e = _new("data", "KO Withdrawal", w.id, "", w.ko_id, w.amount or 0, w.txn_date, "", "KO Withdrawal",
                 disc=_row_disc("KO Withdrawal", w))
        r = p01_by_ko.get(w.ko_id)
        if r:
            e.update(status=_P01_UNIFIED_STATUS.get(r.status, r.status), process="P01",
                     counterpart=f"Bank settled {_r(r.bank_settled)}", result_id=r.id,
                     result_process="p01", _result=r)
        entries.append(e)

    # ---- data side: KO deposits (pair-picker only; OFF for the unified page + report) ----
    if include_deposits:
        for w in db.query(SBIKOLimits).filter(SBIKOLimits.txn_type == "KO Deposit",
                                              SBIKOLimits.txn_date == recon_date).all():
            entries.append(_new("data", "KO Deposit", w.id, "", w.ko_id, w.amount or 0, w.txn_date, "", "KO Deposit",
                                disc=_row_disc("KO Deposit", w)))

    return _apply_pairs(db, entries)


@router.get("/unified")
def get_unified(
    recon_date: str,
    upload_date: Optional[str] = None,
    side: Optional[str] = None,        # bank | data
    status: Optional[str] = None,      # Matched | Unmatched | Reversal | ...
    process: Optional[str] = None,     # p01 | p02 | p03
    search: Optional[str] = None,      # ref / KO / CSP
    page: int = 1,
    page_size: int = 100,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Unified SBI ledger: EVERY source entry — bank statement lines AND data lines
    (transaction reports, KO withdrawals) — each tagged with its reconciliation
    status, which process (P01/P02/P03) reconciled it, and its counterpart. Built
    read-only from the P0x result tables; SRC + manual-match act on the mapped
    result row via the existing endpoints (result_id + result_process on each row).
    Scoped to one date (sources by upload_date≈recon_date; behaviour-contract #17).
    P03 has no source FK, so P03 links are best-effort on CSP+amount.
    Sources by the transaction's BUSINESS date (recon_date == txn_date), matching the runs.
    """
    entries = _unified_entries(db, recon_date, include_deposits=False)

    # ---- filters ----
    if side:    entries = [e for e in entries if e["side"] == side]
    if process: entries = [e for e in entries if (e["process"] or "").lower() == process.lower()]
    if status:  entries = [e for e in entries if (e["status"] or "").lower() == status.lower()]
    if search:
        s = search.lower()
        entries = [e for e in entries if s in (e["ref"] or "").lower() or s in (e["ko_csp"] or "").lower()]

    from collections import Counter
    status_counts = dict(Counter(e["status"] for e in entries))
    side_counts   = dict(Counter(e["side"] for e in entries))
    total = len(entries)

    entries.sort(key=lambda e: (e["ko_csp"] or "", e["ref"] or "", e["side"]))
    page_rows = entries[(page - 1) * page_size: page * page_size]

    # SRC-tag overlay for the page rows, from the mapped result's stable key
    sa = {(a.process, a.match_key): a for a in db.query(SBISrcAssignment).filter(
          SBISrcAssignment.recon_date == recon_date).all()}
    for e in page_rows:
        r = e.pop("_result", None)
        if r is not None and e["result_process"]:
            rd = {c.name: getattr(r, c.name) for c in r.__table__.columns}
            key = _src_key(e["result_process"], rd)
            a = sa.get((e["result_process"], key)) if key else None
            if a:
                e["src_code"], e["src_note"] = a.src_code, a.src_note
    for e in entries:                # drop the non-serialisable ref on any non-page rows
        e.pop("_result", None)

    return {"rows": page_rows, "total": total, "page": page, "page_size": page_size,
            "status_counts": status_counts, "side_counts": side_counts,
            "recon_date": recon_date, "upload_date": upload_date}


# ── Export ─────────────────────────────────────────────────────────────────────
# Each SBI process reconciles in its OWN result table (NOT the core `transactions`
# ledger), so the generic /api/reports exports return blank for partner=kiosk. These
# export the real recon results, with the operator-facing names the team uses.
#   p01 = "Bank vs Settlement"   p02 = "Bank vs Transaction"
#   p03 = "CSP-Txn-Bank"         p04 = "Wallet Balance"
_SBI_EXPORTS = {
    "p01": ("Bank vs Settlement", SBIP01Result, "ko_id",
            ['recon_date', 'ko_id', 'wallet_withdrawn', 'bank_settled', 'difference',
             'status', 'deduct_date', 'bank_txn_date', 'notes']),
    "p02": ("Bank vs Transaction", SBIP02Result, "reference_number",
            ['recon_date', 'reference_number', 'ko_id', 'bank_amount', 'bank_type',
             'report_amount', 'report_txn_type', 'match_status', 'reversal_type',
             'success_status', 'notes']),
    "p03": ("CSP-Txn-Bank", SBIP03Result, "csp_code",
            ['recon_date', 'csp_code', 'mode', 'ref_number', 'txn_amount', 'txn_date',
             'bank_credit_date', 'bank_amount', 'match_status', 'date_shift',
             'match_priority', 'notes']),
    "p04": ("Wallet Balance", SBIP04Result, "csp_code",
            ['recon_date', 'csp_code', 'failed_amount', 'closing_balance',
             'expected_balance', 'difference', 'action_required', 'action_amount',
             'action_done', 'notes']),
}


def _build_sbi_export(db, which, recon_date=None, match_status=None,
                      date_from=None, date_to=None) -> io.BytesIO:
    """Build a styled multi-sheet SBI recon workbook; one sheet per process in `which`.
    Always writes a header row (a valid file even with zero data rows).
    recon_date = single date (unchanged behaviour); date_from/date_to = range."""
    from openpyxl.styles import PatternFill, Font as XFont
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for p in which:
            sheet, model, order_col, cols = _SBI_EXPORTS[p]
            q = db.query(model)
            if recon_date:
                q = q.filter(model.recon_date == recon_date)
            else:
                if date_from: q = q.filter(model.recon_date >= date_from)
                if date_to:   q = q.filter(model.recon_date <= date_to)
            # match_status only applies to the processes that have that column (p02/p03)
            if match_status and hasattr(model, "match_status"):
                q = q.filter(model.match_status == match_status)
            rows = q.order_by(getattr(model, order_col)).all()
            recs = [{c: getattr(r, c) for c in cols} for r in rows]
            if p in ("p02", "p03"):
                recs = _apply_manual_matches(db, p, recon_date, recs)
            df = pd.DataFrame(recs, columns=cols)
            df.to_excel(writer, sheet_name=sheet[:31], index=False)
            ws = writer.sheets[sheet[:31]]
            fill = PatternFill("solid", fgColor="094053")
            for cell in ws[1]:
                cell.fill = fill
                cell.font = XFont(color="FFFFFF", bold=True)
            for i, c in enumerate(cols, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, min(42, len(str(c)) + 4))
    output.seek(0)
    return output


@router.get("/export")
def export_sbi(
    process: str = Query("all", description="p01 | p02 | p03 | p04 | all"),
    recon_date: Optional[str] = None,
    match_status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Export SBI Kiosk recon results to a styled Excel workbook.

    `process=all` → one workbook with one sheet per process (Bank vs Settlement,
    Bank vs Transaction, CSP-Txn-Bank, Wallet Balance). A specific process → a
    single-sheet file. Always returns a valid file (header row even when empty) —
    this replaces the old core-ledger export that came back blank because SBI
    reconciles in its own tables.
    """
    process = (process or "all").lower()
    which = list(_SBI_EXPORTS) if process == "all" else [process]
    if any(p not in _SBI_EXPORTS for p in which):
        raise HTTPException(status_code=400, detail="process must be one of p01, p02, p03, p04, all")
    # Never hand back a silently-blank file: if a recon_date was requested but has no
    # rows for any requested process (wrong date, or recon not run that day), fall back
    # to the latest date that DOES have data. `used_date` is echoed in a header so the
    # UI can tell the operator which date was actually exported (or that none exists).
    used_date = recon_date
    if recon_date:
        has_here = any(
            db.query(_SBI_EXPORTS[p][1].id).filter(_SBI_EXPORTS[p][1].recon_date == recon_date).first()
            for p in which
        )
        if not has_here:
            latest = None
            for p in which:
                m = _SBI_EXPORTS[p][1]
                d = db.query(func.max(m.recon_date)).filter(m.recon_date.isnot(None)).scalar()
                if d and (latest is None or d > latest):
                    latest = d
            used_date = latest  # None only when the process has no data on any date
    output = _build_sbi_export(db, which, used_date, match_status, date_from, date_to)
    _tag = used_date or (f"{date_from or 'start'}_to_{date_to or 'end'}"
                         if (date_from or date_to) else "all")
    fname = f"sbi_{process}_{_tag}.xlsx"
    # Header semantics the UI relies on: the actual date exported, or the range/"all"
    # tag when exporting a span WITH data, or "none" only when there is truly no data.
    if used_date:
        _hdr = used_date
    else:
        _has = any(_rng(db.query(_SBI_EXPORTS[p][1].id), _SBI_EXPORTS[p][1],
                        None, date_from, date_to).first() for p in which)
        _hdr = _tag if _has else "none"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Recon-Date": str(_hdr),
            "Access-Control-Expose-Headers": "X-Recon-Date",
        },
    )


# ── Report library ─────────────────────────────────────────────────────────────
# The full set of SBI report options beyond the raw per-process exports. All are
# READ-ONLY over the result/source/overlay tables — no matching logic involved.
# Range filters compare the zero-padded recon_date strings lexicographically
# (app-wide convention). Single-date reports fall back to the latest date with
# data and echo it in X-Recon-Date (same pattern as /export).

def _style_report_ws(ws, ncols: int):
    from openpyxl.styles import PatternFill, Font as XFont
    fill = PatternFill("solid", fgColor="094053")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = XFont(color="FFFFFF", bold=True)
    for i in range(1, ncols + 1):
        c = ws.cell(row=1, column=i)
        ws.column_dimensions[c.column_letter].width = max(12, min(46, len(str(c.value or "")) + 4))


def _sheet(writer, name: str, rows: list, cols: list):
    """Write one styled sheet; always valid (header row even with zero rows)."""
    df = pd.DataFrame(rows, columns=cols)
    df.to_excel(writer, sheet_name=name[:31], index=False)
    _style_report_ws(writer.sheets[name[:31]], len(cols))


def _rng(q, model, recon_date, date_from, date_to):
    if recon_date:
        return q.filter(model.recon_date == recon_date)
    if date_from:
        q = q.filter(model.recon_date >= date_from)
    if date_to:
        q = q.filter(model.recon_date <= date_to)
    return q


def _latest_sbi_date(db, models=None) -> Optional[str]:
    latest = None
    for m in (models or [SBIP01Result, SBIP02Result, SBIP03Result, SBIP04Result]):
        d = db.query(func.max(m.recon_date)).scalar()
        if d and (latest is None or d > latest):
            latest = d
    return latest


def _bank_desc_map(db, ids: list) -> dict:
    """id → (description, txn_date) for bank rows, chunked (IN-list safe on MySQL)."""
    out = {}
    ids = [i for i in ids if i]
    for i in range(0, len(ids), 900):
        for _id, _d, _dt in db.query(SBIBankTransaction.id, SBIBankTransaction.description,
                                     SBIBankTransaction.txn_date
                                     ).filter(SBIBankTransaction.id.in_(ids[i:i + 900])).all():
            out[_id] = (_d or '', _dt or '')
    return out


def _load_process_recs(db, p: str, recon_date, date_from=None, date_to=None,
                       with_bank_desc: bool = False) -> list:
    """Load one process's result rows as dicts (export columns) with the manual-match
    and SRC overlays applied — the same read-time view the results endpoints serve."""
    sheet, model, order_col, cols = _SBI_EXPORTS[p]
    q = _rng(db.query(model), model, recon_date, date_from, date_to)
    rows = q.order_by(getattr(model, order_col)).all()
    recs = [{c: getattr(r, c) for c in cols} for r in rows]
    if p == "p02" and with_bank_desc and rows:
        bm = _bank_desc_map(db, [r.bank_txn_id for r in rows])
        for rec, r in zip(recs, rows):
            d = bm.get(r.bank_txn_id) or ('', '')
            rec["bank_description"], rec["bank_txn_date"] = d[0], d[1]
    if p in ("p02", "p03"):
        recs = _apply_manual_matches(db, p, recon_date, recs)
    recs = _apply_src_assignments(db, p, recon_date, recs)
    return recs


def _proc_overview(p: str, recs: list) -> dict:
    """Per-process roll-up row for the Overview sheet."""
    total = len(recs)
    if p == "p01":
        matched = sum(1 for r in recs if r.get("status") == "matched")
        amt_all = sum(r.get("wallet_withdrawn") or 0 for r in recs)
        amt_ok = sum(r.get("wallet_withdrawn") or 0 for r in recs if r.get("status") == "matched")
        note = "matched = settlement found (by amount + description date); amounts = wallet withdrawals"
    elif p == "p04":
        matched = sum(1 for r in recs if r.get("action_required") == "NONE")
        amt_all = sum(abs(r.get("action_amount") or 0) for r in recs)
        amt_ok = 0.0
        note = "matched = no action needed; amount = pending corrections"
    else:
        matched = sum(1 for r in recs if r.get("match_status") in ("Matched", "Manual_Matched"))
        amt_field = "bank_amount" if p == "p02" else "txn_amount"
        amt_all = sum((r.get(amt_field) or r.get("bank_amount") or 0) for r in recs)
        amt_ok = sum((r.get(amt_field) or r.get("bank_amount") or 0) for r in recs
                     if r.get("match_status") in ("Matched", "Manual_Matched"))
        note = "matched incl. manual matches"
    return {"process": p.upper(), "sheet": _SBI_EXPORTS[p][0], "total": total,
            "matched": matched, "open": total - matched,
            "match_rate_pct": round(matched / total * 100, 1) if total else None,
            "amount_total": round(amt_all, 2), "amount_matched": round(amt_ok, 2),
            "amount_open": round(amt_all - amt_ok, 2), "notes": note}


_P02_XCOLS = ['recon_date', 'reference_number', 'ko_id', 'bank_amount', 'bank_type',
              'report_amount', 'report_txn_type', 'match_status', 'reversal_type',
              'success_status', 'bank_txn_date', 'bank_description', 'src_code', 'src_note', 'notes']
_P03_XCOLS = ['recon_date', 'csp_code', 'mode', 'ref_number', 'txn_amount', 'txn_date',
              'bank_credit_date', 'bank_amount', 'match_status', 'date_shift',
              'match_priority', 'src_code', 'src_note', 'notes']
_P01_XCOLS = ['recon_date', 'ko_id', 'wallet_withdrawn', 'bank_settled', 'difference',
              'status', 'deduct_date', 'bank_txn_date', 'src_code', 'src_note', 'notes']
_P04_XCOLS = ['recon_date', 'csp_code', 'failed_amount', 'closing_balance', 'expected_balance',
              'difference', 'action_required', 'action_amount', 'action_done',
              'src_code', 'src_note', 'notes']

_SBI_REPORTS = [
    {"id": "daily_pack",  "label": "Daily Recon Pack",          "scope": "date",
     "desc": "One workbook for the day: overview, all four processes, exceptions, reversals, SRC & manual logs."},
    {"id": "exceptions",  "label": "Exceptions / Work List",    "scope": "range",
     "desc": "Only what needs action: P01 not credited, P02/P03 unmatched or partial, P04 pending wallet actions."},
    {"id": "summary",     "label": "MIS Summary (trend)",       "scope": "range",
     "desc": "Per date × process: totals, matched, open, match rate and amounts — management trend view."},
    {"id": "ko_wise",     "label": "KO / CSP-wise Summary",     "scope": "range",
     "desc": "Per agent roll-up across P01–P03: volumes, matched vs open counts and amounts."},
    {"id": "unified",     "label": "All Entries (unified ledger)", "scope": "date",
     "desc": "Every bank and data entry with its status, which process reconciled it, and its counterpart."},
    {"id": "p01_lines",   "label": "P01 Settlement Lines",      "scope": "date",
     "desc": "Line-level P01: each KO withdrawal and each bank settlement line behind every KO total."},
    {"id": "reversals",   "label": "Reversals Report",          "scope": "range",
     "desc": "All P02 reversal legs (same reference DR + CR) with success/fail state."},
    {"id": "p04_actions", "label": "Wallet Action Tracker (P04)", "scope": "range",
     "desc": "Deposit/withdrawal corrections with done vs pending status."},
    {"id": "src_log",     "label": "SRC Disposition Log",       "scope": "range",
     "desc": "Every SRC tag on SBI rows: code, note, who and when."},
    {"id": "manual_log",  "label": "Manual Match Log",          "scope": "range",
     "desc": "Every SBI manual match: key, counterpart, remark, who and when."},
]


@router.get("/report-options")
def sbi_report_options(current_user=Depends(get_current_user)):
    """The report library (drives the Reports UI — additions are backend-only)."""
    return {"reports": _SBI_REPORTS}


def build_sbi_report_xlsx(db, type, recon_date=None, date_from=None, date_to=None, current_user=None):
    """Build one SBI report-library workbook. SHARED by the /sbi/report route AND the
    scheduled sbi_* reports so both produce the same file. Returns (bytes, filename, tag).
    current_user is only threaded to get_unified/get_p01_lines, which use it purely as an
    auth dependency (never in their body), so the scheduler can pass None."""
    t = (type or "").lower()
    if t not in {r["id"] for r in _SBI_REPORTS}:
        raise HTTPException(status_code=400, detail="Unknown report type — see /sbi/report-options")

    # Single-date reports: fall back to the latest date with data (echoed in header).
    used_date = recon_date
    scope = next(r["scope"] for r in _SBI_REPORTS if r["id"] == t)
    if scope == "date":
        if used_date:
            has = any(db.query(m.id).filter(m.recon_date == used_date).first()
                      for m in (SBIP01Result, SBIP02Result, SBIP03Result, SBIP04Result))
            if not has:
                used_date = _latest_sbi_date(db)
        else:
            used_date = _latest_sbi_date(db)
        rd, df_, dt_ = used_date, None, None
        tag = used_date or "none"
    else:
        rd, df_, dt_ = recon_date, date_from, date_to
        tag = rd or (f"{df_ or 'start'}_to_{dt_ or 'end'}" if (df_ or dt_) else "all")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as w:

        if t == "daily_pack":
            recs = {p: _load_process_recs(db, p, rd, with_bank_desc=(p == "p02")) for p in _SBI_EXPORTS}
            _sheet(w, "Overview", [_proc_overview(p, recs[p]) for p in _SBI_EXPORTS],
                   ["process", "sheet", "total", "matched", "open", "match_rate_pct",
                    "amount_total", "amount_matched", "amount_open", "notes"])
            _sheet(w, "P01 Settlement", recs["p01"], _P01_XCOLS)
            _sheet(w, "P02 Bank vs Txn", recs["p02"], _P02_XCOLS)
            _sheet(w, "P03 CSP-Txn-Bank", recs["p03"], _P03_XCOLS)
            _sheet(w, "P04 Wallet", recs["p04"], _P04_XCOLS)
            _sheet(w, "Exceptions P01",
                   [r for r in recs["p01"] if r.get("status") != "matched"], _P01_XCOLS)
            _sheet(w, "Exceptions P02",
                   [r for r in recs["p02"] if r.get("match_status") not in ("Matched", "Manual_Matched")], _P02_XCOLS)
            _sheet(w, "Exceptions P03",
                   [r for r in recs["p03"] if r.get("match_status") not in ("Matched", "Manual_Matched")], _P03_XCOLS)
            _sheet(w, "P04 Pending",
                   [r for r in recs["p04"] if r.get("action_required") != "NONE" and not r.get("action_done")], _P04_XCOLS)
            _sheet(w, "Reversals",
                   [r for r in recs["p02"] if r.get("reversal_type") in ("Reversal Debit", "Reversal Credit")], _P02_XCOLS)
            srcs = db.query(SBISrcAssignment).filter(SBISrcAssignment.recon_date == rd).all() if rd else []
            _sheet(w, "SRC Log",
                   [{"recon_date": s.recon_date, "process": s.process, "match_key": s.match_key,
                     "src_code": s.src_code, "src_note": s.src_note, "by": s.created_by,
                     "at": str(s.created_at or "")} for s in srcs],
                   ["recon_date", "process", "match_key", "src_code", "src_note", "by", "at"])
            mms = db.query(SBIManualMatch).filter(SBIManualMatch.recon_date == rd).all() if rd else []
            _sheet(w, "Manual Log",
                   [{"recon_date": m.recon_date, "process": m.process, "match_key": m.match_key,
                     "counterpart_ref": m.counterpart_ref, "remark": m.remark, "by": m.created_by,
                     "at": str(m.created_at or "")} for m in mms],
                   ["recon_date", "process", "match_key", "counterpart_ref", "remark", "by", "at"])

        elif t == "exceptions":
            p01 = [r for r in _load_process_recs(db, "p01", rd, df_, dt_) if r.get("status") != "matched"]
            p02 = [r for r in _load_process_recs(db, "p02", rd, df_, dt_, with_bank_desc=True)
                   if r.get("match_status") not in ("Matched", "Manual_Matched")]
            p03 = [r for r in _load_process_recs(db, "p03", rd, df_, dt_)
                   if r.get("match_status") not in ("Matched", "Manual_Matched")]
            p04 = [r for r in _load_process_recs(db, "p04", rd, df_, dt_)
                   if r.get("action_required") != "NONE" and not r.get("action_done")]
            _sheet(w, "P01 Not Credited", p01, _P01_XCOLS)
            _sheet(w, "P02 Open", p02, _P02_XCOLS)
            _sheet(w, "P03 Open", p03, _P03_XCOLS)
            _sheet(w, "P04 Pending Actions", p04, _P04_XCOLS)

        elif t == "summary":
            grid, totals = [], {}
            for p in _SBI_EXPORTS:
                recs = _load_process_recs(db, p, rd, df_, dt_)
                by_date = {}
                for r in recs:
                    by_date.setdefault(r.get("recon_date") or "", []).append(r)
                for d in sorted(by_date):
                    row = _proc_overview(p, by_date[d]); row["date"] = d
                    grid.append(row)
                totals[p] = _proc_overview(p, recs)
            _sheet(w, "By Date x Process", grid,
                   ["date", "process", "total", "matched", "open", "match_rate_pct",
                    "amount_total", "amount_matched", "amount_open"])
            _sheet(w, "Totals by Process", [totals[p] for p in _SBI_EXPORTS],
                   ["process", "sheet", "total", "matched", "open", "match_rate_pct",
                    "amount_total", "amount_matched", "amount_open", "notes"])

        elif t == "ko_wise":
            def _agg(recs, key, amt_field, ok_states):
                out = {}
                for r in recs:
                    k = r.get(key) or "(blank)"
                    a = out.setdefault(k, {"ko_csp": k, "total": 0, "matched": 0, "open": 0,
                                           "amount_total": 0.0, "amount_open": 0.0})
                    a["total"] += 1
                    ok = (r.get("match_status") or r.get("status")) in ok_states
                    a["matched"] += 1 if ok else 0
                    a["open"] += 0 if ok else 1
                    amt = r.get(amt_field) or r.get("bank_amount") or 0
                    a["amount_total"] = round(a["amount_total"] + amt, 2)
                    if not ok:
                        a["amount_open"] = round(a["amount_open"] + amt, 2)
                return out
            p01r = _load_process_recs(db, "p01", rd, df_, dt_)
            p02a = _agg(_load_process_recs(db, "p02", rd, df_, dt_), "ko_id", "bank_amount",
                        ("Matched", "Manual_Matched"))
            p03a = _agg(_load_process_recs(db, "p03", rd, df_, dt_), "csp_code", "txn_amount",
                        ("Matched", "Manual_Matched"))
            p01a = {}
            for r in p01r:
                k = r.get("ko_id") or "(blank)"
                a = p01a.setdefault(k, {"ko_csp": k, "wallet_withdrawn": 0.0, "bank_settled": 0.0,
                                        "difference": 0.0, "days": 0, "days_credited": 0})
                a["wallet_withdrawn"] = round(a["wallet_withdrawn"] + (r.get("wallet_withdrawn") or 0), 2)
                a["bank_settled"] = round(a["bank_settled"] + (r.get("bank_settled") or 0), 2)
                a["difference"] = round(a["difference"] + (r.get("difference") or 0), 2)
                a["days"] += 1
                a["days_credited"] += 1 if r.get("status") == "matched" else 0
            all_kos = sorted(set(p01a) | set(p02a) | set(p03a))
            overview = []
            for k in all_kos:
                o1, o2, o3 = p01a.get(k, {}), p02a.get(k, {}), p03a.get(k, {})
                overview.append({
                    "ko_csp": k,
                    "p01_wallet": o1.get("wallet_withdrawn", 0), "p01_settled": o1.get("bank_settled", 0),
                    "p01_diff": o1.get("difference", 0),
                    "p02_total": o2.get("total", 0), "p02_open": o2.get("open", 0),
                    "p02_open_amt": o2.get("amount_open", 0),
                    "p03_total": o3.get("total", 0), "p03_open": o3.get("open", 0),
                    "p03_open_amt": o3.get("amount_open", 0),
                    "total_open_amt": round((o2.get("amount_open", 0) or 0) + (o3.get("amount_open", 0) or 0), 2),
                })
            overview.sort(key=lambda r: -r["total_open_amt"])
            _sheet(w, "KO Overview", overview,
                   ["ko_csp", "p01_wallet", "p01_settled", "p01_diff", "p02_total", "p02_open",
                    "p02_open_amt", "p03_total", "p03_open", "p03_open_amt", "total_open_amt"])
            _sheet(w, "P01 by KO", sorted(p01a.values(), key=lambda r: r["ko_csp"]),
                   ["ko_csp", "wallet_withdrawn", "bank_settled", "difference", "days", "days_credited"])
            _sheet(w, "P02 by KO", sorted(p02a.values(), key=lambda r: -r["amount_open"]),
                   ["ko_csp", "total", "matched", "open", "amount_total", "amount_open"])
            _sheet(w, "P03 by CSP", sorted(p03a.values(), key=lambda r: -r["amount_open"]),
                   ["ko_csp", "total", "matched", "open", "amount_total", "amount_open"])

        elif t == "unified":
            res = get_unified(recon_date=rd, upload_date=None, side=None, status=None,
                              process=None, search=None, page=1, page_size=10 ** 9,
                              db=db, current_user=current_user) if rd else {"rows": []}
            cols = ["side", "source", "ref", "ko_csp", "amount", "date", "drcr", "status",
                    "process", "counterpart", "also_p03", "src_code", "src_note", "narration"]
            _sheet(w, "All Entries", res["rows"], cols)
            sc = res.get("status_counts") or {}
            _sheet(w, "Status Counts",
                   [{"status": k, "entries": v} for k, v in sorted(sc.items(), key=lambda x: -x[1])],
                   ["status", "entries"])

        elif t == "p01_lines":
            res = get_p01_lines(recon_date=rd, upload_date=None, ko_id=None, status=None,
                                db=db, current_user=current_user) if rd else {"kos": []}
            kos, wlines, slines = [], [], []
            for g in res["kos"]:
                kos.append({k: g.get(k) for k in ("ko_id", "status", "wallet_withdrawn", "bank_settled",
                                                  "difference", "deduct_date", "bank_txn_date",
                                                  "src_code", "src_note")})
                for x in g.get("withdrawals") or []:
                    wlines.append({"ko_id": g["ko_id"], **x})
                for x in g.get("settlements") or []:
                    slines.append({"ko_id": g["ko_id"], **x})
            _sheet(w, "KO Summary", kos, ["ko_id", "status", "wallet_withdrawn", "bank_settled",
                                          "difference", "deduct_date", "bank_txn_date", "src_code", "src_note"])
            _sheet(w, "Withdrawal Lines", wlines, ["ko_id", "amount", "txn_date", "datetime", "configured_by"])
            _sheet(w, "Settlement Lines", slines, ["ko_id", "amount", "txn_date", "deduct_date", "ref", "description"])

        elif t == "reversals":
            recs = [r for r in _load_process_recs(db, "p02", rd, df_, dt_, with_bank_desc=True)
                    if r.get("reversal_type") in ("Reversal Debit", "Reversal Credit")]
            recs.sort(key=lambda r: (r.get("reference_number") or "", r.get("bank_type") or ""))
            _sheet(w, "Reversals", recs, _P02_XCOLS)

        elif t == "p04_actions":
            recs = _load_process_recs(db, "p04", rd, df_, dt_)
            _sheet(w, "All Actions", recs, _P04_XCOLS)
            _sheet(w, "Pending", [r for r in recs if r.get("action_required") != "NONE"
                                  and not r.get("action_done")], _P04_XCOLS)
            _sheet(w, "Done", [r for r in recs if r.get("action_done")], _P04_XCOLS)

        elif t == "src_log":
            q = _rng(db.query(SBISrcAssignment), SBISrcAssignment, rd, df_, dt_)
            rows = [{"recon_date": s.recon_date, "process": s.process, "match_key": s.match_key,
                     "src_code": s.src_code, "src_note": s.src_note, "by": s.created_by,
                     "at": str(s.created_at or "")} for s in q.order_by(SBISrcAssignment.created_at.desc()).all()]
            _sheet(w, "SRC Log", rows, ["recon_date", "process", "match_key", "src_code", "src_note", "by", "at"])

        elif t == "manual_log":
            q = _rng(db.query(SBIManualMatch), SBIManualMatch, rd, df_, dt_)
            rows = [{"recon_date": m.recon_date, "process": m.process, "match_key": m.match_key,
                     "counterpart_ref": m.counterpart_ref, "remark": m.remark, "by": m.created_by,
                     "at": str(m.created_at or "")} for m in q.order_by(SBIManualMatch.created_at.desc()).all()]
            _sheet(w, "Manual Matches", rows,
                   ["recon_date", "process", "match_key", "counterpart_ref", "remark", "by", "at"])

    return output.getvalue(), f"sbi_{t}_{tag}.xlsx", str(tag)


@router.get("/report")
def sbi_report(
    type: str = Query(..., description="one of the /sbi/report-options ids"),
    recon_date: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Build one report from the SBI report library as a styled Excel workbook."""
    data, fname, tag = build_sbi_report_xlsx(db, type, recon_date, date_from, date_to, current_user)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Recon-Date": tag,
            "Access-Control-Expose-Headers": "X-Recon-Date",
        },
    )


# ── Operator output workbooks (reference-based bank↔source reconciliation) ──────
# These are the two files finance ops reconcile against by hand, reproduced from our
# data by core/sbi_reports.py (read-only; does NOT touch P01-P04 results).

def _resolve_bank_date(db, recon_date):
    """Use the requested date if it has bank rows; else fall back to the latest date
    that does (so a download never silently returns an empty workbook)."""
    if recon_date and db.query(SBIBankTransaction.id).filter(
            SBIBankTransaction.txn_date == recon_date).first():
        return recon_date
    return db.query(func.max(SBIBankTransaction.txn_date)).filter(
        SBIBankTransaction.txn_date.like("20%")).scalar()


# Range cap for the operator workbooks: each date is a full independent reconcile pass
# on a single worker, so an accidental all-time range must not freeze the app
# (the sync-full-recon-on-upload lesson). A calendar month fits.
_MAX_WORKBOOK_DATES = 31


def _resolve_workbook_dates(db, date_from, date_to):
    """Distinct business dates with SBI bank, txn-report OR KO-limits rows inside
    [from, to] (either bound optional), sentinel dates excluded. Sorted ascending.
    KO-limits-only dates (e.g. a bank holiday with limit activity) must be included
    or their rows silently vanish from the range Limit & Settlement sheet."""
    if date_from and date_to and date_from > date_to:
        raise HTTPException(status_code=422, detail="date_from is after date_to")
    def _dates(model):
        q = db.query(model.txn_date).filter(model.txn_date.like("20%"))
        if date_from:
            q = q.filter(model.txn_date >= date_from)
        if date_to:
            q = q.filter(model.txn_date <= date_to)
        return {d for (d,) in q.distinct().all() if d}
    dates = sorted(_dates(SBIBankTransaction) | _dates(SBITxnReport) | _dates(SBIKOLimits))
    if len(dates) > _MAX_WORKBOOK_DATES:
        raise HTTPException(
            status_code=422,
            detail=f"The range covers {len(dates)} business dates — max "
                   f"{_MAX_WORKBOOK_DATES} per workbook. Narrow the date range.")
    return dates


@router.get("/reports/reconciliation")
def sbi_reconciliation_report(
    recon_date: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Bank-statement-centric reconciliation workbook: every bank row annotated with the
    source product/type it matched (by 20-digit reference), plus unmatched (both sides,
    Success-first) and duplicate/reversal sheets.

    Single date (recon_date, latest-with-data fallback) or a date range
    (date_from/date_to → one workbook, each date reconciled independently)."""
    from core.sbi_reports import (build_reconciliation_report,
                                  build_reconciliation_report_range)
    if date_from or date_to:
        dates = _resolve_workbook_dates(db, date_from, date_to)
        if not dates:
            raise HTTPException(status_code=404, detail="No SBI data in the selected date range")
        if len(dates) == 1:      # one date in range → the exact single-date workbook
            out, tag = build_reconciliation_report(db, dates[0]), dates[0]
        else:
            out, tag = build_reconciliation_report_range(db, dates), f"{dates[0]}_to_{dates[-1]}"
        return StreamingResponse(
            out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="Reconciliation_Report_{tag}.xlsx"',
                     "X-Recon-Date": str(tag), "Access-Control-Expose-Headers": "X-Recon-Date"})
    d = _resolve_bank_date(db, recon_date)
    if not d:
        raise HTTPException(status_code=404, detail="No SBI bank statement data to report on")
    out = build_reconciliation_report(db, d)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Reconciliation_Report_{d}.xlsx"',
                 "X-Recon-Date": str(d), "Access-Control-Expose-Headers": "X-Recon-Date"})


@router.get("/reports/source-match")
def sbi_source_match_report(
    recon_date: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Source-file-centric match-status workbook: a Summary split by Success/Failure, a
    Limit & Settlement sheet, and one sheet per source product with each row's match status
    highlighted (green matched / orange Success-unmatched / red Failure).

    Single date (recon_date, latest-with-data fallback) or a date range
    (date_from/date_to → one workbook, each date reconciled independently)."""
    from core.sbi_reports import (build_source_match_report,
                                  build_source_match_report_range)
    if date_from or date_to:
        dates = _resolve_workbook_dates(db, date_from, date_to)
        if not dates:
            raise HTTPException(status_code=404, detail="No SBI data in the selected date range")
        if len(dates) == 1:      # one date in range → the exact single-date workbook
            out, tag = build_source_match_report(db, dates[0]), dates[0]
        else:
            out, tag = build_source_match_report_range(db, dates), f"{dates[0]}_to_{dates[-1]}"
        return StreamingResponse(
            out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="Source_Files_Match_Status_{tag}.xlsx"',
                     "X-Recon-Date": str(tag), "Access-Control-Expose-Headers": "X-Recon-Date"})
    d = _resolve_bank_date(db, recon_date)
    if not d:
        raise HTTPException(status_code=404, detail="No SBI data to report on")
    out = build_source_match_report(db, d)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Source_Files_Match_Status_{d}.xlsx"',
                 "X-Recon-Date": str(d), "Access-Control-Expose-Headers": "X-Recon-Date"})


def _build_sbi_all_entries(db, dates, current_user) -> io.BytesIO:
    """One 'All Entries' sheet: every BANK row and every INTERNAL row (Txn Reports + KO
    Withdrawals) across the given business dates, each with its recon status, process and
    counterpart. Rows are the per-date unified view concatenated (never crosses a date
    boundary — same guarantee as the operator-workbook range builders). Plus a
    'Status Counts' summary sheet."""
    cols = ["date", "side", "source", "ref", "ko_csp", "amount", "drcr", "status",
            "process", "counterpart", "also_p03", "src_code", "src_note", "narration"]
    all_rows, sc = [], {}
    for d in dates:
        res = get_unified(recon_date=d, upload_date=None, side=None, status=None,
                          process=None, search=None, page=1, page_size=10 ** 9,
                          db=db, current_user=current_user)
        all_rows += res.get("rows", [])
        for k, v in (res.get("status_counts") or {}).items():
            sc[k] = sc.get(k, 0) + v
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        _sheet(w, "All Entries", all_rows, cols)
        _sheet(w, "Status Counts",
               [{"status": k, "entries": v} for k, v in sorted(sc.items(), key=lambda x: -x[1])],
               ["status", "entries"])
    out.seek(0)
    return out


@router.get("/reports/all-entries")
def sbi_all_entries_report(
    recon_date: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """One sheet: ALL bank-side rows vs ALL internal rows (Txn Reports + KO Withdrawals)
    with each row's recon status/process/counterpart. Single date (recon_date, latest
    with data as fallback) OR a date range (date_from/date_to → concatenated, max 31 days).
    Covers the full bank↔internal match universe (P01/P02/P03); the P04 wallet-balance
    sources (KO Deposits / Cash Holding / Limit Failures) have no bank counterpart and are
    not listed."""
    if date_from or date_to:
        dates = _resolve_workbook_dates(db, date_from, date_to)
        if not dates:
            raise HTTPException(status_code=404, detail="No SBI data in the selected date range")
        tag = dates[0] if len(dates) == 1 else f"{dates[0]}_to_{dates[-1]}"
    else:
        # single date: use it if it has any SBI data, else the latest date that does
        if recon_date and _resolve_workbook_dates(db, recon_date, recon_date):
            d = recon_date
        else:
            d = max((x for x in (
                db.query(func.max(m.txn_date)).filter(m.txn_date.like("20%")).scalar()
                for m in (SBIBankTransaction, SBITxnReport, SBIKOLimits)) if x), default=None)
        if not d:
            raise HTTPException(status_code=404, detail="No SBI data to report on")
        dates, tag = [d], d
    out = _build_sbi_all_entries(db, dates, current_user)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="SBI_All_Entries_{tag}.xlsx"',
                 "X-Recon-Date": str(tag), "Access-Control-Expose-Headers": "X-Recon-Date"})


# ── Run all four processes in sequence (QoL orchestration — same code paths) ──

def _backfill_dateless_txn_dates(db):
    """Some 'Deposit' transaction-report files ingest with NO per-row date (their date column
    header differs), so those rows never fall into any recon date and their bank counterparts
    stay Unmatched despite an exact 20-digit reference + amount. The reference is globally
    unique, so heal each dateless report by copying the txn date from the bank row carrying the
    same reference. Idempotent (only touches still-dateless rows); returns the set of business
    dates it assigned so the caller reconciles them too. Never raises to the caller's flow."""
    dateless = (db.query(SBITxnReport)
                .filter((SBITxnReport.txn_date.is_(None)) | (SBITxnReport.txn_date == ''),
                        SBITxnReport.reference_number.isnot(None)).all())
    refs = sorted({(r.reference_number or "").strip() for r in dateless if (r.reference_number or "").strip()})
    if not refs:
        return set()
    bank_date = {}                          # reference -> bank txn_date (first non-empty wins)
    for ref, td in (db.query(SBIBankTransaction.ref_number, SBIBankTransaction.txn_date)
                    .filter(SBIBankTransaction.ref_number.in_(refs)).all()):
        r = (ref or "").strip()
        if r and td and r not in bank_date:
            bank_date[r] = td
    healed = set()
    for rpt in dateless:
        d = bank_date.get((rpt.reference_number or "").strip())
        if d:
            rpt.txn_date = d
            healed.add(d)
    if healed:
        db.commit()
    return healed


def _run_all_dates(db, current_user, dates):
    """Run P01→P04 for each business date in `dates`. Each process is wipe-guarded, so a
    date missing its counterpart file is a no-op (never a wipe). A failure in one
    process/date is recorded but never blocks the rest."""
    # First heal any date-less transaction reports (e.g. the 'Deposit' file) from their bank
    # counterpart's date, and reconcile those business dates too — otherwise their exact-ref
    # bank matches would be silently missed.
    try:
        healed = _backfill_dateless_txn_dates(db)
        if healed:
            dates = sorted(set(dates) | healed)
    except Exception:
        try: db.rollback()
        except Exception: pass
    per_date = {}
    for d in dates:
        r = {}
        for name, fn in (("p01", run_p01), ("p02", run_p02), ("p03", run_p03), ("p04", run_p04)):
            try:
                r[name] = fn(recon_date=d, upload_date=None, db=db, current_user=current_user)
            except Exception as e:
                try: db.rollback()
                except Exception: pass
                r[name] = {"error": str(e)[:200]}
        per_date[d] = r
    return per_date


@router.post("/run/all")
def run_all(
    recon_date: Optional[str] = None,
    upload_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("run_recon")),
):
    """Run P01 → P02 → P03 → P04. Pass `recon_date` for a single business date, or omit it
    to reconcile EVERY business date present in the data (so one bulk upload reconciles
    each day it contains). Each process is wipe-guarded — a date with no source data is
    skipped, never wiped. A failure in one process/date is reported, not fatal."""
    dates = [recon_date] if recon_date else _sbi_business_dates(db)
    per_date = _run_all_dates(db, current_user, dates)
    _audit(db, current_user, "sbi_run_all",
           {"dates": dates, "count": len(dates), "single": bool(recon_date)})
    return ({"recon_date": recon_date, "results": per_date.get(recon_date, {})}
            if recon_date else {"dates": dates, "results": per_date})


@router.get("/readiness")
def readiness(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Per business date: which of the source files are present + the current P02 match
    rate. This makes the confusion visible — a low rate is almost always a MISSING FILE
    (no transaction report uploaded that day), not broken reconciliation."""
    from sqlalchemy import func as F

    def by_date(model, col):
        out = {}
        for d, n in db.query(col, F.count(model.id)).filter(col.isnot(None)).group_by(col).all():
            if d and len(d) >= 10:
                out[d[:10]] = out.get(d[:10], 0) + n
        return out

    bank  = by_date(SBIBankTransaction, SBIBankTransaction.txn_date)
    txn   = by_date(SBITxnReport, SBITxnReport.txn_date)
    kolim = by_date(SBIKOLimits, SBIKOLimits.txn_date)
    cash  = by_date(SBIKOCashHolding, SBIKOCashHolding.report_date)
    fail  = by_date(SBILimitFailure, SBILimitFailure.txn_date)

    # Which of the SIX source files are present per business date. Finance ops upload them
    # separately, and it's common for one (often the small Deposit file) to be missing while
    # the rest are in — a partial-report day that reconciles low for a benign reason.
    from core.sbi_reports import canonical_product, PRODUCTS
    src_files = {}   # date -> {product: count}
    for d, sf, n in db.query(SBITxnReport.txn_date, SBITxnReport.source_file,
                             F.count(SBITxnReport.id)).filter(
                             SBITxnReport.txn_date.isnot(None)).group_by(
                             SBITxnReport.txn_date, SBITxnReport.source_file).all():
        if d and len(d) >= 10:
            src_files.setdefault(d[:10], {})
            prod = canonical_product(sf)
            src_files[d[:10]][prod] = src_files[d[:10]].get(prod, 0) + n

    p02 = {}
    for d, st, n in db.query(SBIP02Result.recon_date, SBIP02Result.match_status,
                             F.count(SBIP02Result.id)).group_by(
                             SBIP02Result.recon_date, SBIP02Result.match_status).all():
        if d:
            p02.setdefault(d[:10], {})[st] = n

    dates = sorted(set(bank) | set(txn) | set(kolim) | set(cash) | set(fail))
    rows = []
    for d in dates:
        p = p02.get(d, {}); tot = sum(p.values())
        m = p.get("Matched", 0); rev = p.get("Reversal", 0)
        reconciled = m + rev   # a reversal is a net-zero DR+CR pair — reconciled, not open
        sf = src_files.get(d, {})
        present = [pr for pr in PRODUCTS if sf.get(pr)]
        missing_files = [pr for pr in PRODUCTS if not sf.get(pr)]
        rows.append({
            "date": d,
            "bank": bank.get(d, 0), "txn_report": txn.get(d, 0),
            "ko_limits": kolim.get(d, 0), "cash_holding": cash.get(d, 0),
            "limit_failures": fail.get(d, 0),
            # the six source files: present / missing (a partial-report day reconciles low benignly)
            "source_files_present": present, "source_files_missing": missing_files,
            "source_files_have": len(present), "source_files_total": len(PRODUCTS),
            # P02 can only reconcile a day that has BOTH a bank statement and a txn report
            "p02_ready": bool(bank.get(d) and txn.get(d)),
            "p02_total": tot, "p02_matched": m, "p02_reversal": rev,
            "p02_reconciled": reconciled,
            # rate counts reversals as reconciled (they net to zero) — they no longer drag it
            "p02_rate": round(100 * reconciled / tot, 1) if tot else None,
            "missing": [name for name, avail in (
                ("Bank statement", bank.get(d)), ("Transaction report", txn.get(d)))
                if not avail],
        })

    from models.database import SBICSPMaster
    p02_m = sum(p.get("Matched", 0) for p in p02.values())
    p02_rev = sum(p.get("Reversal", 0) for p in p02.values())
    p02_recon = p02_m + p02_rev
    p02_t = sum(sum(p.values()) for p in p02.values())
    totals = {
        "bank": sum(bank.values()), "txn_report": sum(txn.values()),
        "ko_limits": sum(kolim.values()), "cash_holding": sum(cash.values()),
        "limit_failures": sum(fail.values()),
        "csp_master": db.query(F.count(SBICSPMaster.id)).scalar() or 0,
        "days_with_data": len(dates),
        "days_missing_report": sum(1 for r in rows if r["bank"] and not r["txn_report"]),
        "p02_matched": p02_m, "p02_reversal": p02_rev, "p02_reconciled": p02_recon,
        "p02_total": p02_t,
        "p02_rate": round(100 * p02_recon / p02_t, 1) if p02_t else None,
    }
    return {"dates": rows, "latest": dates[-1] if dates else None, "totals": totals}


def _auto_run_after_upload(db, current_user, dates=None):
    """Auto-recon after every SBI upload. Reconciles ONLY the business dates the upload
    touched — NOT every date, which (at 22 dates × P01-P04) took minutes and froze the
    single worker on each upload. `dates=None` means "unknown → reconcile all" (kept for
    callers that can't scope, e.g. reference-data uploads); an EMPTY list means "this file
    had no dates → reconcile nothing" (must NOT fall through to all). Each process is
    wipe-guarded, so a date missing a counterpart file is a harmless no-op the next upload
    completes. Every failure is swallowed so an upload never blocks."""
    try:
        if dates is None:
            ds = _sbi_business_dates(db)
        else:
            ds = sorted({d[:10] for d in dates if d and len(d) >= 10})
        if not ds:
            return {"dates": [], "note": "no business dates to reconcile"}
        _run_all_dates(db, current_user, ds)
        return {"dates": ds}
    except Exception as e:
        try: db.rollback()
        except Exception: pass
        return {"error": str(e)[:200]}


# ── Manual match (persistent overlay across re-runs) ──────────────────────────
# P02/P03 results are delete-and-recreated each run (#17), so a manual match is
# stored keyed by the row's STABLE business key and overlaid at READ time onto the
# results + exports — surviving re-runs without touching the run logic. P01/P04
# don't carry a bank↔report pairing, so manual match applies to P02 and P03.

def _manual_key(process: str, row: dict):
    """Stable business key for a result row (survives delete-and-recreate)."""
    p = (process or "").lower()
    if p == "p02":
        ref = row.get("reference_number")
        return f"{ref}|{row.get('bank_type') or ''}" if ref else None
    if p == "p03":
        csp, ref = row.get("csp_code"), row.get("ref_number")
        return f"{csp or ''}|{ref or ''}" if (csp or ref) else None
    return None


def _apply_manual_matches(db, process: str, recon_date, rows: list) -> list:
    """Overlay persisted manual matches onto serialized result dicts (read-time)."""
    p = (process or "").lower()
    if p not in ("p02", "p03") or not rows:
        return rows
    q = db.query(SBIManualMatch).filter(SBIManualMatch.process == p)
    if recon_date:
        q = q.filter(SBIManualMatch.recon_date == recon_date)
    mm = {(m.recon_date, m.match_key): m for m in q.all()}
    if not mm:
        return rows
    for r in rows:
        key = _manual_key(p, r)
        m = mm.get((r.get("recon_date"), key)) if key else None
        if m:
            r["match_status"] = "Manual_Matched"
            r["manual_remark"] = m.remark
            r["manual_counterpart"] = m.counterpart_ref
            r["manual_match_id"] = m.id
    return rows


def _src_key(process: str, row: dict):
    """Stable business key for an SRC tag (survives delete-and-recreate, per #17)."""
    p = (process or "").lower()
    if p == "p01":
        return row.get("ko_id") or None
    if p == "p02":
        ref = row.get("reference_number")
        return f"{ref}|{row.get('bank_type') or ''}" if ref else None
    if p == "p03":
        csp, ref = row.get("csp_code"), row.get("ref_number")
        return f"{csp or ''}|{ref or ''}" if (csp or ref) else None
    if p == "p04":
        return row.get("csp_code") or None
    return None


def _apply_src_assignments(db, process: str, recon_date, rows: list) -> list:
    """Overlay persisted SRC tags onto serialized result dicts (read-time).

    Adds src_code/src_note to every row (None when untagged). Survives re-runs because
    the tag lives in SBISrcAssignment keyed by the stable business key, not the row id."""
    p = (process or "").lower()
    sa = {}
    if p in ("p01", "p02", "p03", "p04") and rows:
        q = db.query(SBISrcAssignment).filter(SBISrcAssignment.process == p)
        if recon_date:
            q = q.filter(SBISrcAssignment.recon_date == recon_date)
        sa = {(a.recon_date, a.match_key): a for a in q.all()}
    for r in rows:
        key = _src_key(p, r)
        a = sa.get((r.get("recon_date"), key)) if key else None
        r["src_code"] = a.src_code if a else None
        r["src_note"] = a.src_note if a else None
    return rows


class ManualMatchIn(BaseModel):
    process: str
    result_id: str
    counterpart_ref: Optional[str] = None
    remark: str = ""


@router.post("/manual-match")
def create_manual_match(
    body: ManualMatchIn,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("src_assign")),
):
    """Manually resolve an unmatched P02/P03 row. Persists across re-runs (overlay)."""
    p = (body.process or "").lower()
    model = {"p02": SBIP02Result, "p03": SBIP03Result}.get(p)
    if not model:
        raise HTTPException(status_code=400, detail="Manual match supports process p02 or p03")
    if len((body.remark or "").strip()) < 5:
        raise HTTPException(status_code=400, detail="A remark (≥5 characters) is required")
    row = db.query(model).filter(model.id == body.result_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Result row not found")
    row_dict = {c.name: getattr(row, c.name) for c in model.__table__.columns}
    key = _manual_key(p, row_dict)
    if not key:
        raise HTTPException(status_code=400, detail="Row has no stable key to match on")
    existing = db.query(SBIManualMatch).filter(
        SBIManualMatch.recon_date == row.recon_date,
        SBIManualMatch.process == p,
        SBIManualMatch.match_key == key,
    ).first()
    queued = maker_checker.intercept(
        db, current_user, "sbi_manual_match",
        payload={"process": body.process, "result_id": body.result_id,
                 "counterpart_ref": body.counterpart_ref, "remark": body.remark},
        summary=f"SBI Kiosk manual match: {p} {body.result_id}",
        partner="kiosk")
    if queued:
        return queued
    if existing:
        existing.counterpart_ref = body.counterpart_ref
        existing.remark = body.remark.strip()
        existing.created_by = current_user.username
        mm = existing
    else:
        mm = SBIManualMatch(recon_date=row.recon_date, process=p, match_key=key,
                            counterpart_ref=body.counterpart_ref, remark=body.remark.strip(),
                            created_by=current_user.username)
        db.add(mm)
    try:
        db.add(AuditLog(user_id=current_user.id, username=current_user.username,
                        action="sbi_manual_match", action_type="human",
                        entity_type="sbi", entity_id=mm.id,
                        detail=json.dumps({"process": p, "recon_date": row.recon_date,
                                           "match_key": key, "counterpart_ref": body.counterpart_ref,
                                           "remark": body.remark.strip()})))
    except Exception:
        pass
    db.commit()
    return {"id": mm.id, "match_status": "Manual_Matched", "match_key": key}


@router.delete("/manual-match/{mm_id}")
def delete_manual_match(
    mm_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("src_assign")),
):
    """Undo a manual match (reverts the row to its algorithm-computed status)."""
    mm = db.query(SBIManualMatch).filter(SBIManualMatch.id == mm_id).first()
    if not mm:
        raise HTTPException(status_code=404, detail="Manual match not found")
    queued = maker_checker.intercept(
        db, current_user, "sbi_delete_manual_match",
        payload={"mm_id": mm_id},
        summary=f"SBI Kiosk undo manual match: {mm_id}",
        partner="kiosk")
    if queued:
        return queued
    db.delete(mm)
    try:
        db.add(AuditLog(user_id=current_user.id, username=current_user.username,
                        action="sbi_manual_unmatch", action_type="human",
                        entity_type="sbi", entity_id=mm_id,
                        detail=json.dumps({"process": mm.process, "match_key": mm.match_key})))
    except Exception:
        pass
    db.commit()
    return {"deleted": mm_id}


@router.get("/manual-match")
def list_manual_matches(
    recon_date: Optional[str] = None,
    process: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = db.query(SBIManualMatch)
    if recon_date: q = q.filter(SBIManualMatch.recon_date == recon_date)
    if process:    q = q.filter(SBIManualMatch.process == process.lower())
    rows = q.order_by(SBIManualMatch.created_at.desc()).all()
    return {"rows": [{"id": m.id, "recon_date": m.recon_date, "process": m.process,
                      "match_key": m.match_key, "counterpart_ref": m.counterpart_ref,
                      "remark": m.remark, "by": m.created_by, "at": str(m.created_at)} for m in rows],
            "count": len(rows)}


# ── Two-sided manual PAIR (SBI Kiosk pair-picker) ─────────────────────────────
# A first-class bank-row↔internal-row link (unlike the one-sided SBIManualMatch above).
# Stored in SBIManualPair keyed by each side's stable business key, overlaid onto the
# unified read model (_apply_pairs) so both rows flip to Manual_Matched and the exports
# reconcile. Free-form: any bank row may pair with any internal row; amount/bucket
# mismatches are surfaced as a warning, not blocked. Internal side spans Txn Reports (each
# labelled by its source file), KO Withdrawals and KO Deposits — NOT Limit Failures /
# Cash Holding (they have no bank counterpart).

@router.get("/manual-match/open-items")
def manual_pair_open_items(
    side: str,                       # bank | data
    date_from: str,
    date_to: Optional[str] = None,
    search: Optional[str] = None,    # ref / KO / CSP / file
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    page: int = 1,
    page_size: int = 200,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Open (unmatched) SBI Kiosk rows for ONE side across a date range — feeds the
    pair-picker. Bank side = settlements + statement + credits; data side = Txn Reports
    (each carrying its originating file), KO Withdrawals and KO Deposits."""
    if side not in ("bank", "data"):
        raise HTTPException(status_code=400, detail="side must be 'bank' or 'data'")
    dates = _resolve_workbook_dates(db, date_from, date_to or date_from)   # 422 on bad range / >31 dates
    rows = []
    for d in dates:
        for e in _unified_entries(db, d, include_deposits=True):
            e.pop("_result", None)
            if e["side"] != side or e["status"] in _MATCHED_UNIFIED:
                continue
            rows.append(e)
    if search:
        s = search.lower()
        rows = [e for e in rows if s in (e["ref"] or "").lower() or s in (e["ko_csp"] or "").lower()
                or s in (e["file"] or "").lower()]
    if amount_min is not None:
        rows = [e for e in rows if float(e["amount"] or 0) >= amount_min]
    if amount_max is not None:
        rows = [e for e in rows if float(e["amount"] or 0) <= amount_max]
    rows.sort(key=lambda e: (e["date"] or "", e["ko_csp"] or "", e["ref"] or ""))
    total = len(rows)
    page_rows = rows[(page - 1) * page_size: page * page_size]
    return {"items": page_rows, "total": total, "page": page, "page_size": page_size, "dates": dates}


def _bank_descriptor(b):
    """Resolve a bank source row to its stable, id-independent pair descriptor."""
    src = "Bank Settlement" if b.is_settlement else "Bank Statement"
    drcr = "DR" if (b.debit or 0) > 0 else ("CR" if (b.credit or 0) > 0 else "")
    amt = (b.debit or 0) if (b.debit or 0) > 0 else (b.credit or 0)
    # normalise the ref the SAME way _unified_entries does (placeholder → '') so the write-time
    # key matches the read-time key computed in _apply_pairs.
    key = _pair_key("bank", src, b.ko_id, _clean_refno(b.ref_number), b.txn_date, drcr, amt, _row_disc(src, b))
    return {"key": key, "source": src, "amount": float(amt or 0), "date": b.txn_date}


def _data_descriptor(row, data_source):
    """Resolve an internal source row to its stable, id-independent pair descriptor."""
    if data_source == "Txn Report":
        ko, ref, date, amt = row.ko_id, row.reference_number, row.txn_date, (row.amount or 0)
    else:  # KO Withdrawal | KO Deposit
        ko, ref, date, amt = row.ko_id, "", row.txn_date, (row.amount or 0)
    key = _pair_key("data", data_source, ko, ref, date, "", amt, _row_disc(data_source, row))
    return {"key": key, "source": data_source, "amount": float(amt or 0), "date": date}


class ManualPairIn(BaseModel):
    data_source: str                 # Txn Report | KO Withdrawal | KO Deposit
    # A fresh submit sends the two source-row ids (resolved to stable keys server-side). A
    # maker-checker replay instead carries the RESOLVED keys below, so an approved pair never
    # depends on regenerable source ids (behavior-contract #17 — ids change on re-upload).
    bank_id: Optional[str] = None
    data_id: Optional[str] = None
    bank_key: Optional[str] = None
    data_key: Optional[str] = None
    bank_source: Optional[str] = None
    bank_amount: Optional[float] = None
    data_amount: Optional[float] = None
    bank_date: Optional[str] = None
    data_date: Optional[str] = None


class ManualPairBulkIn(BaseModel):
    pairs: List[ManualPairIn]
    remark: str = Field("", max_length=500)


@router.post("/manual-match/pair")
def create_manual_pairs(
    body: ManualPairBulkIn,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("src_assign")),
):
    """Link chosen (bank row, internal row) pairs. Free-form; amount/bucket mismatches are
    warned, not blocked. Persists as a business-key overlay (survives re-runs + re-uploads),
    shows as Manual_Matched in the unified view + exports. Remark (5-500) required; maker-checker
    gated; a bank OR internal row already paired is rejected (1:1)."""
    if not body.pairs:
        raise HTTPException(status_code=400, detail="No pairs submitted")
    remark = (body.remark or "").strip()
    if len(remark) < 5:
        raise HTTPException(status_code=400, detail="A remark (≥5 characters) is required")

    # Phase 1 — resolve every pair to STABLE business keys (id-independent) BEFORE the
    # maker-checker gate, so a queued/approved pair survives a re-upload (#17). A replayed
    # pair already carries its keys and skips the source-row lookup entirely.
    existing = db.query(SBIManualPair).all()
    used_bank = {p.bank_key for p in existing}
    used_data = {p.data_key for p in existing}
    resolved, results, errors, warned = [], [], 0, 0
    for pr in body.pairs:
        try:
            if pr.bank_key and pr.data_key:                      # replay: keys already resolved
                bd = {"key": pr.bank_key, "source": pr.bank_source,
                      "amount": float(pr.bank_amount or 0), "date": pr.bank_date}
                dd = {"key": pr.data_key, "source": pr.data_source,
                      "amount": float(pr.data_amount or 0), "date": pr.data_date}
            else:                                                # fresh submit: resolve source rows
                b = db.query(SBIBankTransaction).filter(SBIBankTransaction.id == pr.bank_id).first()
                if not b:
                    raise ValueError("Bank row not found")
                bd = _bank_descriptor(b)
                ds = pr.data_source
                if ds == "Txn Report":
                    row = db.query(SBITxnReport).filter(SBITxnReport.id == pr.data_id).first()
                elif ds in ("KO Withdrawal", "KO Deposit"):
                    row = db.query(SBIKOLimits).filter(SBIKOLimits.id == pr.data_id,
                                                       SBIKOLimits.txn_type == ds).first()
                else:
                    raise ValueError(f"Unsupported data source '{ds}'")
                if not row:
                    raise ValueError(f"{ds} row not found")
                dd = _data_descriptor(row, ds)

            if bd["key"] in used_bank:
                raise ValueError("Bank row already manually matched")
            if dd["key"] in used_data:
                raise ValueError("Internal row already manually matched")
            used_bank.add(bd["key"])
            used_data.add(dd["key"])
            diff = round(abs(bd["amount"] - dd["amount"]), 2)
            resolved.append({"bank_key": bd["key"], "data_key": dd["key"],
                             "bank_source": bd["source"], "data_source": dd["source"],
                             "bank_amount": bd["amount"], "data_amount": dd["amount"],
                             "bank_date": bd["date"], "data_date": dd["date"], "_res_idx": len(results)})
            results.append({"bank_id": pr.bank_id, "data_id": pr.data_id, "status": "ok",
                            "pair_id": None, "amount_diff": diff, "warn": diff > TOL})
            if diff > TOL:
                warned += 1
        except Exception as ex:
            errors += 1
            results.append({"bank_id": pr.bank_id, "data_id": pr.data_id,
                            "status": "error", "error": str(ex)[:200]})

    if not resolved:
        return {"paired": 0, "errors": errors, "warned": 0, "results": results}

    # Phase 2 — maker-checker gate on the RESOLVED descriptors (no volatile source ids).
    queued = maker_checker.intercept(
        db, current_user, "sbi_manual_pair",
        payload={"pairs": [{"data_source": r["data_source"], "bank_key": r["bank_key"],
                            "data_key": r["data_key"], "bank_source": r["bank_source"],
                            "bank_amount": r["bank_amount"], "data_amount": r["data_amount"],
                            "bank_date": r["bank_date"], "data_date": r["data_date"]} for r in resolved],
                 "remark": remark},
        summary=f"SBI Kiosk manual pair: {len(resolved)} pair(s)",
        partner="kiosk")
    if queued:
        return queued

    # Phase 3 — persist. Guard the commit so a DB error rolls back cleanly instead of
    # leaving a poisoned session that 500s with nothing saved.
    try:
        for r in resolved:
            mp = SBIManualPair(
                bank_date=r["bank_date"], data_date=r["data_date"], bank_key=r["bank_key"],
                data_key=r["data_key"], bank_source=r["bank_source"], data_source=r["data_source"],
                bank_amount=r["bank_amount"], data_amount=r["data_amount"],
                remark=remark, created_by=current_user.username)
            db.add(mp)
            db.flush()
            results[r["_res_idx"]]["pair_id"] = mp.id
            try:
                db.add(AuditLog(user_id=current_user.id, username=current_user.username,
                                action="sbi_manual_pair", action_type="human",
                                entity_type="sbi", entity_id=mp.id,
                                detail=json.dumps({"bank_key": r["bank_key"], "data_key": r["data_key"],
                                                   "bank_source": r["bank_source"], "data_source": r["data_source"],
                                                   "bank_amount": r["bank_amount"], "data_amount": r["data_amount"],
                                                   "remark": remark})))
            except Exception:
                pass
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to save pairs: {str(ex)[:200]}")

    return {"paired": len(resolved), "errors": errors, "warned": warned, "results": results}


@router.delete("/manual-match/pair/{pair_id}")
def delete_manual_pair(
    pair_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("src_assign")),
):
    """Undo a two-sided manual pair — both rows revert to their algorithm-computed status."""
    mp = db.query(SBIManualPair).filter(SBIManualPair.id == pair_id).first()
    if not mp:
        raise HTTPException(status_code=404, detail="Manual pair not found")
    queued = maker_checker.intercept(
        db, current_user, "sbi_delete_manual_pair",
        payload={"pair_id": pair_id},
        summary=f"SBI Kiosk undo manual pair: {pair_id}",
        partner="kiosk")
    if queued:
        return queued
    bk, dk = mp.bank_key, mp.data_key
    db.delete(mp)
    try:
        db.add(AuditLog(user_id=current_user.id, username=current_user.username,
                        action="sbi_manual_unpair", action_type="human",
                        entity_type="sbi", entity_id=pair_id,
                        detail=json.dumps({"bank_key": bk, "data_key": dk})))
    except Exception:
        pass
    db.commit()
    return {"deleted": pair_id}


@router.get("/manual-match/pair")
def list_manual_pairs(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """List two-sided manual pairs touching the date range (either side)."""
    q = db.query(SBIManualPair)
    if date_from:
        q = q.filter(or_(SBIManualPair.bank_date >= date_from, SBIManualPair.data_date >= date_from))
    if date_to:
        q = q.filter(or_(SBIManualPair.bank_date <= date_to, SBIManualPair.data_date <= date_to))
    rows = q.order_by(SBIManualPair.created_at.desc()).all()
    return {"rows": [{"id": m.id, "bank_date": m.bank_date, "data_date": m.data_date,
                      "bank_source": m.bank_source, "data_source": m.data_source,
                      "bank_amount": m.bank_amount, "data_amount": m.data_amount,
                      "amount_diff": round(abs(float(m.bank_amount or 0) - float(m.data_amount or 0)), 2),
                      "remark": m.remark, "by": m.created_by, "at": str(m.created_at)} for m in rows],
            "count": len(rows)}


# ── SRC disposition (overlay; parity with core-ledger /recon/assign-src) ───────
class SBISRCIn(BaseModel):
    process: str       # p01 | p02 | p03 | p04
    result_id: str
    src_code: str
    src_note: str = ""


@router.post("/assign-src")
def assign_src(
    body: SBISRCIn,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("src_assign")),
):
    """Tag an SBI result row with an SRC code + note. Persists across re-runs (overlay)."""
    from routes.recon import is_valid_src_code
    p = (body.process or "").lower()
    model = {"p01": SBIP01Result, "p02": SBIP02Result, "p03": SBIP03Result, "p04": SBIP04Result}.get(p)
    if not model:
        raise HTTPException(status_code=400, detail="process must be one of p01..p04")
    if not is_valid_src_code(db, body.src_code):
        raise HTTPException(status_code=400, detail=f"Invalid SRC code: {body.src_code}")
    row = db.query(model).filter(model.id == body.result_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Result row not found")
    row_dict = {c.name: getattr(row, c.name) for c in model.__table__.columns}
    key = _src_key(p, row_dict)
    if not key:
        raise HTTPException(status_code=400, detail="Row has no stable key to tag")
    existing = db.query(SBISrcAssignment).filter(
        SBISrcAssignment.recon_date == row.recon_date,
        SBISrcAssignment.process == p,
        SBISrcAssignment.match_key == key,
    ).first()
    if existing:
        existing.src_code = body.src_code
        existing.src_note = (body.src_note or "")[:500]
        existing.created_by = current_user.username
        sa = existing
    else:
        sa = SBISrcAssignment(recon_date=row.recon_date, process=p, match_key=key,
                              src_code=body.src_code, src_note=(body.src_note or "")[:500],
                              created_by=current_user.username)
        db.add(sa)
    try:
        db.add(AuditLog(user_id=current_user.id, username=current_user.username,
                        action="sbi_assign_src", action_type="human",
                        entity_type="sbi", entity_id=sa.id,
                        detail=json.dumps({"process": p, "recon_date": row.recon_date,
                                           "match_key": key, "src_code": body.src_code,
                                           "src_note": body.src_note})))
    except Exception:
        pass
    db.commit()
    return {"id": sa.id, "src_code": body.src_code, "match_key": key}


@router.delete("/assign-src/{sa_id}")
def delete_src_assignment(
    sa_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("src_assign")),
):
    """Remove an SBI SRC tag (revert the row to untagged)."""
    sa = db.query(SBISrcAssignment).filter(SBISrcAssignment.id == sa_id).first()
    if not sa:
        raise HTTPException(status_code=404, detail="SRC assignment not found")
    db.delete(sa)
    try:
        db.add(AuditLog(user_id=current_user.id, username=current_user.username,
                        action="sbi_unassign_src", action_type="human",
                        entity_type="sbi", entity_id=sa_id,
                        detail=json.dumps({"process": sa.process, "match_key": sa.match_key})))
    except Exception:
        pass
    db.commit()
    return {"deleted": sa_id}


class SBIBulkSRCItem(BaseModel):
    process: str        # p01 | p02 | p03 | p04
    result_id: str


class SBIBulkSRCIn(BaseModel):
    items: List[SBIBulkSRCItem]
    src_code: str
    src_note: str = ""


@router.post("/assign-src-bulk")
def assign_src_bulk(
    body: SBIBulkSRCIn,
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("src_assign")),
):
    """Tag MANY SBI result rows with one SRC code + note in a single shot (overlay,
    parity with core /recon/assign-src-bulk). Each item is {process, result_id};
    rows that can't be keyed/found are skipped and counted, not fatal."""
    from routes.recon import is_valid_src_code
    if not is_valid_src_code(db, body.src_code):
        raise HTTPException(status_code=400, detail=f"Invalid SRC code: {body.src_code}")
    if not body.items:
        raise HTTPException(status_code=400, detail="No rows provided")
    models = {"p01": SBIP01Result, "p02": SBIP02Result, "p03": SBIP03Result, "p04": SBIP04Result}
    note = (body.src_note or "")[:500]
    updated = skipped = 0
    seen = set()   # two unified rows can map to one P0x result — tag it once
    for it in body.items:
        p = (it.process or "").lower()
        if (p, it.result_id) in seen:
            continue
        seen.add((p, it.result_id))
        model = models.get(p)
        if not model:
            skipped += 1; continue
        row = db.query(model).filter(model.id == it.result_id).first()
        if not row:
            skipped += 1; continue
        row_dict = {c.name: getattr(row, c.name) for c in model.__table__.columns}
        key = _src_key(p, row_dict)
        if not key:
            skipped += 1; continue
        existing = db.query(SBISrcAssignment).filter(
            SBISrcAssignment.recon_date == row.recon_date,
            SBISrcAssignment.process == p,
            SBISrcAssignment.match_key == key,
        ).first()
        if existing:
            existing.src_code = body.src_code
            existing.src_note = note
            existing.created_by = current_user.username
        else:
            db.add(SBISrcAssignment(recon_date=row.recon_date, process=p, match_key=key,
                                    src_code=body.src_code, src_note=note,
                                    created_by=current_user.username))
        updated += 1
    try:
        db.add(AuditLog(user_id=current_user.id, username=current_user.username,
                        action="sbi_assign_src_bulk", action_type="human",
                        entity_type="sbi", entity_id=None,
                        detail=json.dumps({"src_code": body.src_code, "updated": updated, "skipped": skipped})))
    except Exception:
        pass
    db.commit()
    return {"updated": updated, "skipped": skipped, "src_code": body.src_code}


@router.get("/summary")
def get_summary(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Dashboard roll-up for SBI Kiosk. Data-presence counts + an overall recon
    health rate across the four processes (P01 settlement, P02 bank↔txn,
    P03 CSP↔txn↔bank, P04 wallet balance). Date-range aware on the result tables."""
    from sqlalchemy import func as _f
    GOOD = {"credited", "matched", "reconciled", "ok", "none", "no_action",
            "balanced", "success", "settled", "no action"}

    counts = {
        "bank_statement":  db.query(SBIBankTransaction).count(),
        "txn_reports":     db.query(SBITxnReport).count(),
        "ko_limits":       db.query(SBIKOLimits).count(),
        "ko_cash_holding": db.query(SBIKOCashHolding).count(),
        "limit_failures":  db.query(SBILimitFailure).count(),
        "csp_master":      db.query(SBICSPMaster).count(),
    }

    def _proc(model, status_col):
        q = db.query(status_col, _f.count())
        if date_from: q = q.filter(model.recon_date >= date_from)
        if date_to:   q = q.filter(model.recon_date <= date_to)
        by = {(s or "").strip().lower(): c for s, c in q.group_by(status_col).all()}
        total   = sum(by.values())
        matched = sum(c for s, c in by.items() if s in GOOD)
        return {"total": total, "matched": matched,
                "exceptions": total - matched, "by_status": by}

    p01 = _proc(SBIP01Result, SBIP01Result.status)
    p02 = _proc(SBIP02Result, SBIP02Result.match_status)
    p03 = _proc(SBIP03Result, SBIP03Result.match_status)
    p04 = _proc(SBIP04Result, SBIP04Result.action_required)

    tot     = p01["total"] + p02["total"] + p03["total"] + p04["total"]
    matched = p01["matched"] + p02["matched"] + p03["matched"] + p04["matched"]
    return {
        "counts":      counts,
        "has_data":    any(counts.values()),
        "processes":   {"p01": p01, "p02": p02, "p03": p03, "p04": p04},
        "total":       tot,
        "matched":     matched,
        "exceptions":  tot - matched,
        "match_rate":  round(matched / tot * 100, 1) if tot else None,
    }


@router.get("/upload-status")
def get_upload_status(
    upload_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Show count of uploaded records per data type for the given date."""
    today = upload_date or str(datetime.date.today())
    return {
        "upload_date": today,
        "bank_statement":  db.query(SBIBankTransaction).filter(SBIBankTransaction.upload_date == today).count(),
        "txn_reports":     db.query(SBITxnReport).filter(SBITxnReport.upload_date == today).count(),
        "ko_limits":       db.query(SBIKOLimits).filter(SBIKOLimits.upload_date == today).count(),
        "ko_cash_holding": db.query(SBIKOCashHolding).filter(SBIKOCashHolding.upload_date == today).count(),
        "limit_failures":  db.query(SBILimitFailure).filter(SBILimitFailure.upload_date == today).count(),
        "csp_master":      db.query(SBICSPMaster).filter(SBICSPMaster.upload_date == today).count(),
    }


@router.delete("/clear")
def clear_sbi_data(
    table: Optional[str] = Query(None, description="bank|txn|limits|cash|failures|csp|p01|p02|p03|p04|all"),
    upload_date: Optional[str] = None,
    recon_date: Optional[str] = None,    # business date (preferred over upload_date)
    dry_run: bool = False,               # preview: exact counts, deletes NOTHING
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("clear_data")),
):
    """Clear SBI data per table — recycle-binned, audited, business-date-aware.

    Rebuilt (2026-07-23): was a HARD delete with only an upload_date filter and the weak
    'upload' permission, and it left P0x result rows claiming Matched after their source
    rows were gone (unrecomputable — the wipe-guard rightly refuses an empty date). Now:
    every deletion goes through the recycle bin (one batch, restorable), clearing bank/txn
    sources also clears the affected business dates' results, and dry_run gives the UI an
    exact-count preview."""
    from core import recycle_bin
    tables = {
        "bank": SBIBankTransaction, "txn": SBITxnReport,
        "limits": SBIKOLimits, "cash": SBIKOCashHolding,
        "failures": SBILimitFailure, "csp": SBICSPMaster,
        "p01": SBIP01Result, "p02": SBIP02Result,
        "p03": SBIP03Result, "p04": SBIP04Result,
    }
    date_col = {"bank": "txn_date", "txn": "txn_date", "limits": "txn_date",
                "cash": "report_date", "failures": "txn_date",
                "p01": "recon_date", "p02": "recon_date", "p03": "recon_date",
                "p04": "recon_date"}
    to_clear = tables if table in (None, "all") else {table: tables[table]} if table in tables else {}
    if not to_clear:
        raise HTTPException(status_code=400, detail=f"Unknown table: {table}")
    batch = generate_id()
    counts = {}
    stale_dates = set()      # business dates whose bank/txn sources are being removed
    for name, model in to_clear.items():
        q = db.query(model)
        if upload_date and hasattr(model, 'upload_date'):
            q = q.filter(model.upload_date == upload_date)
        if recon_date and name in date_col:
            q = q.filter(getattr(model, date_col[name]) == recon_date)
        if name in ("bank", "txn"):
            stale_dates.update(d for (d,) in q.with_entities(
                getattr(model, date_col[name])).distinct().all() if d)
        if dry_run:
            n = q.count()
        else:
            n = recycle_bin.soft_delete(
                db, q, model, module="sbi", user=current_user.username,
                filters={k: v for k, v in {"table": name, "upload_date": upload_date,
                                           "recon_date": recon_date}.items() if v},
                reason="sbi per-table clear", batch_id=batch)
        if n:
            counts[name] = n
    # Sources gone ⇒ those dates' results are fiction — clear them in the same batch.
    if stale_dates:
        for rname in ("p01", "p02", "p03", "p04"):
            if rname in to_clear and table not in (None, "all"):
                continue                  # explicitly selected → already handled above
            rm = tables[rname]
            q = db.query(rm).filter(rm.recon_date.in_(stale_dates))
            if dry_run:
                n = q.count()
            else:
                n = recycle_bin.soft_delete(
                    db, q, rm, module="sbi", user=current_user.username,
                    filters={"stale_after": sorted(stale_dates)[:10]},
                    reason="sbi clear (source cleared — results stale)", batch_id=batch)
            if n:
                counts[rname] = counts.get(rname, 0) + n
    if dry_run:
        return {"dry_run": True, "would_delete": counts,
                "total": sum(counts.values()),
                "stale_result_dates": sorted(stale_dates)}
    db.commit()
    _audit(db, current_user, "sbi_clear",
           {"table": table or "all", "upload_date": upload_date, "recon_date": recon_date,
            "deleted": counts, "recycle_batch_id": batch})
    return {"deleted": counts, "total": sum(counts.values()), "recycle_batch_id": batch}
