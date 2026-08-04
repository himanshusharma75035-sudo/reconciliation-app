"""
routes/evalue.py

E-Value (wallet load) reconciliation API.

Flow:
  1. Upload the internal dump (SMB_BANK_LOADIN)        → POST /api/evalue/upload-internal
  2. Select bank → account → upload its statement       → POST /api/evalue/upload-bank
  3. Run reconciliation for an account (or all)         → POST /api/evalue/run-recon
  4. View results / summary / export                    → GET  /api/evalue/...
"""

import io
import os
import json
import datetime
import logging

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from models.database import (
    get_db, generate_id, AuditLog,
    EvalueAccount, EvalueWalletLoad, EvalueBankTxn, Transaction,
)

# The date a wallet load RECONCILES on = its VALUE date (the bank credits an E-Value
# load on its value date), falling back to the transaction date only when value date
# is blank. Used for every load date filter / ordering; transaction_date stays stored
# and is still shown alongside. Mirrors evalue_engine.load_date() and the model's
# recon_date_effective property. (Finance ops / Rajendra — behavior-contract item 13.)
_LOAD_DATE = func.coalesce(func.nullif(EvalueWalletLoad.value_date, ""),
                           EvalueWalletLoad.transaction_date)
from core.auth import get_current_user, require_product_access
from core import evalue_engine as EV
from core import maker_checker
from pydantic import BaseModel
from typing import Optional, List

logger = logging.getLogger("eko_recon")
router = APIRouter(prefix="/api/evalue", tags=["evalue"],
                   dependencies=[Depends(require_product_access("evalue"))])

_TODAY = lambda: datetime.date.today().isoformat()


# ─── Accounts ─────────────────────────────────────────────────────────────────
@router.get("/accounts")
def list_accounts(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Return active accounts grouped by bank for the upload dropdowns."""
    rows = db.query(EvalueAccount).filter(EvalueAccount.is_active == True)\
             .order_by(EvalueAccount.bank_name, EvalueAccount.reco_acc_no).all()
    banks = {}
    for r in rows:
        banks.setdefault(r.bank_name, []).append({
            "reco_acc_no": r.reco_acc_no,
            "account_number": r.account_number,
        })
    return {
        "banks": [{"bank_name": b, "accounts": a} for b, a in banks.items()],
        "supported_parsers": sorted(EV.BANK_PARSERS.keys()),
    }


# ─── Account management (Configuration) ───────────────────────────────────────
class AccountIn(BaseModel):
    bank_name: str
    account_number: str
    reco_acc_no: str


@router.get("/accounts/all")
def accounts_all(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Flat list of all E-Value accounts (incl. inactive) for Configuration."""
    rows = db.query(EvalueAccount).order_by(EvalueAccount.bank_name, EvalueAccount.reco_acc_no).all()
    return [{"id": r.id, "bank_name": r.bank_name, "account_number": r.account_number,
             "reco_acc_no": r.reco_acc_no, "is_active": r.is_active} for r in rows]


@router.post("/accounts")
def add_account(body: AccountIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if db.query(EvalueAccount).filter(EvalueAccount.reco_acc_no == body.reco_acc_no.strip()).first():
        raise HTTPException(400, f"Account {body.reco_acc_no} already exists")
    a = EvalueAccount(bank_name=body.bank_name.strip().upper(),
                      account_number=body.account_number.strip(),
                      reco_acc_no=body.reco_acc_no.strip())
    db.add(a)
    _audit_simple(db, user, "evalue_account_add", {"reco_acc_no": a.reco_acc_no})
    db.commit()
    return {"message": "Account added", "reco_acc_no": a.reco_acc_no}


@router.patch("/accounts/{acct_id}/toggle")
def toggle_account(acct_id: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
    a = db.query(EvalueAccount).filter(EvalueAccount.id == acct_id).first()
    if not a:
        raise HTTPException(404, "Account not found")
    a.is_active = not a.is_active
    db.commit()
    return {"reco_acc_no": a.reco_acc_no, "is_active": a.is_active}


def _audit_simple(db, user, action, detail):
    db.add(AuditLog(id=generate_id(), user_id=getattr(user, "id", None),
                    username=getattr(user, "username", None), action=action, action_type="human",
                    entity_type="evalue_account", detail=json.dumps(detail, default=str)))


# ══════════════════════════════════════════════════════════════════════════════
#  SHARED INGEST CORE — the SINGLE copy of the replace/upsert logic, used by BOTH
#  the interactive upload routes AND the auto-pickup helpers, so the two ingest
#  paths can never diverge again (behavior-contract item 10). Neither commits nor
#  audits — the caller does.
# ══════════════════════════════════════════════════════════════════════════════
def _replace_bank_txns(db, acct, bank_name, reco_acc_no, rows, filename):
    """Per-date, preserve-matched replace + insert of E-Value bank rows. Refuses (422)
    if no row carries a transaction date instead of full-wiping the account. Keeps rows
    with a preserved disposition (EVX-/EVIBT-/EVMAN-/src_assigned) and skips re-inserting
    the file's copy of a kept row (Counter over a strong signature). Returns inserted."""
    file_dates = {r["txn_date"] for r in rows if r.get("txn_date")}
    if not file_dates:
        raise HTTPException(
            422,
            f"[FORMAT] Parsed {len(rows)} rows from '{filename}' but none had a readable "
            f"transaction date, so this file cannot safely replace prior data for {reco_acc_no}. "
            f"Check the statement format/columns and re-upload.",
        )
    _PRESERVED_PREFIXES = ("EVX-", "EVIBT-", "EVMAN-")
    def _preserved(b):
        return b.recon_status == "src_assigned" or (b.match_id or "").startswith(_PRESERVED_PREFIXES)
    existing = db.query(EvalueBankTxn).filter(
        EvalueBankTxn.reco_acc_no == reco_acc_no,
        EvalueBankTxn.txn_date.in_(file_dates)).all()
    kept = [b for b in existing if _preserved(b)]
    for b in existing:
        if not _preserved(b):
            db.delete(b)
    from collections import Counter as _Counter
    def _norm(s):
        return (s or "").strip()
    def _sig(txn_date, amount, dr_cr, ref_no, utr, atm_ref, description):
        return (txn_date or "", round(float(amount or 0), 2), _norm(dr_cr),
                _norm(ref_no), _norm(utr), _norm(atm_ref), _norm(description))
    kept_sigs = _Counter(_sig(b.txn_date, b.amount, b.dr_cr, b.ref_no, b.utr, b.atm_ref, b.description) for b in kept)
    upload_date = _TODAY()
    inserted = 0
    for r in rows:
        _s = _sig(r["txn_date"], r["amount"], r["dr_cr"], r.get("ref_no"), r.get("utr"),
                  r.get("atm_ref"), r.get("description"))
        if kept_sigs.get(_s, 0) > 0:
            kept_sigs[_s] -= 1
            continue
        db.add(EvalueBankTxn(
            id=generate_id(), upload_date=upload_date, bank_name=bank_name.strip().upper(),
            reco_acc_no=reco_acc_no, account_number=(acct.account_number if acct else ""),
            txn_date=r["txn_date"], value_date=r["value_date"], description=r["description"],
            dr_cr=r["dr_cr"], amount=r["amount"], balance=r.get("balance"),
            ref_no=r.get("ref_no"), utr=r.get("utr"), atm_ref=r.get("atm_ref"),
            branch=r.get("branch"), mobile=r.get("mobile"), channel=r.get("channel"),
        ))
        inserted += 1
    return inserted


def _upsert_wallet_loads(db, loads):
    """Upsert-per-eko_trxn_id insert of E-Value wallet loads. Rows with a blank
    eko_trxn_id are appended (they cannot be upserted) — both callers SHA-256-guard the
    file first, so a re-applied file can't double-append them. Returns (inserted, replaced)."""
    upload_date = _TODAY()
    _EV_PAIRED = ("matched_online", "matched_cash", "matched_manual", "wrong_amount",
                  "twice_credit", "interbank_matched")
    inserted, replaced = 0, 0
    for l in loads:
        eko = l.get("eko_trxn_id") or ""
        if eko:
            ex = db.query(EvalueWalletLoad).filter(EvalueWalletLoad.eko_trxn_id == eko).first()
            if ex:
                if ex.match_id:
                    # Replacing a MATCHED load orphans its bank leg. Manual / preserved
                    # dispositions (EVMAN-/EVX-/EVIBT-) never self-heal (re-runs skip
                    # them by design), so reset the surviving bank row(s) to unmatched;
                    # auto EV- legs also reset here and simply re-match on the next run.
                    db.query(EvalueBankTxn).filter(
                        EvalueBankTxn.match_id == ex.match_id,
                        EvalueBankTxn.recon_status.in_(_EV_PAIRED),
                    ).update({"recon_status": "unmatched_bank", "match_id": None,
                              "match_note": "load leg replaced by re-upload — auto-reverted"},
                             synchronize_session=False)
                db.delete(ex); replaced += 1
        db.add(EvalueWalletLoad(
            id=generate_id(), upload_date=upload_date,
            reco_acc_no=l["reco_acc_no"], eko_trxn_id=eko,
            account_trxn_id=l.get("account_trxn_id"), csp_code=l.get("csp_code"),
            merchant_name=l.get("merchant_name"), cell_number=l.get("cell_number"),
            amount=l.get("amount", 0), transaction_date=l.get("transaction_date"),
            value_date=l.get("value_date"), status=l.get("status"),
            typename=l.get("typename"), dr_cr=l.get("dr_cr"), mode=l.get("mode"),
            load_mode=EV.classify_load_mode(l),
            utr_number=l.get("utr_number"), tid_chequeno=l.get("tid_chequeno"),
            bank_ref=l.get("bank_ref"),
            cdm_txn_number=l.get("cdm_txn_number"), cdm_branch=l.get("cdm_branch"),
            branch_name=l.get("branch_name"), branch_code=l.get("branch_code"),
            remarks=l.get("remarks"), comments=l.get("comments"),
            provider_type=l.get("provider_type"),
        ))
        inserted += 1
    return inserted, replaced


# ─── Upload internal dump ─────────────────────────────────────────────────────
@router.post("/upload-internal")
async def upload_internal(
    file: UploadFile = File(...),
    force: bool = Form(False),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    raw = await file.read()
    from core.file_hash_guard import guard_duplicate_file
    # force= restores rows removed since the original upload. The old "internal can never
    # lose data" rationale is FALSE: the Clear Data screen bulk-deletes wallet loads (it
    # did, three times on 2026-07-22), after which the lingering hash blocked the
    # legitimate re-upload forever — the documented hash-outlives-its-data trap. The
    # upsert-per-eko_trxn_id makes a forced re-apply idempotent for id-carrying rows; the
    # one hazard (blank-eko rows append, so a replay would double them) is closed below
    # by skipping eko-less rows on a forced re-apply.
    guard_duplicate_file(db, "evalue", "internal", raw, file.filename, user, force=force)
    try:
        loads = EV.parse_internal_dump(raw, file.filename)
    except Exception as e:
        raise HTTPException(400, f"Could not parse internal dump: {e}")
    if not loads:
        raise HTTPException(400, "No wallet-load rows found in the dump.")
    skipped_ekoless = 0
    if force:
        n0 = len(loads)
        loads = [l for l in loads if l.get("eko_trxn_id")]
        skipped_ekoless = n0 - len(loads)
        if not loads:
            raise HTTPException(400, "All rows in this dump lack an eko_trxn_id — a forced "
                                     "re-apply would duplicate them, so nothing was ingested.")

    inserted, replaced = _upsert_wallet_loads(db, loads)
    db.add(AuditLog(id=generate_id(), user_id=getattr(user, "id", None),
                    username=getattr(user, "username", None), action="evalue_upload_internal",
                    action_type="human", entity_type="evalue_wallet_load",
                    detail=json.dumps({"file": file.filename, "rows": inserted, "replaced": replaced,
                                       "force": force, "skipped_ekoless": skipped_ekoless})))
    db.commit()
    from collections import Counter
    by_acct = Counter(l["reco_acc_no"] for l in loads)
    # Auto-recon every account this dump touched (never blocks the upload).
    auto_recon = _auto_recon_after_upload(db, sorted(by_acct))
    return {"message": "Internal dump ingested", "rows": inserted, "replaced": replaced,
            "accounts": len(by_acct), "top_accounts": by_acct.most_common(10),
            "auto_recon": auto_recon}


# ─── Upload bank statement ────────────────────────────────────────────────────
@router.post("/upload-bank")
async def upload_bank(
    file: UploadFile = File(...),
    bank_name: str = Form(...),
    reco_acc_no: str = Form(...),
    force: bool = Form(False),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    acct = db.query(EvalueAccount).filter(EvalueAccount.reco_acc_no == reco_acc_no).first()
    if not acct:
        raise HTTPException(404, f"Unknown account {reco_acc_no}")
    raw = await file.read()
    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "evalue", f"bank:{reco_acc_no}", raw, file.filename, user, force=force)
    try:
        rows = EV.parse_bank_statement(bank_name, raw, file.filename)
    except Exception as e:
        raise HTTPException(400, f"Could not parse {bank_name} statement: {e}")
    if not rows:
        raise HTTPException(400, "No transactions found in the statement.")

    # Per-date, preserve-matched replace — the SINGLE shared copy (see _replace_bank_txns).
    _replace_bank_txns(db, acct, bank_name, reco_acc_no, rows, file.filename)
    db.add(AuditLog(id=generate_id(), user_id=getattr(user, "id", None),
                    username=getattr(user, "username", None), action="evalue_upload_bank",
                    action_type="human", entity_type="evalue_bank_txn",
                    detail=json.dumps({"bank": bank_name, "account": reco_acc_no,
                                       "file": file.filename, "rows": len(rows)})))
    db.commit()

    # ── Funds-position: EOD balance snapshots for this account (never blocks) ──
    try:
        from core.funds import record_snapshots
        record_snapshots(db, "evalue", bank_name.strip().lower(), reco_acc_no,
                         [{"date": (r.get("txn_date") or ""), "balance": r.get("balance"),
                           "dr": (r["amount"] or 0) if r["dr_cr"] == "DR" else 0.0,
                           "cr": (r["amount"] or 0) if r["dr_cr"] == "CR" else 0.0}
                          for r in rows],
                         uploaded_by=getattr(user, "username", None))
    except Exception:
        pass

    cr = sum(1 for r in rows if r["dr_cr"] == "CR")
    # Auto-recon this account (never blocks the upload).
    auto_recon = _auto_recon_after_upload(db, [reco_acc_no])
    return {"message": "Bank statement ingested", "bank": bank_name, "account": reco_acc_no,
            "rows": len(rows), "credits": cr, "debits": len(rows) - cr,
            "auto_recon": auto_recon}


# ─── Run reconciliation ───────────────────────────────────────────────────────
def _auto_recon_after_upload(db: Session, accounts) -> dict:
    """Auto-recon after an E-Value upload — parity with the core products' post-upload
    chain and the kiosk auto-run: per-account recon for every touched account, then
    the global cross-account reference pass, with EVERY failure swallowed so an
    upload is never blocked (an account missing its counterpart file is normal).
    Manual (EVMAN-), interbank (EVIBT-), cross-account (EVX-) and SRC dispositions
    are preserved by _run_one's exclusion filter."""
    out = {}
    for a in accounts:
        try:
            r = _run_one(db, a)
            out[a] = r.get("error") or "ok"
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            out[a] = f"skipped: {str(e)[:120]}"
    try:
        x = _cross_account_reference_match(db)
        out["cross_account"] = x.get("cross_account_matched", 0)
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        out["cross_account"] = f"skipped: {str(e)[:120]}"
    return out


def _run_one(db: Session, reco_acc_no: str) -> dict:
    all_bank = db.query(EvalueBankTxn).filter(EvalueBankTxn.reco_acc_no == reco_acc_no).all()
    if not all_bank:
        return {"reco_acc_no": reco_acc_no, "error": "no bank statement uploaded"}
    # SRC-disposed rows are excluded from re-matching and left untouched — parity with
    # core run_reconciliation (which only feeds unmatched rows). This preserves a manual
    # SRC tag (src_code/src_note + status) across a recon re-run instead of clobbering it.
    # EVX- rows are cross-account matches (counterpart in ANOTHER account), EVIBT- are
    # interbank links (counterpart in the CORE ledger), EVMAN- are human decisions —
    # a per-account re-run must not dissolve any of them (one-sided orphans otherwise;
    # core parity: run_reconciliation never re-feeds already-matched rows).
    _PRESERVED = ("EVX-", "EVIBT-", "EVMAN-")

    def _keep(r):
        return r.recon_status != "src_assigned" and not (r.match_id or "").startswith(_PRESERVED)
    bank_objs = [b for b in all_bank if _keep(b)]
    load_objs = [l for l in db.query(EvalueWalletLoad).filter(
                    EvalueWalletLoad.reco_acc_no == reco_acc_no).all()
                 if _keep(l)]

    # Build dicts carrying a backref to the DB object
    bank_rows = []
    for b in bank_objs:
        bank_rows.append({"_db": b, "txn_date": b.txn_date, "value_date": b.value_date or "",
                          "description": b.description or "", "dr_cr": b.dr_cr or "",
                          "amount": float(b.amount or 0), "balance": b.balance,
                          "ref_no": b.ref_no or "", "utr": b.utr or "", "atm_ref": b.atm_ref or "",
                          "branch": b.branch or "", "mobile": b.mobile or "", "channel": b.channel or "OTHER"})
    loads = []
    for l in load_objs:
        loads.append({"_db": l, "reco_acc_no": l.reco_acc_no, "amount": float(l.amount or 0),
                      "transaction_date": l.transaction_date or "", "value_date": l.value_date or "",
                      "status": l.status or "",
                      "dr_cr": l.dr_cr or "CR", "utr_number": l.utr_number or "",
                      "tid_chequeno": l.tid_chequeno or "", "bank_ref": getattr(l, "bank_ref", "") or "",
                      "cdm_txn_number": l.cdm_txn_number or "",
                      "cdm_branch": l.cdm_branch or "", "branch_name": l.branch_name or "",
                      "branch_code": l.branch_code or "", "cell_number": l.cell_number or "",
                      "remarks": l.remarks or "", "comments": l.comments or "", "mode": l.mode or ""})

    bank_name = bank_objs[0].bank_name if bank_objs else ""
    res = EV.reconcile(bank_rows, loads, reco_acc_no, bank_name=bank_name)

    # Write back statuses to DB objects
    for d in res["bank_rows"]:
        b = d["_db"]
        b.recon_status = d.get("recon_status", "unmatched_bank")
        b.match_id = d.get("match_id", "")
        b.match_note = (d.get("match_note", "") or "")[:300]
    for d in res["loads"]:
        l = d["_db"]
        l.recon_status = d.get("recon_status", "unmatched_load")
        l.match_id = d.get("match_id", "")
        l.match_note = (d.get("match_note", "") or "")[:300]
        l.load_mode = d.get("_mode", l.load_mode)
    db.commit()
    return res["summary"]


@router.post("/run-recon")
def run_recon(
    reco_acc_no: str = Query(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    summary = _run_one(db, reco_acc_no)
    if summary.get("error"):
        raise HTTPException(400, summary["error"])
    # Global reference pass picks up credits whose load is tagged to another account.
    xacc = _cross_account_reference_match(db)
    summary.update(xacc)
    db.add(AuditLog(id=generate_id(), user_id=getattr(user, "id", None),
                    username=getattr(user, "username", None), action="evalue_run_recon",
                    action_type="human", entity_type="evalue", entity_id=reco_acc_no,
                    detail=json.dumps(summary, default=str)))
    db.commit()
    return summary


def _next_evx_number(db: Session) -> int:
    """Next free EVX sequence number = MAX over every EVX id already worn by EITHER
    side + 1. Never COUNT: E-Value uploads delete-and-replace rows, so a count
    shrinks below the highest issued number and recycles live match IDs (the
    EVX-00039..43 collision of 2026-07-02). Match IDs are audit references — they
    must never be reused."""
    import re as _re
    mx = 0
    ids = [r[0] for r in db.query(EvalueBankTxn.match_id)
           .filter(EvalueBankTxn.match_id.like("EVX-%")).distinct().all()]
    ids += [r[0] for r in db.query(EvalueWalletLoad.match_id)
            .filter(EvalueWalletLoad.match_id.like("EVX-%")).distinct().all()]
    for mid in ids:
        m = _re.fullmatch(r"EVX-(\d+)", mid or "")
        if m:
            mx = max(mx, int(m.group(1)))
    return mx + 1


def _cross_account_reference_match(db: Session) -> dict:
    """Global reference-based match across accounts (run after per-account recon).

    Per finance ops: the bank Particulars reference (e.g. HDFCR…796815) is the true
    transaction identity. Some loads are tagged to a different reco_acc_no than the
    statement they actually belong to. This pass pairs any still-unmatched bank
    credit with any still-unmatched load whose TID_ChequeNo / Is_Weekend_Loan
    shares the same RRN — even across accounts — and FLAGS the account mismatch in
    the match note so it stays auditable.
    """
    credits = db.query(EvalueBankTxn).filter(
        EvalueBankTxn.recon_status == "unmatched_bank",
        EvalueBankTxn.dr_cr == "CR",
    ).all()
    loads = db.query(EvalueWalletLoad).filter(
        EvalueWalletLoad.recon_status == "unmatched_load",
    ).all()
    if not credits or not loads:
        return {"cross_account_matched": 0}

    # Index loads by each long RRN found in their reference fields
    by_run = {}
    for l in loads:
        for run in EV._ref_digits(l.tid_chequeno, l.bank_ref, l.utr_number):
            by_run.setdefault(run, []).append(l)

    used = set()
    n = 0; flagged = 0
    base = _next_evx_number(db) - 1
    for b in credits:
        runs = EV._ref_digits(b.utr, b.description, b.ref_no)
        cand = None
        for run in runs:
            for l in by_run.get(run, []):
                if l.id in used:
                    continue
                cand = l; break
            if cand:
                break
        if not cand:
            continue
        # amount AND date must match EXACTLY — full parity with the per-account passes
        # (E-Value tolerance is exact, behavior-contract item 7; Rajendra 2026-07-14).
        # A non-exact amount OR a differing date is left OPEN, never recovered here —
        # otherwise a sub-₹1 mismatch the per-account pass rejected would slip through
        # as a clean matched_online with no flag.
        if round(float(b.amount or 0), 2) != round(float(cand.amount or 0), 2):
            continue
        # Internal side reconciles on the VALUE date (fallback transaction date) —
        # parity with the per-account passes' load_date() (Rajendra, 2026-07-15).
        if (b.txn_date or "") != (cand.value_date or cand.transaction_date or ""):
            continue
        used.add(cand.id)
        n += 1
        base += 1
        mid = f"EVX-{base:05d}"
        same = (cand.reco_acc_no == b.reco_acc_no)
        note = "cross-account reference" if not same else "reference"
        if not same:
            flagged += 1
            note = f"⚠ account mismatch — load acct {cand.reco_acc_no} vs bank {b.reco_acc_no} (matched on reference)"
        b.recon_status = "matched_online"; b.match_id = mid; b.match_note = note[:300]
        cand.recon_status = "matched_online"; cand.match_id = mid; cand.match_note = note[:300]
    if n:
        db.commit()
    return {"cross_account_matched": n, "account_mismatch_flagged": flagged}


@router.post("/run-recon-all")
def run_recon_all(db: Session = Depends(get_db), user=Depends(get_current_user)):
    accts = [r[0] for r in db.query(EvalueBankTxn.reco_acc_no).distinct().all()]
    out = []
    for a in accts:
        try:
            out.append(_run_one(db, a))
        except Exception as e:
            out.append({"reco_acc_no": a, "error": str(e)})
    xacc = _cross_account_reference_match(db)
    return {"accounts": len(out), "results": out, **xacc}


# ─── Summary & results ────────────────────────────────────────────────────────
def _ev_date(q, col, date_from, date_to):
    if date_from: q = q.filter(col >= date_from)
    if date_to:   q = q.filter(col <= date_to)
    return q


@router.get("/summary")
def summary(date_from: str = Query(""), date_to: str = Query(""),
            db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Per-account roll-up of the latest recon state (SQL aggregates — no full-table loads)."""
    from sqlalchemy import func
    from core.dash_cache import dash_key, dash_get, dash_put
    _ck = dash_key("evalue_summary", db, date_from=date_from, date_to=date_to)
    _hit = dash_get(_ck)
    if _hit is not None:
        return _hit
    STAT = ["matched_online", "matched_cash", "matched_manual", "interbank_matched",
            "twice_credit", "wrong_amount", "unmatched_bank", "transaction_fee", "bank_debit",
            "src_assigned"]

    accts = [r[0] for r in db.query(EvalueBankTxn.reco_acc_no).distinct().all()]

    sq = db.query(EvalueBankTxn.reco_acc_no, EvalueBankTxn.recon_status, func.count(EvalueBankTxn.id))
    sq = _ev_date(sq, EvalueBankTxn.txn_date, date_from, date_to)
    status_counts = sq.group_by(EvalueBankTxn.reco_acc_no, EvalueBankTxn.recon_status).all()

    cq = db.query(EvalueBankTxn.reco_acc_no, func.count(EvalueBankTxn.id)).filter(EvalueBankTxn.dr_cr == "CR")
    cq = _ev_date(cq, EvalueBankTxn.txn_date, date_from, date_to)
    credit_counts = dict(cq.group_by(EvalueBankTxn.reco_acc_no).all())

    tq = db.query(EvalueBankTxn.reco_acc_no, func.count(EvalueBankTxn.id))
    tq = _ev_date(tq, EvalueBankTxn.txn_date, date_from, date_to)
    total_counts = dict(tq.group_by(EvalueBankTxn.reco_acc_no).all())

    lq = db.query(EvalueWalletLoad.reco_acc_no, func.count(EvalueWalletLoad.id))
    lq = _ev_date(lq, _LOAD_DATE, date_from, date_to)
    load_counts = dict(lq.group_by(EvalueWalletLoad.reco_acc_no).all())

    uq = db.query(EvalueWalletLoad.reco_acc_no, func.count(EvalueWalletLoad.id)) \
           .filter(EvalueWalletLoad.recon_status == "unmatched_load")
    uq = _ev_date(uq, _LOAD_DATE, date_from, date_to)
    unmatched_load_counts = dict(uq.group_by(EvalueWalletLoad.reco_acc_no).all())

    by_acct = {}
    for (a, s, c) in status_counts:
        by_acct.setdefault(a, {})[s] = c

    rows = []
    for a in accts:
        c = {s: by_acct.get(a, {}).get(s, 0) for s in STAT}
        credits = credit_counts.get(a, 0)
        matched = c["matched_online"] + c["matched_cash"] + c["matched_manual"] + c["interbank_matched"]
        rows.append({
            "reco_acc_no": a, "bank_rows": total_counts.get(a, 0), "bank_credits": credits,
            "loads": load_counts.get(a, 0),
            "unmatched_load": unmatched_load_counts.get(a, 0),
            **c, "match_rate": round(matched / (credits or 1) * 100, 1),
        })
    rows.sort(key=lambda r: r["reco_acc_no"])
    return dash_put(_ck, {"accounts": rows})


@router.get("/results")
def results(
    reco_acc_no: str = Query(...),
    side: str = Query("bank", pattern="^(bank|load)$"),
    status: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if side == "bank":
        q = db.query(EvalueBankTxn).filter(EvalueBankTxn.reco_acc_no == reco_acc_no)
        if status:
            q = q.filter(EvalueBankTxn.recon_status == status)
        q = _ev_date(q, EvalueBankTxn.txn_date, date_from, date_to)
        rows = q.order_by(EvalueBankTxn.txn_date, EvalueBankTxn.amount).all()
        return [{
            "id": b.id,
            "txn_date": b.txn_date, "description": b.description, "dr_cr": b.dr_cr,
            "amount": b.amount, "channel": b.channel, "utr": b.utr, "atm_ref": b.atm_ref,
            "branch": b.branch, "mobile": b.mobile, "recon_status": b.recon_status,
            "match_id": b.match_id, "match_note": b.match_note,
            "src_code": b.src_code, "src_note": b.src_note,
        } for b in rows]
    else:
        q = db.query(EvalueWalletLoad).filter(EvalueWalletLoad.reco_acc_no == reco_acc_no)
        if status:
            q = q.filter(EvalueWalletLoad.recon_status == status)
        q = _ev_date(q, _LOAD_DATE, date_from, date_to)
        rows = q.order_by(_LOAD_DATE, EvalueWalletLoad.amount).all()
        return [{
            "id": l.id,
            # recon_date = the load's VALUE date (fallback txn) — the date it reconciles
            # on. transaction_date kept for reference. (behavior-contract item 13)
            "recon_date": l.recon_date_effective,
            "transaction_date": l.transaction_date, "value_date": l.value_date,
            "csp_code": l.csp_code,
            "merchant_name": l.merchant_name, "cell_number": l.cell_number,
            "amount": l.amount, "load_mode": l.load_mode, "utr_number": l.utr_number,
            "eko_trxn_id": l.eko_trxn_id, "recon_status": l.recon_status,
            "match_id": l.match_id, "match_note": l.match_note,
            "src_code": l.src_code, "src_note": l.src_note,
        } for l in rows]


# ─── Export ───────────────────────────────────────────────────────────────────
@router.get("/export")
def export(
    reco_acc_no: str = Query(...),
    date_from: str = Query(""),
    date_to: str = Query(""),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    bank = _ev_date(db.query(EvalueBankTxn).filter(EvalueBankTxn.reco_acc_no == reco_acc_no),
                    EvalueBankTxn.txn_date, date_from, date_to).all()
    loads = _ev_date(db.query(EvalueWalletLoad).filter(EvalueWalletLoad.reco_acc_no == reco_acc_no),
                     _LOAD_DATE, date_from, date_to).all()
    bank_df = pd.DataFrame([{
        "Txn Date": b.txn_date, "Description": b.description, "Dr/Cr": b.dr_cr,
        "Amount": b.amount, "Channel": b.channel, "UTR": b.utr, "ATM/CDM Ref": b.atm_ref,
        "Branch": b.branch, "Mobile": b.mobile, "Recon Status": b.recon_status,
        "Match ID": b.match_id, "Note": b.match_note,
    } for b in bank])
    load_df = pd.DataFrame([{
        "Value Date": l.value_date, "Txn Date": l.transaction_date, "CSP Code": l.csp_code,
        "Merchant": l.merchant_name,
        "Mobile": l.cell_number, "Amount": l.amount, "Mode": l.load_mode,
        "UTR": l.utr_number, "Eko TID": l.eko_trxn_id, "Recon Status": l.recon_status,
        "Match ID": l.match_id, "Note": l.match_note,
    } for l in loads])

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        (bank_df if not bank_df.empty else pd.DataFrame({"info": ["no bank rows"]})).to_excel(
            w, index=False, sheet_name="Bank Statement")
        (load_df if not load_df.empty else pd.DataFrame({"info": ["no loads"]})).to_excel(
            w, index=False, sheet_name="Wallet Loads")
    out.seek(0)
    fn = f"evalue_recon_{reco_acc_no}_{_TODAY()}.xlsx"
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ══════════════════════════════════════════════════════════════════════════════
#  MANUAL RECONCILIATION
# ══════════════════════════════════════════════════════════════════════════════
EV_OVERRIDE_STATUSES = [
    "written_off", "twice_credit", "wrong_amount", "adhoc_settlement",
    "under_review", "ignore", "unmatched_bank", "unmatched_load",
]


def _ev_audit(db, user, action, detail, entity_id=None, prev=None):
    db.add(AuditLog(
        id=generate_id(), user_id=getattr(user, "id", None),
        username=getattr(user, "username", None), action=action, action_type="human",
        entity_type="evalue", entity_id=entity_id,
        detail=json.dumps(detail, default=str),
        previous_state=json.dumps(prev, default=str) if prev else None))


def _ev_mid(prefix, reco):
    return f"{prefix}-{EV._norm_ref(reco) or 'X'}-{generate_id()[:6].upper()}"


class ManualMatchIn(BaseModel):
    bank_txn_id: str
    load_id: str


@router.post("/manual-match")
def manual_match(body: ManualMatchIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Manually pair a bank credit with a wallet load (same or cross account).
    Amount guard (parity with core manual_match): a pair whose amounts differ by
    more than ₹1 is linked but flagged 'wrong_amount' — never a clean manual match."""
    b = db.query(EvalueBankTxn).filter(EvalueBankTxn.id == body.bank_txn_id).first()
    l = db.query(EvalueWalletLoad).filter(EvalueWalletLoad.id == body.load_id).first()
    if not b or not l:
        raise HTTPException(404, "Bank transaction or wallet load not found")
    cross = (b.reco_acc_no != l.reco_acc_no)
    diff = round(abs(float(b.amount or 0) - float(l.amount or 0)), 2)
    status = "matched_manual" if diff <= 1.0 else "wrong_amount"
    queued = maker_checker.intercept(
        db, user, "evalue_manual_match",
        payload={"bank_txn_id": body.bank_txn_id, "load_id": body.load_id},
        summary=f"E-Value manual match: bank {body.bank_txn_id} <-> load {body.load_id}",
        partner="evalue")
    if queued:
        return queued
    mid = _ev_mid("EVMAN", b.reco_acc_no)
    note = "manual match" + (f" — cross-account (bank {b.reco_acc_no} / load {l.reco_acc_no})" if cross else "")
    if status == "wrong_amount":
        note = f"manual match; bank {b.amount} vs load {l.amount} (Δ {diff})" + \
               (f" — cross-account (bank {b.reco_acc_no} / load {l.reco_acc_no})" if cross else "")
    for row in (b, l):
        row.recon_status = status; row.match_id = mid; row.match_note = note[:300]
    _ev_audit(db, user, "evalue_manual_match",
              {"bank": b.id, "load": l.id, "match_id": mid, "cross_account": cross,
               "amount": b.amount, "status": status, "amount_diff": diff})
    db.commit()
    return {"message": ("Matched" if status == "matched_manual"
                        else f"Linked but flagged wrong_amount — amounts differ by ₹{diff}"),
            "match_id": mid, "cross_account": cross, "status": status}


@router.post("/unmatch")
def unmatch(match_id: str = Query(None), bank_txn_id: str = Query(None),
            load_id: str = Query(None), db: Session = Depends(get_db),
            user=Depends(get_current_user)):
    """Break a match back to unmatched. Accepts a match_id, or a single bank/load id."""
    mid = match_id
    if not mid:
        ref = (db.query(EvalueBankTxn).filter(EvalueBankTxn.id == bank_txn_id).first()
               if bank_txn_id else
               db.query(EvalueWalletLoad).filter(EvalueWalletLoad.id == load_id).first())
        if not ref:
            raise HTTPException(404, "Row not found")
        mid = ref.match_id
    queued = maker_checker.intercept(
        db, user, "evalue_unmatch",
        payload={"match_id": match_id, "bank_txn_id": bank_txn_id, "load_id": load_id},
        summary=f"E-Value unmatch: {mid or bank_txn_id or load_id}",
        partner="evalue")
    if queued:
        return queued
    n = 0
    if mid:
        for b in db.query(EvalueBankTxn).filter(EvalueBankTxn.match_id == mid).all():
            b.recon_status = "unmatched_bank"; b.match_id = ""; b.match_note = "unmatched (manual)"; n += 1
        for l in db.query(EvalueWalletLoad).filter(EvalueWalletLoad.match_id == mid).all():
            l.recon_status = "unmatched_load"; l.match_id = ""; l.match_note = "unmatched (manual)"; n += 1
        # cross-product (interbank) counterpart in the main ledger
        for t in db.query(Transaction).filter(Transaction.match_id == mid).all():
            t.recon_status = "unmatched"; t.match_id = None; t.matched_with_id = None; n += 1
    else:
        # single unmatched row with no match
        row = (db.query(EvalueBankTxn).filter(EvalueBankTxn.id == bank_txn_id).first()
               if bank_txn_id else
               db.query(EvalueWalletLoad).filter(EvalueWalletLoad.id == load_id).first())
        if row:
            row.recon_status = "unmatched_bank" if bank_txn_id else "unmatched_load"; n = 1
    _ev_audit(db, user, "evalue_unmatch", {"match_id": mid, "rows_reset": n})
    db.commit()
    return {"message": "Unmatched", "rows_reset": n, "match_id": mid}


class OverrideIn(BaseModel):
    id: str
    side: str          # "bank" | "load"
    status: str
    note: str


@router.post("/override-status")
def override_status(body: OverrideIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Manually set a row's status with a mandatory remark (reason tag)."""
    _note = (body.note or "").strip()
    if len(_note) < 10:
        raise HTTPException(422, "A remark/reason of at least 10 characters is required")
    if body.status not in EV_OVERRIDE_STATUSES:
        raise HTTPException(400, f"Invalid status. Allowed: {EV_OVERRIDE_STATUSES}")
    Model = EvalueBankTxn if body.side == "bank" else EvalueWalletLoad
    row = db.query(Model).filter(Model.id == body.id).first()
    if not row:
        raise HTTPException(404, "Row not found")
    queued = maker_checker.intercept(
        db, user, "evalue_override",
        payload={"id": body.id, "side": body.side, "status": body.status, "note": body.note},
        summary=f"E-Value override {body.side} {body.id} -> {body.status}",
        partner="evalue")
    if queued:
        return queued
    prev = {"recon_status": row.recon_status, "match_id": row.match_id}
    row.prev_recon_status = row.recon_status
    row.recon_status = body.status
    row.match_note = f"override: {_note}"
    row.override_by = getattr(user, "username", None)
    row.override_at = datetime.datetime.utcnow()
    _ev_audit(db, user, "evalue_override_status",
              {"id": body.id, "side": body.side, "to": body.status, "note": body.note},
              entity_id=body.id, prev=prev)
    db.commit()
    return {"message": "Status updated", "id": body.id, "status": body.status}


# ── SRC disposition (parity with core-ledger /recon/assign-src) ────────────────
# Tags an unmatched E-Value row with a source code + note → recon_status='src_assigned'.
# Same access level as override-status (any authed user) so no permission regression.
# Preserved across recon re-runs by _run_one (src_assigned rows are not re-fed).
class EvalueSRCIn(BaseModel):
    id: str
    side: str          # "bank" | "load"
    src_code: str
    src_note: str = ""


class EvalueBulkSRCIn(BaseModel):
    ids: List[str]
    side: str
    src_code: str
    src_note: str = ""


@router.post("/assign-src")
def assign_src(body: EvalueSRCIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Assign an SRC code + note to one E-Value row (bank credit or wallet load)."""
    from routes.recon import is_valid_src_code
    if not is_valid_src_code(db, body.src_code):
        raise HTTPException(400, f"Invalid SRC code: {body.src_code}")
    Model = EvalueBankTxn if body.side == "bank" else EvalueWalletLoad
    row = db.query(Model).filter(Model.id == body.id).first()
    if not row:
        raise HTTPException(404, "Row not found")
    queued = maker_checker.intercept(
        db, user, "evalue_assign_src",
        payload={"id": body.id, "side": body.side, "src_code": body.src_code, "src_note": body.src_note},
        summary=f"E-Value SRC {body.src_code} -> {body.side} {body.id}",
        partner="evalue")
    if queued:
        return queued
    prev = {"recon_status": row.recon_status, "src_code": row.src_code, "match_id": row.match_id}
    row.prev_recon_status = row.recon_status
    row.recon_status = "src_assigned"
    row.src_code = body.src_code
    row.src_note = (body.src_note or "")[:500]
    row.override_by = getattr(user, "username", None)
    row.override_at = datetime.datetime.utcnow()
    _ev_audit(db, user, "evalue_assign_src",
              {"id": body.id, "side": body.side, "src_code": body.src_code, "src_note": body.src_note},
              entity_id=body.id, prev=prev)
    db.commit()
    return {"message": "SRC assigned", "id": body.id, "src_code": body.src_code}


@router.post("/assign-src-bulk")
def assign_src_bulk(body: EvalueBulkSRCIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Assign one SRC code + note to many E-Value rows in one shot."""
    from routes.recon import is_valid_src_code
    if not is_valid_src_code(db, body.src_code):
        raise HTTPException(400, f"Invalid SRC code: {body.src_code}")
    Model = EvalueBankTxn if body.side == "bank" else EvalueWalletLoad
    rows = db.query(Model).filter(Model.id.in_(body.ids or [])).all()
    queued = maker_checker.intercept(
        db, user, "evalue_assign_src_bulk",
        payload={"ids": body.ids, "side": body.side, "src_code": body.src_code, "src_note": body.src_note},
        summary=f"E-Value bulk SRC {body.src_code} -> {len(body.ids or [])} {body.side} rows",
        partner="evalue")
    if queued:
        return queued
    for row in rows:
        row.prev_recon_status = row.recon_status
        row.recon_status = "src_assigned"
        row.src_code = body.src_code
        row.src_note = (body.src_note or "")[:500]
        row.override_by = getattr(user, "username", None)
        row.override_at = datetime.datetime.utcnow()
    _ev_audit(db, user, "evalue_assign_src_bulk",
              {"side": body.side, "src_code": body.src_code, "count": len(rows)})
    db.commit()
    return {"updated": len(rows), "src_code": body.src_code}


class EvalueSRCRemoveIn(BaseModel):
    id: str
    side: str            # bank | internal (wallet load)


class EvalueBulkSRCRemoveIn(BaseModel):
    ids: list
    side: str


def _ev_open_default(side):
    return "unmatched_bank" if side == "bank" else "unmatched_load"


@router.post("/remove-src")
def remove_src(body: EvalueSRCRemoveIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Remove the SRC tag from one E-Value row and revert it to the status it held before the
    tag (falls back to unmatched_bank / unmatched_load for pre-existing rows)."""
    Model = EvalueBankTxn if body.side == "bank" else EvalueWalletLoad
    row = db.query(Model).filter(Model.id == body.id).first()
    if not row:
        raise HTTPException(404, "Row not found")
    if row.recon_status != "src_assigned":
        raise HTTPException(400, "Row is not SRC-tagged")
    queued = maker_checker.intercept(
        db, user, "evalue_remove_src",
        payload={"id": body.id, "side": body.side},
        summary=f"E-Value remove SRC -> {body.side} {body.id}", partner="evalue")
    if queued:
        return queued
    prev = row.prev_recon_status
    if not prev or prev == "src_assigned":
        prev = _ev_open_default(body.side)
    old_code = row.src_code
    row.recon_status = prev
    row.src_code = None
    row.src_note = None
    row.prev_recon_status = None
    row.override_by = getattr(user, "username", None)
    row.override_at = datetime.datetime.utcnow()
    _ev_audit(db, user, "evalue_remove_src",
              {"id": body.id, "side": body.side, "reverted_to": prev, "removed_src_code": old_code},
              entity_id=body.id)
    db.commit()
    return {"message": "SRC removed", "id": body.id, "reverted_to": prev}


@router.post("/remove-src-bulk")
def remove_src_bulk(body: EvalueBulkSRCRemoveIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Remove SRC tags from many E-Value rows; each reverts to its own prior status."""
    Model = EvalueBankTxn if body.side == "bank" else EvalueWalletLoad
    rows = db.query(Model).filter(Model.id.in_(body.ids or []),
                                  Model.recon_status == "src_assigned").all()
    queued = maker_checker.intercept(
        db, user, "evalue_remove_src_bulk",
        payload={"ids": body.ids, "side": body.side},
        summary=f"E-Value remove SRC -> {len(body.ids or [])} {body.side} rows", partner="evalue")
    if queued:
        return queued
    default = _ev_open_default(body.side)
    for row in rows:
        prev = row.prev_recon_status
        if not prev or prev == "src_assigned":
            prev = default
        row.recon_status = prev
        row.src_code = None
        row.src_note = None
        row.prev_recon_status = None
        row.override_by = getattr(user, "username", None)
        row.override_at = datetime.datetime.utcnow()
    _ev_audit(db, user, "evalue_remove_src_bulk", {"side": body.side, "count": len(rows)})
    db.commit()
    return {"reverted": len(rows), "side": body.side}


# ══════════════════════════════════════════════════════════════════════════════
#  CROSS-PRODUCT INTERBANK MATCHING
#  An E-Value bank entry can correspond to a transfer that appears as a credit /
#  fund-transfer in another product's ledger. Match on UTR (+ amount tolerance).
# ══════════════════════════════════════════════════════════════════════════════
_IBT_TXN_OPEN = ["unmatched", "fund_transfer", "settlement_credit"]
_EV_IBT_CANDIDATE = ["unmatched_bank", "bank_debit"]
_AMT_TOL = 1.0


def _txn_brief(t: Transaction) -> dict:
    return {"id": t.id, "partner": t.partner, "side": t.side, "eko_tid": t.eko_tid,
            "utr_number": t.utr_number, "tracking_number": t.tracking_number,
            "amount": t.amount, "dr_cr": t.dr_cr, "transaction_date": t.transaction_date,
            "recon_status": t.recon_status, "recon_date": t.recon_date}


@router.get("/interbank/candidates")
def interbank_candidates(bank_txn_id: str = Query(...), db: Session = Depends(get_db),
                         user=Depends(get_current_user)):
    """Candidate ledger entries (other products) for an E-Value bank entry, by UTR / amount."""
    b = db.query(EvalueBankTxn).filter(EvalueBankTxn.id == bank_txn_id).first()
    if not b:
        raise HTTPException(404, "Bank transaction not found")
    q = db.query(Transaction).filter(Transaction.recon_status.in_(_IBT_TXN_OPEN))
    cands = []
    for t in q.limit(5000).all():
        amt_ok = abs(float(t.amount or 0) - float(b.amount or 0)) <= _AMT_TOL
        utr_ok = bool(b.utr) and (EV._utr_match(b.utr, t.utr_number) or EV._utr_match(b.utr, t.tracking_number))
        if utr_ok or (amt_ok and b.txn_date and t.transaction_date and b.txn_date == (t.transaction_date or "")[:10]):
            d = _txn_brief(t); d["match_on"] = "UTR" if utr_ok else "amount+date"; d["amount_ok"] = amt_ok
            cands.append(d)
    cands.sort(key=lambda d: (d["match_on"] != "UTR", not d["amount_ok"]))
    return {"bank_txn": {"id": b.id, "amount": b.amount, "utr": b.utr, "txn_date": b.txn_date,
                         "description": b.description, "reco_acc_no": b.reco_acc_no},
            "candidates": cands[:100]}


def _do_interbank_link(db, b: EvalueBankTxn, t: Transaction, user, how: str):
    mid = _ev_mid("EVIBT", b.reco_acc_no)
    b.recon_status = "interbank_matched"; b.match_id = mid
    b.match_note = f"interbank: {t.partner}/{t.side} {t.eko_tid or ''} ({how})"
    t.recon_status = "interbank_matched"; t.match_id = mid
    _ev_audit(db, user, "evalue_interbank_match",
              {"evalue_bank": b.id, "txn": t.id, "partner": t.partner, "match_id": mid,
               "amount": b.amount, "how": how}, entity_id=b.id)
    return mid


class IbtManualIn(BaseModel):
    bank_txn_id: str
    transaction_id: str


@router.post("/interbank/manual-match")
def interbank_manual(body: IbtManualIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Amount guard (parity with the auto interbank scan, which requires UTR+amount):
    a manual interbank pair with amounts differing by more than ₹1 is linked but
    flagged wrong_amount / amount_mismatch instead of interbank_matched."""
    b = db.query(EvalueBankTxn).filter(EvalueBankTxn.id == body.bank_txn_id).first()
    t = db.query(Transaction).filter(Transaction.id == body.transaction_id).first()
    if not b or not t:
        raise HTTPException(404, "Bank txn or ledger transaction not found")
    diff = round(abs(float(b.amount or 0) - float(t.amount or 0)), 2)
    queued = maker_checker.intercept(
        db, user, "evalue_interbank_match",
        payload={"bank_txn_id": body.bank_txn_id, "transaction_id": body.transaction_id},
        summary=f"E-Value interbank match: bank {body.bank_txn_id} <-> txn {body.transaction_id}",
        partner="evalue")
    if queued:
        return queued
    mid = _do_interbank_link(db, b, t, user, "manual")
    if diff > 1.0:
        note = f"interbank manual; evalue {b.amount} vs {t.partner} {t.amount} (Δ {diff})"
        b.recon_status = "wrong_amount"; b.match_note = note[:300]
        t.recon_status = "amount_mismatch"    # core-ledger vocabulary for the same flag
    db.commit()
    return {"message": ("Interbank matched" if diff <= 1.0
                        else f"Linked but flagged — amounts differ by ₹{diff}"),
            "match_id": mid, "amount_diff": diff}


@router.post("/interbank/scan")
def interbank_scan(reco_acc_no: str = Query(None), db: Session = Depends(get_db),
                   user=Depends(get_current_user)):
    """Auto-match open E-Value bank entries to other-product ledger entries on UTR + amount."""
    bq = db.query(EvalueBankTxn).filter(EvalueBankTxn.recon_status.in_(_EV_IBT_CANDIDATE))
    if reco_acc_no:
        bq = bq.filter(EvalueBankTxn.reco_acc_no == reco_acc_no)
    bank_rows = [b for b in bq.all() if EV._norm_ref(b.utr)]
    ledger = db.query(Transaction).filter(Transaction.recon_status.in_(_IBT_TXN_OPEN)).all()
    used = set()
    matched = 0
    for b in bank_rows:
        for t in ledger:
            if t.id in used:
                continue
            if EV._utr_match(b.utr, t.utr_number) or EV._utr_match(b.utr, t.tracking_number):
                if abs(float(t.amount or 0) - float(b.amount or 0)) <= _AMT_TOL:
                    _do_interbank_link(db, b, t, user, "auto UTR+amount")
                    used.add(t.id); matched += 1
                    break
    db.commit()
    return {"message": "Interbank scan complete", "matched": matched,
            "scanned_bank_entries": len(bank_rows)}


@router.get("/interbank/list")
def interbank_list(db: Session = Depends(get_db), user=Depends(get_current_user)):
    bank = db.query(EvalueBankTxn).filter(EvalueBankTxn.recon_status == "interbank_matched").all()
    out = []
    for b in bank:
        t = db.query(Transaction).filter(Transaction.match_id == b.match_id).first()
        out.append({
            "match_id": b.match_id, "reco_acc_no": b.reco_acc_no, "txn_date": b.txn_date,
            "amount": b.amount, "utr": b.utr, "bank_txn_id": b.id, "note": b.match_note,
            "counterpart": _txn_brief(t) if t else None,
        })
    return {"matches": out, "count": len(out)}


# ══════════════════════════════════════════════════════════════════════════════
#  REPORTS  (filterable + Excel export)
# ══════════════════════════════════════════════════════════════════════════════
def _hdr_style(ws, fill="1E3A5F"):
    from openpyxl.styles import PatternFill, Font as XFont
    for col in ws.columns:
        m = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(m + 2, 48)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = XFont(bold=True, color="FFFFFF")


def _ev_filtered_bank(db, frm, to, bank, reco, statuses=None):
    q = db.query(EvalueBankTxn)
    if reco: q = q.filter(EvalueBankTxn.reco_acc_no == reco)
    if bank: q = q.filter(EvalueBankTxn.bank_name == bank.strip().upper())
    if frm:  q = q.filter(EvalueBankTxn.txn_date >= frm)
    if to:   q = q.filter(EvalueBankTxn.txn_date <= to)
    if statuses: q = q.filter(EvalueBankTxn.recon_status.in_(statuses))
    return q.order_by(EvalueBankTxn.reco_acc_no, EvalueBankTxn.txn_date).all()


def _ev_filtered_loads(db, frm, to, reco, statuses=None):
    q = db.query(EvalueWalletLoad)
    if reco: q = q.filter(EvalueWalletLoad.reco_acc_no == reco)
    if frm:  q = q.filter(_LOAD_DATE >= frm)   # VALUE date (fallback txn) — item 13
    if to:   q = q.filter(_LOAD_DATE <= to)
    if statuses: q = q.filter(EvalueWalletLoad.recon_status.in_(statuses))
    return q.order_by(EvalueWalletLoad.reco_acc_no, _LOAD_DATE).all()


def _bank_row(b):
    return {"Account": b.reco_acc_no, "Bank": b.bank_name, "Txn Date": b.txn_date,
            "Description": b.description, "Dr/Cr": b.dr_cr, "Amount": b.amount,
            "Channel": b.channel, "UTR": b.utr, "ATM/CDM Ref": b.atm_ref, "Branch": b.branch,
            "Mobile": b.mobile, "Recon Status": b.recon_status, "Match ID": b.match_id, "Note": b.match_note}


def _load_row(l):
    return {"Account": l.reco_acc_no, "Value Date": l.value_date, "Txn Date": l.transaction_date,
            "CSP Code": l.csp_code,
            "Merchant": l.merchant_name, "Mobile": l.cell_number, "Amount": l.amount,
            "Mode": l.load_mode, "UTR": l.utr_number, "Eko TID": l.eko_trxn_id,
            "Recon Status": l.recon_status, "Match ID": l.match_id, "Note": l.match_note}


REPORT_TYPES = {
    "summary":        "E-Value Summary",
    "matched":        "Matched Pairs",
    "unmatched_bank": "Unmatched Bank Credits",
    "unmatched_load": "Unmatched Wallet Loads",
    "twice_credit":   "Twice Credit",
    "wrong_amount":   "Wrong Amount Credits",
    "fees_debits":    "Transaction Fees & Bank Debits",
    "interbank":      "Interbank Matches",
    "exceptions":     "All Exceptions",
    "full":           "Full Classified Export",
}


@router.get("/report/types")
def report_types(user=Depends(get_current_user)):
    return {"types": [{"key": k, "label": v} for k, v in REPORT_TYPES.items()]}


@router.get("/report/summary")
def report_summary(frm: str = Query(None, alias="from"), to: str = Query(None),
                   bank: str = Query(None), reco_acc_no: str = Query(None),
                   db: Session = Depends(get_db), user=Depends(get_current_user)):
    """All-accounts + per-account roll-up with date/bank/account filters and grand totals."""
    bank_rows = _ev_filtered_bank(db, frm, to, bank, reco_acc_no)
    loads = _ev_filtered_loads(db, frm, to, reco_acc_no)
    STAT = ["matched_online", "matched_cash", "matched_manual", "interbank_matched",
            "twice_credit", "wrong_amount", "unmatched_bank", "transaction_fee", "bank_debit",
            "src_assigned"]
    from collections import defaultdict
    accts = sorted(set([b.reco_acc_no for b in bank_rows] + [l.reco_acc_no for l in loads]))
    rows = []
    for a in accts:
        bk = [b for b in bank_rows if b.reco_acc_no == a]
        ld = [l for l in loads if l.reco_acc_no == a]
        c = {s: sum(1 for b in bk if b.recon_status == s) for s in STAT}
        credits = sum(1 for b in bk if b.dr_cr == "CR")
        matched = c["matched_online"] + c["matched_cash"] + c["matched_manual"] + c["interbank_matched"]
        amt = sum(float(b.amount or 0) for b in bk if b.dr_cr == "CR")
        rows.append({"reco_acc_no": a, "bank_credits": credits, "credit_amount": round(amt, 2),
                     "loads": len(ld),
                     "unmatched_load": sum(1 for l in ld if l.recon_status == "unmatched_load"),
                     **c, "matched": matched, "match_rate": round(matched / (credits or 1) * 100, 1)})
    grand = {"bank_credits": sum(r["bank_credits"] for r in rows),
             "credit_amount": round(sum(r["credit_amount"] for r in rows), 2),
             "loads": sum(r["loads"] for r in rows),
             "matched": sum(r["matched"] for r in rows)}
    for s in STAT + ["unmatched_load"]:
        grand[s] = sum(r.get(s, 0) for r in rows)
    grand["match_rate"] = round(grand["matched"] / (grand["bank_credits"] or 1) * 100, 1)
    return {"accounts": rows, "grand": grand, "filters": {"from": frm, "to": to, "bank": bank, "reco_acc_no": reco_acc_no}}


@router.get("/report/export")
def report_export(report: str = Query("summary"), frm: str = Query(None, alias="from"),
                  to: str = Query(None), bank: str = Query(None), reco_acc_no: str = Query(None),
                  db: Session = Depends(get_db), user=Depends(get_current_user)):
    if report not in REPORT_TYPES:
        raise HTTPException(400, f"Invalid report. Valid: {list(REPORT_TYPES)}")
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        def sheet(name, rows, fill="1E3A5F"):
            df = pd.DataFrame(rows) if rows else pd.DataFrame({"info": ["no rows"]})
            df.to_excel(w, index=False, sheet_name=name[:31])
            _hdr_style(w.sheets[name[:31]], fill)

        if report == "summary":
            data = report_summary.__wrapped__ if hasattr(report_summary, "__wrapped__") else None
            s = report_summary(frm=frm, to=to, bank=bank, reco_acc_no=reco_acc_no, db=db, user=user)
            sheet("Summary", s["accounts"])
            sheet("Grand Total", [s["grand"]])
        elif report == "matched":
            bank_rows = _ev_filtered_bank(db, frm, to, bank, reco_acc_no,
                                          ["matched_online", "matched_cash", "matched_manual"])
            pairs = []
            for b in bank_rows:
                l = db.query(EvalueWalletLoad).filter(EvalueWalletLoad.match_id == b.match_id).first() if b.match_id else None
                pairs.append({"Account": b.reco_acc_no, "Match ID": b.match_id, "Status": b.recon_status,
                              "Bank Date": b.txn_date, "Bank Amount": b.amount, "Bank UTR": b.utr,
                              "Bank Channel": b.channel,
                              "Load Date": (l.recon_date_effective if l else ""), "Load Amount": l.amount if l else "",
                              "CSP": l.csp_code if l else "", "Load UTR": l.utr_number if l else "",
                              "Eko TID": l.eko_trxn_id if l else ""})
            sheet("Matched Pairs", pairs)
        elif report in ("unmatched_bank", "twice_credit", "wrong_amount"):
            sheet(REPORT_TYPES[report][:31], [_bank_row(b) for b in
                  _ev_filtered_bank(db, frm, to, bank, reco_acc_no, [report])], "B45309")
        elif report == "unmatched_load":
            sheet("Unmatched Loads", [_load_row(l) for l in
                  _ev_filtered_loads(db, frm, to, reco_acc_no, ["unmatched_load"])], "B45309")
        elif report == "fees_debits":
            sheet("Fees & Debits", [_bank_row(b) for b in
                  _ev_filtered_bank(db, frm, to, bank, reco_acc_no, ["transaction_fee", "bank_debit"])], "475569")
        elif report == "interbank":
            sheet("Interbank Matches", [_bank_row(b) for b in
                  _ev_filtered_bank(db, frm, to, bank, reco_acc_no, ["interbank_matched"])])
        elif report == "exceptions":
            sheet("Bank Exceptions", [_bank_row(b) for b in
                  _ev_filtered_bank(db, frm, to, bank, reco_acc_no,
                  ["twice_credit", "wrong_amount", "unmatched_bank"])], "B45309")
            sheet("Unmatched Loads", [_load_row(l) for l in
                  _ev_filtered_loads(db, frm, to, reco_acc_no, ["unmatched_load"])], "B45309")
        elif report == "full":
            sheet("Bank Statement", [_bank_row(b) for b in _ev_filtered_bank(db, frm, to, bank, reco_acc_no)])
            sheet("Wallet Loads", [_load_row(l) for l in _ev_filtered_loads(db, frm, to, reco_acc_no)])
    out.seek(0)
    fn = f"evalue_{report}_{_TODAY()}.xlsx"
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ══════════════════════════════════════════════════════════════════════════════
#  AGEING & ESCALATION
# ══════════════════════════════════════════════════════════════════════════════
def _age_days(date_str: str) -> int:
    try:
        d = datetime.date.fromisoformat((date_str or "")[:10])
        return (datetime.date.today() - d).days
    except Exception:
        return 0


def _bucket(days: int) -> str:
    if days <= 0:  return "current"
    if days == 1:  return "d1"
    if days <= 3:  return "d3"
    if days <= 7:  return "d7"
    return "d7plus"


@router.get("/ageing")
def ageing(reco_acc_no: str = Query(None), db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Ageing of unmatched E-Value items (bank credits + wallet loads) by bucket."""
    bq = db.query(EvalueBankTxn).filter(EvalueBankTxn.recon_status == "unmatched_bank")
    lq = db.query(EvalueWalletLoad).filter(EvalueWalletLoad.recon_status == "unmatched_load")
    if reco_acc_no:
        bq = bq.filter(EvalueBankTxn.reco_acc_no == reco_acc_no)
        lq = lq.filter(EvalueWalletLoad.reco_acc_no == reco_acc_no)
    buckets = ["current", "d1", "d3", "d7", "d7plus"]
    agg = {b: {"bank": 0, "load": 0, "bank_amount": 0.0, "load_amount": 0.0} for b in buckets}
    escalation = []
    for b in bq.all():
        days = _age_days(b.txn_date); bk = _bucket(days)
        agg[bk]["bank"] += 1; agg[bk]["bank_amount"] += float(b.amount or 0)
        if days > 7:
            escalation.append({"side": "bank", "id": b.id, "reco_acc_no": b.reco_acc_no,
                               "date": b.txn_date, "age_days": days, "amount": b.amount,
                               "utr": b.utr, "description": b.description})
    for l in lq.all():
        # age from the VALUE date (the date the load reconciles on), fallback txn — item 13
        days = _age_days(l.recon_date_effective); bk = _bucket(days)
        agg[bk]["load"] += 1; agg[bk]["load_amount"] += float(l.amount or 0)
        if days > 7:
            escalation.append({"side": "load", "id": l.id, "reco_acc_no": l.reco_acc_no,
                               "date": l.recon_date_effective, "age_days": days, "amount": l.amount,
                               "csp_code": l.csp_code, "utr": l.utr_number})
    escalation.sort(key=lambda x: -x["age_days"])
    return {"buckets": agg, "escalation_count": len(escalation), "escalation": escalation}


# ══════════════════════════════════════════════════════════════════════════════
#  BULK ACTIONS
# ══════════════════════════════════════════════════════════════════════════════
class BulkOverrideIn(BaseModel):
    ids: List[str]
    side: str                # "bank" | "load"
    status: str
    note: str


@router.post("/bulk-override")
def bulk_override(body: BulkOverrideIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if len((body.note or "").strip()) < 10:
        raise HTTPException(422, "A remark/reason of at least 10 characters is required")
    if body.status not in EV_OVERRIDE_STATUSES:
        raise HTTPException(400, f"Invalid status. Allowed: {EV_OVERRIDE_STATUSES}")
    Model = EvalueBankTxn if body.side == "bank" else EvalueWalletLoad
    queued = maker_checker.intercept(
        db, user, "evalue_bulk_override",
        payload={"ids": body.ids, "side": body.side, "status": body.status, "note": body.note},
        summary=f"E-Value bulk override {len(body.ids or [])} {body.side} rows -> {body.status}",
        partner="evalue")
    if queued:
        return queued
    n = 0
    for rid in body.ids:
        row = db.query(Model).filter(Model.id == rid).first()
        if not row:
            continue
        row.prev_recon_status = row.recon_status
        row.recon_status = body.status
        row.match_note = f"override: {body.note.strip()}"
        row.override_by = getattr(user, "username", None)
        row.override_at = datetime.datetime.utcnow()
        n += 1
    _ev_audit(db, user, "evalue_bulk_override",
              {"side": body.side, "status": body.status, "count": n, "note": body.note})
    db.commit()
    return {"message": "Bulk override applied", "updated": n}


@router.post("/bulk-unmatch")
def bulk_unmatch(match_ids: List[str] = None, db: Session = Depends(get_db), user=Depends(get_current_user)):
    match_ids = match_ids or []
    queued = maker_checker.intercept(
        db, user, "evalue_bulk_unmatch",
        payload={"match_ids": match_ids},
        summary=f"E-Value bulk unmatch: {len(match_ids)} match(es)",
        partner="evalue")
    if queued:
        return queued
    n = 0
    for mid in match_ids:
        for b in db.query(EvalueBankTxn).filter(EvalueBankTxn.match_id == mid).all():
            b.recon_status = "unmatched_bank"; b.match_id = ""; b.match_note = "unmatched (bulk)"; n += 1
        for l in db.query(EvalueWalletLoad).filter(EvalueWalletLoad.match_id == mid).all():
            l.recon_status = "unmatched_load"; l.match_id = ""; l.match_note = "unmatched (bulk)"; n += 1
        for t in db.query(Transaction).filter(Transaction.match_id == mid).all():
            t.recon_status = "unmatched"; t.match_id = None; t.matched_with_id = None; n += 1
    _ev_audit(db, user, "evalue_bulk_unmatch", {"match_ids": match_ids, "rows_reset": n})
    db.commit()
    return {"message": "Bulk unmatch complete", "rows_reset": n}


# ══════════════════════════════════════════════════════════════════════════════
#  TWICE-CREDIT RECOVERY WORKFLOW
# ══════════════════════════════════════════════════════════════════════════════
RECOVERY_STATES = ["recovery_pending", "recovered", "waived"]


class RecoveryIn(BaseModel):
    bank_txn_id: str
    recovery_status: str
    note: str = ""


@router.post("/recovery/mark")
def recovery_mark(body: RecoveryIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Track recovery of an excess (twice/wrong-amount) credit from the CSP."""
    if body.recovery_status not in RECOVERY_STATES:
        raise HTTPException(400, f"Invalid recovery_status. Allowed: {RECOVERY_STATES}")
    b = db.query(EvalueBankTxn).filter(EvalueBankTxn.id == body.bank_txn_id).first()
    if not b:
        raise HTTPException(404, "Bank transaction not found")
    b.recovery_status = body.recovery_status
    b.recovery_note = (body.note or "").strip()[:300]
    _ev_audit(db, user, "evalue_recovery_mark",
              {"bank": b.id, "recovery_status": body.recovery_status, "amount": b.amount},
              entity_id=b.id)
    db.commit()
    return {"message": "Recovery updated", "id": b.id, "recovery_status": body.recovery_status}


# ══════════════════════════════════════════════════════════════════════════════
#  REUSABLE INGEST HELPERS (used by the upload routes AND the auto-upload job)
# ══════════════════════════════════════════════════════════════════════════════
def ingest_internal_bytes(raw: bytes, filename: str, db: Session, username=None) -> dict:
    """Parse + upsert the SMB_BANK_LOADIN internal dump (shared _upsert_wallet_loads).
    SHA-256-guards the file first so a folder re-scan can't re-append eko-less rows — a
    duplicate raises 409, which evalue_auto_pickup catches and records as skipped."""
    loads = EV.parse_internal_dump(raw, filename)
    if not loads:
        return {"rows": 0, "replaced": 0, "accounts": 0}
    # Guard AFTER the zero-row check so a file that parses to nothing records NO hash and
    # stays retryable (a pending hash on a success-return would leak and later commit); a
    # real duplicate still raises 409, caught by evalue_auto_pickup.
    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "evalue", "internal", raw, filename, None)
    inserted, replaced = _upsert_wallet_loads(db, loads)
    db.add(AuditLog(id=generate_id(), username=username, action="evalue_auto_internal",
                    action_type="app", entity_type="evalue_wallet_load",
                    detail=json.dumps({"file": filename, "rows": inserted, "replaced": replaced})))
    db.commit()
    from collections import Counter
    return {"rows": inserted, "replaced": replaced, "accounts": len(Counter(l["reco_acc_no"] for l in loads))}


def ingest_bank_bytes(raw: bytes, filename: str, bank_name: str, reco_acc_no: str,
                      db: Session, username=None) -> dict:
    """Parse + per-date preserve-matched replace of a bank statement for one account
    (shared _replace_bank_txns — NOT the old full-account wipe). SHA-256-guards the file
    first so a folder re-scan skips an already-ingested file; a date-less parse now
    refuses (422) instead of wiping history. Both are caught by evalue_auto_pickup."""
    acct = db.query(EvalueAccount).filter(EvalueAccount.reco_acc_no == reco_acc_no).first()
    rows = EV.parse_bank_statement(bank_name, raw, filename)
    if not rows:
        return {"rows": 0, "credits": 0}
    # Guard AFTER the zero-row check so a file that parses to nothing records NO hash and
    # stays retryable; a real duplicate still raises 409, caught by evalue_auto_pickup.
    from core.file_hash_guard import guard_duplicate_file
    guard_duplicate_file(db, "evalue", f"bank:{reco_acc_no}", raw, filename, None)
    _replace_bank_txns(db, acct, bank_name, reco_acc_no, rows, filename)
    db.add(AuditLog(id=generate_id(), username=username, action="evalue_auto_bank",
                    action_type="app", entity_type="evalue_bank_txn",
                    detail=json.dumps({"bank": bank_name, "account": reco_acc_no,
                                       "file": filename, "rows": len(rows)})))
    db.commit()
    return {"rows": len(rows), "credits": sum(1 for r in rows if r["dr_cr"] == "CR")}


def evalue_auto_pickup(folder: str, db: Session, username="auto-upload", run: bool = True) -> dict:
    """
    Scan a folder and ingest E-Value files automatically:
      • any file whose name contains LOADIN / SMB_BANK_LOADIN  → internal dump
      • any file whose name contains a known account number     → that account's bank statement
    Then (optionally) reconcile every affected account. Used by the manual
    trigger and by the scheduled watch-folder job.
    """
    if not folder or not os.path.isdir(folder):
        raise HTTPException(400, f"Folder not found: {folder}")
    accounts = db.query(EvalueAccount).filter(EvalueAccount.is_active == True).all()
    result = {"internal": None, "banks": [], "reconciled": [], "skipped": []}
    affected = set()
    for fn in sorted(os.listdir(folder)):
        path = os.path.join(folder, fn)
        if not os.path.isfile(path):
            continue
        up = fn.upper()
        try:
            raw = open(path, "rb").read()
            if "LOADIN" in up or "SMB_BANK_LOADIN" in up:
                result["internal"] = {"file": fn, **ingest_internal_bytes(raw, fn, db, username)}
            else:
                acct = next((a for a in accounts if a.account_number and a.account_number in fn), None)
                if acct:
                    r = ingest_bank_bytes(raw, fn, acct.bank_name, acct.reco_acc_no, db, username)
                    result["banks"].append({"file": fn, "account": acct.reco_acc_no, **r})
                    affected.add(acct.reco_acc_no)
                else:
                    result["skipped"].append(fn)
        except Exception as e:
            # A guard 409 (already ingested) or a 422 (date-less parse) lands here — skip
            # the file, and roll back any pending rows/hash from the failed ingest so a
            # partial state can't ride the next file's commit.
            try:
                db.rollback()
            except Exception:
                pass
            result["skipped"].append(f"{fn} (error: {e})")
    if run:
        for reco in sorted(affected):
            try:
                result["reconciled"].append(_run_one(db, reco))
            except Exception as e:
                result["reconciled"].append({"reco_acc_no": reco, "error": str(e)})
    return result


@router.post("/auto-pickup")
def auto_pickup(folder: str = Query(...), run: bool = Query(True),
                db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Manually trigger an E-Value folder pickup (also callable on a schedule)."""
    return evalue_auto_pickup(folder, db, getattr(user, "username", "manual"), run)


@router.get("/recovery/list")
def recovery_list(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Excess credits (twice/wrong-amount) and their recovery state."""
    rows = db.query(EvalueBankTxn).filter(
        EvalueBankTxn.recon_status.in_(["twice_credit", "wrong_amount"])).all()
    out = [{"id": b.id, "reco_acc_no": b.reco_acc_no, "txn_date": b.txn_date, "amount": b.amount,
            "recon_status": b.recon_status, "utr": b.utr, "match_note": b.match_note,
            "recovery_status": b.recovery_status or "recovery_pending",
            "recovery_note": b.recovery_note} for b in rows]
    pending = sum(1 for r in out if r["recovery_status"] in ("recovery_pending", None))
    return {"items": out, "count": len(out), "pending": pending,
            "pending_amount": round(sum(r["amount"] or 0 for r in out
                                        if r["recovery_status"] in ("recovery_pending", None)), 2)}
