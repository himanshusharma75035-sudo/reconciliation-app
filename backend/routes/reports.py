import logging
import io, json
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import pandas as pd
from models.database import get_db, Transaction, ReconRun, UploadSession, ReconStatus, User
from core.auth import get_current_user, require_permission

logger = logging.getLogger("eko_recon")

router = APIRouter(prefix="/api/reports", tags=["reports"])

# ── Product → constituent partner slugs ───────────────────────────────────────
# Reports are selected PRODUCT-wise. DMT spans several banks; selecting it filters
# across all of them. Other products map 1:1 to a partner slug. Per-bank slugs
# (fino/axis/…) still work for drill-down.
PRODUCT_PARTNERS = {
    "dmt": ["fino", "airtel", "axis", "levin"],
}


def _partner_slugs(partner):
    """Expand a Reports product/partner selection to concrete partner slugs.
    'dmt' → its banks; ''/'all'/None → None (no filter); else [partner]."""
    if not partner or str(partner).lower() == "all":
        return None
    return PRODUCT_PARTNERS.get(str(partner).lower(), [partner])


def _apply_partner(q, partner):
    """Filter a Transaction query by the (possibly product-level) partner."""
    slugs = _partner_slugs(partner)
    return q.filter(Transaction.partner.in_(slugs)) if slugs else q


def _module_export_row(d: dict) -> dict:
    """Map a module-adapter open-items row (E-Value / BBPS) to the export columns —
    identical in shape and order to the core-ledger mapper below. Module products
    have no human-override workflow, so those two columns stay blank."""
    side = d.get("side")
    return {
        "Partner":          d.get("partner"),
        "Side":             side,
        "Recon Date":       d.get("recon_date"),
        "Transaction Date": d.get("transaction_date"),
        "Recon Status":     d.get("recon_status"),
        "Row Type":         d.get("row_type") or "txn",
        "Eko TID":          d.get("eko_tid")         or "",
        "Tracking Number":  d.get("tracking_number") or "",
        "UTR Number":       d.get("utr_number")      or "",
        "Amount":           d.get("amount"),
        "DR/CR":            d.get("dr_cr")           or "",
        "Txn Status":       d.get("status")          or "",
        "Match ID":         d.get("match_id")        or "",
        "SRC Code":         d.get("src_code")        or "",
        "SRC Note":         d.get("src_note")        or "",
        "Override By":      "",
        "Override Note":    "",
        "Bank Description": (d.get("bank_description") or "") if side == "bank" else "",
        "CSP Code":         (d.get("csp_code") or "") if side == "internal" else "",
        "CSP Name":         (d.get("csp_name") or "") if side == "internal" else "",
    }

@router.get("/open-items/export")
def export_open_items(
    partner:          Optional[str] = None,
    recon_date:       Optional[str] = None,
    # Accept BOTH naming conventions — frontend uses date_from/date_to,
    # older callers may use from_date/to_date. Whichever is set wins.
    date_from:        Optional[str] = None,
    date_to:          Optional[str] = None,
    from_date:        Optional[str] = None,   # alias
    to_date:          Optional[str] = None,   # alias
    side:             Optional[str] = None,
    src_code:         Optional[str] = None,
    eko_tid:          Optional[str] = None,
    tracking_number:  Optional[str] = None,
    match_id:         Optional[str] = None,
    aging:            Optional[str] = None,
    recon_status:     Optional[str] = None,
    bank_description: Optional[str] = None,   # substring search over bank narration
    csp_code:         Optional[str] = None,   # substring search over CSP code (internal rows)
    csp_name:         Optional[str] = None,   # substring search over CSP name (internal rows)
    amount_min:       Optional[float] = None,
    amount_max:       Optional[float] = None,
    bank_account:     Optional[str] = None,   # filter core bank rows by source account
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """
    Export open items / all transactions to Excel — mirrors every filter
    available on the Open Items page so the download always matches what
    the user sees on screen.
    """
    # Resolve date alias: date_from takes precedence over from_date
    _date_from = date_from or from_date
    _date_to   = date_to   or to_date

    # Module products (E-Value, BBPS) live in their OWN tables, not Transaction.
    # Mirror the Open Items list endpoint (recon.get_open_items) so the export
    # always matches what's on screen:
    #   • module partner (evalue/bbps) → that module's rows only
    #   • Partner = All (no partner)   → core ledger + every module product
    #   • core partner (fino/dmt/…)    → core ledger only
    from routes.recon import _module_rows, _all_module_rows, _MODULE_OPEN
    _mod_filters = dict(recon_date=recon_date, date_from=_date_from, date_to=_date_to,
                        side=side, recon_status=recon_status, eko_tid=eko_tid,
                        tracking_number=tracking_number, match_id=match_id,
                        amount_min=amount_min, amount_max=amount_max, aging=aging,
                        bank_description=bank_description, csp_code=csp_code, csp_name=csp_name)
    _is_module = partner in _MODULE_OPEN
    _is_all    = (not partner) or str(partner).lower() == "all"
    if _is_module:
        module_rows = _module_rows(partner, db=db, **_mod_filters)
    elif _is_all:
        module_rows = _all_module_rows(db, **_mod_filters)
    else:
        module_rows = []

    q = db.query(Transaction)

    # ── Row-type gate: skip fee/settlement rows unless user asked for "all" ──
    _skip_row_type = ("fee_matched", "fund_transfer", "interbank_matched", "all")
    if recon_status not in _skip_row_type:
        q = q.filter(Transaction.row_type == "txn")

    # ── Status filter — mirrors get_open_items in recon.py exactly ───────────
    if recon_status == "all":
        pass
    elif recon_status == "matched_all":
        q = q.filter(Transaction.recon_status.in_([ReconStatus.matched, ReconStatus.manual_matched]))
    elif recon_status in (
        "matched", "manual_matched", "unmatched", "src_assigned",
        "fee_matched", "fund_transfer", "amount_mismatch", "duplicate",
        "failed", "interbank_matched", "adhoc_settlement", "reversal_matched",
        "human_override",
    ):
        q = q.filter(Transaction.recon_status == recon_status)
    else:
        # Default: open items (unmatched + src_assigned)
        q = q.filter(Transaction.recon_status.in_([ReconStatus.unmatched, ReconStatus.src_assigned]))

    # ── Other filters ─────────────────────────────────────────────────────────
    if partner: q = _apply_partner(q, partner)
    if side:    q = q.filter(Transaction.side == side)
    if src_code: q = q.filter(Transaction.src_code == src_code)
    if match_id: q = q.filter(Transaction.match_id.ilike(f"%{match_id}%"))
    if bank_description:
        q = q.filter(Transaction.bank_description.ilike(f"%{bank_description}%"))
    if csp_code:
        q = q.filter(Transaction.csp_code.ilike(f"%{csp_code}%"))
    if csp_name:
        q = q.filter(Transaction.csp_name.ilike(f"%{csp_name}%"))
    if bank_account:
        q = q.filter(Transaction.bank_account == bank_account)
    if amount_min is not None:
        q = q.filter(Transaction.amount >= amount_min)
    if amount_max is not None:
        q = q.filter(Transaction.amount <= amount_max)

    # Date: single date takes precedence over range
    if recon_date:
        q = q.filter(Transaction.recon_date == recon_date)
    else:
        if _date_from: q = q.filter(Transaction.recon_date >= _date_from)
        if _date_to:   q = q.filter(Transaction.recon_date <= _date_to)

    # TID / tracking — supports comma-sep list (same as open-items endpoint)
    if eko_tid:
        tids = [t.strip() for t in eko_tid.strip().strip("()[]{} ").split(",") if t.strip()]
        q = q.filter(Transaction.eko_tid.in_(tids)) if len(tids) > 1 \
            else q.filter(Transaction.eko_tid.ilike(f"%{tids[0]}%")) if tids else q
    if tracking_number:
        tracks = [t.strip() for t in tracking_number.strip().strip("()[]{} ").split(",") if t.strip()]
        q = q.filter(Transaction.tracking_number.in_(tracks)) if len(tracks) > 1 \
            else q.filter(Transaction.tracking_number.ilike(f"%{tracks[0]}%")) if tracks else q

    # Aging filter
    if aging:
        import datetime as _dt
        today = _dt.date.today()
        if aging == "d1":
            target = str(today - _dt.timedelta(days=1))
            q = q.filter(Transaction.recon_date == target)
        elif aging == "d3":
            q = q.filter(Transaction.recon_date.between(
                str(today - _dt.timedelta(days=3)), str(today - _dt.timedelta(days=2))))
        elif aging == "d7":
            q = q.filter(Transaction.recon_date <= str(today - _dt.timedelta(days=4)))

    # Core rows are skipped for a module-only partner (its data isn't in Transaction).
    txns = [] if _is_module else q.order_by(
        Transaction.recon_date, Transaction.partner, Transaction.side).all()

    rows = [{
        "Partner":          t.partner,
        "Side":             t.side,
        "Recon Date":       t.recon_date,
        "Transaction Date": t.transaction_date,
        "Recon Status":     t.recon_status,
        "Row Type":         t.row_type,
        "Eko TID":          t.eko_tid         or "",
        "Tracking Number":  t.tracking_number or "",
        "UTR Number":       t.utr_number      or "",
        "Amount":           t.amount,
        "DR/CR":            t.dr_cr           or "",
        "Txn Status":       t.status          or "",
        "Match ID":         t.match_id        or "",
        "SRC Code":         t.src_code        or "",
        "SRC Note":         t.src_note        or "",
        "Override By":      getattr(t, "override_by", "") or "",
        "Override Note":    getattr(t, "override_note", "") or "",
        # Bank narration (bank side only; internal rows stay blank)
        "Bank Description": (getattr(t, "bank_description", "") or "") if t.side == "bank" else "",
        # CSP (retailer) identity (internal side only; bank rows stay blank)
        "CSP Code":         (getattr(t, "csp_code", "") or "") if t.side == "internal" else "",
        "CSP Name":         (getattr(t, "csp_name", "") or "") if t.side == "internal" else "",
    } for t in txns]
    # Append module-product rows (E-Value / BBPS) so the export mirrors the list.
    rows.extend(_module_export_row(d) for d in module_rows)

    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "Partner","Side","Recon Date","Transaction Date","Recon Status","Row Type",
        "Eko TID","Tracking Number","UTR Number","Amount","DR/CR","Txn Status",
        "Match ID","SRC Code","SRC Note","Override By","Override Note","Bank Description",
        "CSP Code","CSP Name"
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "All Transactions" if recon_status == "all" else "Open Items"
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        _autofit(ws, df)
        from openpyxl.styles import PatternFill, Font as XFont
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = XFont(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
    output.seek(0)

    date_part = recon_date or f"{_date_from or 'all'}_to_{_date_to or 'all'}"
    label = "all_transactions" if recon_status == "all" else "open_items"
    fname = f"{label}_{partner or 'all'}_{date_part}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

@router.get("/reconciliation-statement")
def recon_statement(
    partner: str,
    recon_date: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """Bank reconciliation statement — Simplibank balance → adjustments → Bank balance. Supports single date or date range."""
    def _sum(txns):
        return sum(t.amount or 0 for t in txns)

    def _build_q(side):
        q = db.query(Transaction).filter(Transaction.partner.in_(_partner_slugs(partner) or [partner]), Transaction.side == side)
        if recon_date:
            q = q.filter(Transaction.recon_date == recon_date)
        else:
            if from_date: q = q.filter(Transaction.recon_date >= from_date)
            if to_date:   q = q.filter(Transaction.recon_date <= to_date)
        return q

    bank_txns     = _build_q("bank").all()
    internal_txns = _build_q("internal").all()

    # Display label for the date dimension
    if recon_date:
        date_label = recon_date
    else:
        date_label = f"{from_date or '—'} → {to_date or '—'}"

    debited_internal_not_bank = [t for t in internal_txns
                                  if t.recon_status == ReconStatus.unmatched and (t.status or "").lower() in ["debit", "dr"]]
    credited_internal_not_bank = [t for t in internal_txns
                                   if t.recon_status == ReconStatus.unmatched and (t.status or "").lower() in ["credit", "cr"]]
    debited_bank_not_internal = [t for t in bank_txns
                                  if t.recon_status == ReconStatus.unmatched and (t.status or "").lower() in ["debit", "dr"]]
    credited_bank_not_internal = [t for t in bank_txns
                                   if t.recon_status == ReconStatus.unmatched and (t.status or "").lower() in ["credit", "cr"]]

    balance_internal = _sum(internal_txns)
    adj = (
        - _sum(debited_internal_not_bank)
        + _sum(credited_internal_not_bank)
        - _sum(debited_bank_not_internal)
        + _sum(credited_bank_not_internal)
    )
    balance_bank = _sum(bank_txns)
    difference = (balance_internal + adj) - balance_bank

    return {
        "partner": partner,
        "recon_date": date_label,
        "balance_as_per_simplibank": round(balance_internal, 2),
        "less_debited_in_smb_not_bank": round(_sum(debited_internal_not_bank), 2),
        "add_credited_in_smb_not_bank": round(_sum(credited_internal_not_bank), 2),
        "less_debited_in_bank_not_smb": round(_sum(debited_bank_not_internal), 2),
        "add_credited_in_bank_not_smb": round(_sum(credited_bank_not_internal), 2),
        "balance_should_be_as_per_bank": round(balance_internal + adj, 2),
        "balance_as_per_bank": round(balance_bank, 2),
        "difference": round(difference, 2),
        "total_bank_txns": len(bank_txns),
        "total_internal_txns": len(internal_txns),
        "matched": len([t for t in bank_txns if t.recon_status in [ReconStatus.matched, ReconStatus.manual_matched]]),
        "unmatched_bank": len([t for t in bank_txns if t.recon_status == ReconStatus.unmatched]),
        "unmatched_internal": len([t for t in internal_txns if t.recon_status == ReconStatus.unmatched])
    }

@router.get("/summary")
def report_summary(
    partner: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    q = db.query(Transaction)
    if partner: q = _apply_partner(q, partner)
    if from_date: q = q.filter(Transaction.recon_date >= from_date)
    if to_date:   q = q.filter(Transaction.recon_date <= to_date)

    txns = q.all()

    total = len(txns)
    matched = len([t for t in txns if t.recon_status in [ReconStatus.matched, ReconStatus.manual_matched]])
    unmatched = len([t for t in txns if t.recon_status == ReconStatus.unmatched])
    src_assigned = len([t for t in txns if t.recon_status == ReconStatus.src_assigned])

    src_breakdown = {}
    for t in txns:
        if t.src_code:
            src_breakdown[t.src_code] = src_breakdown.get(t.src_code, 0) + 1

    return {
        "total": total, "matched": matched, "unmatched": unmatched,
        "src_assigned": src_assigned,
        "match_rate": round(matched / total * 100, 1) if total > 0 else 0,
        "src_breakdown": src_breakdown
    }

@router.get("/ageing")
def ageing_report(
    partner: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """Ageing of unmatched items grouped by date (aggregated in SQL; None-safe)."""
    from sqlalchemy import func
    q = db.query(Transaction.partner, Transaction.recon_date, Transaction.side,
                 func.count(Transaction.id)) \
          .filter(Transaction.recon_status == ReconStatus.unmatched,
                  Transaction.side.in_(["bank", "internal"]))
    if partner:    q = _apply_partner(q, partner)
    if from_date:  q = q.filter(Transaction.recon_date >= from_date)
    if to_date:    q = q.filter(Transaction.recon_date <= to_date)
    rows = q.group_by(Transaction.partner, Transaction.recon_date, Transaction.side).all()

    ageing: dict = {}
    for (p, d, side, cnt) in rows:
        d = d or ""                       # never let None reach the sorter
        key = f"{p}|{d}"
        if key not in ageing:
            ageing[key] = {"partner": p, "date": d, "bank": 0, "internal": 0}
        ageing[key][side] += cnt
    return sorted(ageing.values(), key=lambda x: str(x["date"]), reverse=True)

@router.get("/pg-settlement/export")
def pg_settlement_export(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """Excel export of the date-wise PG (PayU) settlement summary: gross, net, fees."""
    from sqlalchemy import func
    q = db.query(
        Transaction.recon_date,
        func.count(Transaction.id).label("txn_count"),
        func.sum(Transaction.amount).label("gross_total"),
        func.sum(Transaction.net_amount).label("net_total"),
    ).filter(
        Transaction.partner == "pg",
        Transaction.side == "bank",
        Transaction.row_type == "txn",
    )
    if date_from: q = q.filter(Transaction.recon_date >= date_from)
    if date_to:   q = q.filter(Transaction.recon_date <= date_to)
    rows = q.group_by(Transaction.recon_date).order_by(Transaction.recon_date.desc()).all()

    data = []
    for r in rows:
        gross = round(float(r.gross_total or 0), 2)
        net   = round(float(r.net_total or 0), 2)
        data.append({"Recon Date": r.recon_date or "", "Transactions": r.txn_count,
                     "Gross Amount": gross, "Net Amount": net, "Fees": round(gross - net, 2)})
    if data:
        data.append({"Recon Date": "TOTAL",
                     "Transactions": sum(d["Transactions"] for d in data),
                     "Gross Amount": round(sum(d["Gross Amount"] for d in data), 2),
                     "Net Amount":   round(sum(d["Net Amount"] for d in data), 2),
                     "Fees":         round(sum(d["Fees"] for d in data), 2)})
    df = pd.DataFrame(data) if data else pd.DataFrame({"info": ["no PG settlement data in range"]})

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="PG Settlement")
        ws = writer.sheets["PG Settlement"]
        from openpyxl.styles import PatternFill, Font as XFont
        for col in ws.columns:
            m = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(m + 2, 40)
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor="1E3A5F"); c.font = XFont(bold=True, color="FFFFFF")
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="pg_settlement_{date_from or "all"}_to_{date_to or "all"}.xlsx"'},
    )


@router.get("/export-matched")
def export_matched_pairs(
    partner: str,
    recon_date: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """
    Export matched pairs as Excel — each row = one pair:
    Bank columns on left | Match ID in middle | Internal columns on right.
    Supports single recon_date or from_date/to_date range.
    """
    import datetime

    q = db.query(Transaction).filter(
        Transaction.partner.in_(_partner_slugs(partner) or [partner]),
        Transaction.side == "bank",
        Transaction.recon_status.in_([ReconStatus.matched, ReconStatus.manual_matched]),
        Transaction.matched_with_id.isnot(None)
    )
    if recon_date:
        q = q.filter(Transaction.recon_date == recon_date)
    else:
        if from_date: q = q.filter(Transaction.recon_date >= from_date)
    if to_date:   q = q.filter(Transaction.recon_date <= to_date)

    bank_txns = q.all()

    # Build id → transaction map (bank + internal)
    all_ids = [t.id for t in bank_txns] + [t.matched_with_id for t in bank_txns]
    all_txns = {t.id: t for t in db.query(Transaction).filter(Transaction.id.in_(all_ids)).all()}

    def _t(txn):
        return {
            "Eko TID":          txn.eko_tid or "",
            "Tracking No":      txn.tracking_number or "",
            "UTR No":           txn.utr_number or "",
            "Amount":           txn.amount,
            "DR/CR":            txn.dr_cr or "",
            "Txn Status":       txn.status or "",
            "Txn Date":         txn.transaction_date or "",
            "Recon Status":     txn.recon_status,
        }

    rows = []
    for b in bank_txns:
        internal = all_txns.get(b.matched_with_id)
        brow = _t(b)
        irow = _t(internal) if internal else {k: "" for k in _t(b)}
        row = {}
        for k, v in brow.items():
            row[f"Bank – {k}"] = v
        row["Match ID"]   = b.match_id or ""
        row["Match Type"] = b.recon_status
        for k, v in irow.items():
            row[f"Internal – {k}"] = v
        # Bank narration (bank side only) — appended last so no existing column moves
        row["Bank – Description"] = getattr(b, "bank_description", "") or ""
        # CSP (retailer) identity of the internal side — appended last (additive)
        row["Internal – CSP Code"] = (getattr(internal, "csp_code", "") or "") if internal else ""
        row["Internal – CSP Name"] = (getattr(internal, "csp_name", "") or "") if internal else ""
        rows.append(row)

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Matched Pairs")
        ws = writer.sheets["Matched Pairs"]
        _autofit(ws, df)
        # Color-code header: bank = light blue, match = yellow, internal = light green
        from openpyxl.styles import PatternFill, Font
        blue   = PatternFill("solid", fgColor="DBEAFE")
        yellow = PatternFill("solid", fgColor="FEF9C3")
        green  = PatternFill("solid", fgColor="DCFCE7")
        for cell in ws[1]:
            if cell.value and cell.value.startswith("Bank"):
                cell.fill = blue
            elif cell.value in ("Match ID", "Match Type"):
                cell.fill = yellow
            elif cell.value and cell.value.startswith("Internal"):
                cell.fill = green
            cell.font = Font(bold=True)
    output.seek(0)

    date_part = recon_date or f"{from_date or 'start'}_to_{to_date or 'end'}"
    fname = f"matched_pairs_{partner}_{date_part}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@router.get("/eod-summary/export")
def export_eod_summary(
    partner: str,
    recon_date: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """
    EOD reconciliation summary — one Excel file with 3 sheets:
    1. Summary   — headline numbers (match rate, open items, aging)
    2. Open Items — all unmatched + amount_mismatch rows
    3. SRC Breakdown — count per SRC code
    """
    import datetime, openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    today = datetime.date.today()
    try:
        row_date  = datetime.date.fromisoformat(recon_date)
        days_open = (today - row_date).days
    except Exception:
        days_open = None

    all_txns = db.query(Transaction).filter(
        Transaction.partner.in_(_partner_slugs(partner) or [partner]),
        Transaction.recon_date == recon_date,
        Transaction.row_type == "txn"
    ).all()

    matched        = [t for t in all_txns if t.recon_status in ("matched", "manual_matched")]
    unmatched      = [t for t in all_txns if t.recon_status == "unmatched"]
    src_assigned   = [t for t in all_txns if t.recon_status == "src_assigned"]
    amt_mismatch   = [t for t in all_txns if t.recon_status == "amount_mismatch"]
    total          = len(all_txns)
    match_rate     = round(len(matched) / total * 100, 1) if total > 0 else 0

    # SRC breakdown
    src_counts = {}
    for t in src_assigned:
        src_counts[t.src_code or "UNCLASSIFIED"] = src_counts.get(t.src_code or "UNCLASSIFIED", 0) + 1

    output = io.BytesIO()
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 20

    summary_rows = [
        ("Partner",                  partner.upper()),
        ("Recon Date",               recon_date),
        ("Days Open",                days_open if days_open is not None else "—"),
        ("Generated At",             str(today)),
        ("", ""),
        ("Total Transactions",       total),
        ("Matched",                  len(matched)),
        ("Unmatched",                len(unmatched)),
        ("SRC Assigned",             len(src_assigned)),
        ("Amount Mismatch (Review)", len(amt_mismatch)),
        ("Match Rate",               f"{match_rate}%"),
    ]
    for r, (label, value) in enumerate(summary_rows, start=1):
        ws1.cell(r, 1, label)
        ws1.cell(r, 2, value)
        if label:
            ws1.cell(r, 1).font = Font(bold=True)

    # ── Sheet 2: Open Items ───────────────────────────────────────────────────
    open_items = unmatched + amt_mismatch + src_assigned
    ws2 = wb.create_sheet("Open Items")
    headers2 = ["Side", "Eko TID", "Tracking No", "UTR No", "Amount",
                "DR/CR", "Txn Date", "Recon Status", "SRC Code", "SRC Note",
                "Bank Description", "CSP Code", "CSP Name"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
    for r, t in enumerate(open_items, 2):
        ws2.cell(r, 1, t.side)
        ws2.cell(r, 2, t.eko_tid or "")
        ws2.cell(r, 3, t.tracking_number or "")
        ws2.cell(r, 4, t.utr_number or "")
        ws2.cell(r, 5, t.amount)
        ws2.cell(r, 6, t.dr_cr or "")
        ws2.cell(r, 7, t.transaction_date or "")
        ws2.cell(r, 8, t.recon_status)
        ws2.cell(r, 9, t.src_code or "")
        ws2.cell(r, 10, t.src_note or "")
        ws2.cell(r, 11, (getattr(t, "bank_description", "") or "") if t.side == "bank" else "")
        ws2.cell(r, 12, (getattr(t, "csp_code", "") or "") if t.side == "internal" else "")
        ws2.cell(r, 13, (getattr(t, "csp_name", "") or "") if t.side == "internal" else "")
    for c in range(1, 14):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    # ── Sheet 3: SRC Breakdown ────────────────────────────────────────────────
    ws3 = wb.create_sheet("SRC Breakdown")
    ws3.cell(1, 1, "SRC Code").font = Font(bold=True)
    ws3.cell(1, 2, "Count").font    = Font(bold=True)
    for r, (code, cnt) in enumerate(sorted(src_counts.items()), 2):
        ws3.cell(r, 1, code)
        ws3.cell(r, 2, cnt)
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 10

    wb.save(output)
    output.seek(0)

    fname = f"eod_summary_{partner}_{recon_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@router.get("/summary/export")
def export_summary(
    partner: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """Export summary report as Excel."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    q = db.query(Transaction)
    if partner:   q = _apply_partner(q, partner)
    if from_date: q = q.filter(Transaction.recon_date >= from_date)
    if to_date:   q = q.filter(Transaction.recon_date <= to_date)

    txns = q.all()

    total        = len(txns)
    matched      = len([t for t in txns if t.recon_status in [ReconStatus.matched, ReconStatus.manual_matched]])
    unmatched    = len([t for t in txns if t.recon_status == ReconStatus.unmatched])
    src_assigned = len([t for t in txns if t.recon_status == ReconStatus.src_assigned])
    match_rate   = round(matched / total * 100, 1) if total > 0 else 0

    src_breakdown: dict = {}
    for t in txns:
        if t.src_code:
            src_breakdown[t.src_code] = src_breakdown.get(t.src_code, 0) + 1

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")

    # Summary section
    summary_rows = [
        ("Partner",         partner.upper() if partner else "All"),
        ("From Date",       from_date or "—"),
        ("To Date",         to_date or "—"),
        ("", ""),
        ("Total",           total),
        ("Matched",         matched),
        ("Unmatched",       unmatched),
        ("SRC Assigned",    src_assigned),
        ("Match Rate",      f"{match_rate}%"),
    ]
    for r, (label, value) in enumerate(summary_rows, start=1):
        ws.cell(r, 1, label).font = Font(bold=True) if label else Font()
        ws.cell(r, 2, value)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    # SRC breakdown table below
    if src_breakdown:
        start = len(summary_rows) + 2
        ws.cell(start, 1, "SRC Code").fill = hdr_fill
        ws.cell(start, 1).font = hdr_font
        ws.cell(start, 2, "Count").fill = hdr_fill
        ws.cell(start, 2).font = hdr_font
        for i, (code, cnt) in enumerate(sorted(src_breakdown.items()), 1):
            ws.cell(start + i, 1, code)
            ws.cell(start + i, 2, cnt)

    wb.save(output)
    output.seek(0)

    date_part = f"{from_date or 'start'}_to_{to_date or 'end'}"
    fname = f"summary_{partner or 'all'}_{date_part}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@router.get("/ageing/export")
def export_ageing(
    partner: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """Export ageing analysis as Excel."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    q = db.query(Transaction).filter(Transaction.recon_status == ReconStatus.unmatched)
    if partner:   q = _apply_partner(q, partner)
    if from_date: q = q.filter(Transaction.recon_date >= from_date)
    if to_date:   q = q.filter(Transaction.recon_date <= to_date)

    txns = q.all()

    ageing: dict = {}
    for t in txns:
        key = f"{t.partner}|{t.recon_date}"
        if key not in ageing:
            ageing[key] = {"Partner": t.partner, "Date": t.recon_date, "Bank Open": 0, "Internal Open": 0}
        ageing[key]["Bank Open" if t.side == "bank" else "Internal Open"] += 1

    rows = sorted(ageing.values(), key=lambda x: x["Date"], reverse=True)
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Partner", "Date", "Bank Open", "Internal Open"])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ageing Analysis")
        ws = writer.sheets["Ageing Analysis"]
        _autofit(ws, df)
        from openpyxl.styles import PatternFill, Font as XFont
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = XFont(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
    output.seek(0)

    date_part = f"{from_date or 'start'}_to_{to_date or 'end'}"
    fname = f"ageing_{partner or 'all'}_{date_part}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


def _autofit(ws, df):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)


def _txn_rows(txns):
    """Shared serialiser for full-side and SRC reports."""
    return [{
        "Recon Date":       t.recon_date,
        "Partner":          t.partner,
        "Side":             t.side,
        "Row Type":         t.row_type,
        "Recon Status":     t.recon_status,
        "Eko TID":          t.eko_tid or "",
        "Tracking Number":  t.tracking_number or "",
        "UTR Number":       t.utr_number or "",
        "Amount":           t.amount,
        "DR/CR":            t.dr_cr or "",
        "Txn Status":       t.status or "",
        "Transaction Date": t.transaction_date or "",
        "Match ID":         t.match_id or "",
        "SRC Code":         t.src_code or "",
        "SRC Note":         t.src_note or "",
        # Bank narration (bank side only; internal rows blank)
        "Bank Description": (getattr(t, "bank_description", "") or "") if t.side == "bank" else "",
        # CSP (retailer) identity (internal side only; bank rows blank)
        "CSP Code":         (getattr(t, "csp_code", "") or "") if t.side == "internal" else "",
        "CSP Name":         (getattr(t, "csp_name", "") or "") if t.side == "internal" else "",
    } for t in txns]


def _styled_excel(rows, sheet_name, header_color="1E3A5F") -> io.BytesIO:
    """Build a styled Excel workbook from a list of dicts."""
    from openpyxl.styles import PatternFill, Font as XFont
    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        _autofit(ws, df)
        hdr_fill = PatternFill("solid", fgColor=header_color)
        hdr_font = XFont(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
    output.seek(0)
    return output


# ─── One-side Full Report ─────────────────────────────────────────────────────
@router.get("/side-report/export")
def export_side_report(
    side: str,                         # "bank" or "internal"
    partner: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """
    Complete one-side report: every transaction for bank OR internal side,
    all row types, all recon statuses, no filtering.
    """
    if side not in ("bank", "internal"):
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="side must be 'bank' or 'internal'")

    q = db.query(Transaction).filter(Transaction.side == side)
    if partner:   q = _apply_partner(q, partner)
    if from_date: q = q.filter(Transaction.recon_date >= from_date)
    if to_date:   q = q.filter(Transaction.recon_date <= to_date)

    txns = q.order_by(Transaction.recon_date, Transaction.partner).all()

    rows = _txn_rows(txns)
    color = "1E3A5F" if side == "bank" else "065F46"
    output = _styled_excel(rows, f"{side.title()} Full Report", color)

    date_part = f"{from_date or 'start'}_to_{to_date or 'end'}"
    fname = f"{side}_full_report_{partner or 'all'}_{date_part}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── SRC Code Report ─────────────────────────────────────────────────────────
@router.get("/src-report/export")
def export_src_report(
    src_code: Optional[str] = None,    # specific code, or empty = all SRC-assigned
    partner: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("reports"))
):
    """Export transactions filtered by SRC code (or all SRC-assigned rows)."""
    from models.database import ReconStatus as RS
    q = db.query(Transaction).filter(Transaction.recon_status == RS.src_assigned)
    if src_code:  q = q.filter(Transaction.src_code == src_code)
    if partner:   q = _apply_partner(q, partner)
    if from_date: q = q.filter(Transaction.recon_date >= from_date)
    if to_date:   q = q.filter(Transaction.recon_date <= to_date)

    txns = q.order_by(Transaction.src_code, Transaction.recon_date).all()

    rows = _txn_rows(txns)
    output = _styled_excel(rows, "SRC Report", "92400E")   # amber header

    date_part = f"{from_date or 'start'}_to_{to_date or 'end'}"
    label = src_code or "all_src"
    fname = f"src_report_{label}_{partner or 'all'}_{date_part}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ─── SRC Code Listing (for UI dropdown) ──────────────────────────────────────
@router.get("/src-codes-active")
def get_active_src_codes(
    partner: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Return distinct SRC codes actually used in the DB (for dropdown population)."""
    from sqlalchemy import distinct
    q = db.query(distinct(Transaction.src_code)).filter(
        Transaction.recon_status == "src_assigned",
        Transaction.src_code.isnot(None),
    )
    if partner:   q = _apply_partner(q, partner)
    if from_date: q = q.filter(Transaction.recon_date >= from_date)
    if to_date:   q = q.filter(Transaction.recon_date <= to_date)

    rows = q.all()
    return {"src_codes": sorted([r[0] for r in rows if r[0]])}



# ── Funds Position (EOD) — director/CFO view ───────────────────────────────────
# "What are the funds, product-wise, as per the data given" — latest statement per
# bank account ≤ the chosen date (T+1 aware: the statement date is always shown).
# Data comes from BankBalanceSnapshot, written at ingest by every bank upload path.
# Reporting only — reconciliation never reads this.

@router.get("/funds-position")
def funds_position(
    as_of: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    import datetime as _dt
    from core.funds import get_funds_position
    return get_funds_position(db, as_of or str(_dt.date.today()))


@router.get("/funds-position/export")
def funds_position_export(
    as_of: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    import io as _io
    import datetime as _dt
    import pandas as _pd
    from core.funds import get_funds_position
    from routes.sbi_kiosk import _sheet as _xl_sheet
    d = as_of or str(_dt.date.today())
    pos = get_funds_position(db, d)
    cols = ["product", "partner", "bank_account", "statement_date", "stale",
            "opening_balance", "total_dr", "total_cr", "closing_balance",
            "delta", "prev_date", "txn_count", "source"]
    tot_cols = ["product", "accounts", "closing_total", "stale"]
    totals = [{"product": p, **t} for p, t in sorted(pos["totals_by_product"].items())]
    out = _io.BytesIO()
    with _pd.ExcelWriter(out, engine="openpyxl") as w:
        _xl_sheet(w, "Funds Position", pos["rows"], cols)
        _xl_sheet(w, "Totals by Product", totals, tot_cols)
    out.seek(0)
    from fastapi.responses import StreamingResponse as _SR
    return _SR(out,
               media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
               headers={"Content-Disposition": f'attachment; filename="funds_position_{d}.xlsx"',
                        "X-Recon-Date": d, "Access-Control-Expose-Headers": "X-Recon-Date"})
